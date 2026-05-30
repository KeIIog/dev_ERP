from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Header
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from datetime import datetime
from urllib.parse import urlparse, parse_qs, unquote, quote
from typing import List, Optional, Dict, Any
from jose import jwt
import os, sys, re, mimetypes
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from server.database import get_db, ReceiptItem, PurchaseOrder, PurchaseRequest, User
from server.inspection_report_generator import generate_inspection_report, generate_inspection_report_template_xlsx, generate_inspection_report_from_template_pdf, open_folder_on_server
from server.vendor_import import load_vendor_list, _normalize_vendor_name
from server.routers.users import _get_current_user, SECRET_KEY, ALGORITHM

router = APIRouter()
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PHOTO_DIR = os.path.join(BASE_DIR, 'receipt_photos')
REPORT_DIR = os.path.join(BASE_DIR, 'generated', 'inspection_reports')
LIST_DIR = os.path.join(BASE_DIR, 'generated', 'receipt_lists')
QUALITY_DIR = os.path.join(BASE_DIR, 'generated', 'quality_control')
os.makedirs(PHOTO_DIR, exist_ok=True); os.makedirs(REPORT_DIR, exist_ok=True); os.makedirs(LIST_DIR, exist_ok=True); os.makedirs(QUALITY_DIR, exist_ok=True)


def _join_people_names(values):
    out = []
    for v in values or []:
        n = str(v or '').strip()
        if n and n not in out:
            out.append(n)
    if not out:
        return ''
    return out[0] if len(out) == 1 else f"{out[0]} 외 {len(out)-1}명"


def _short_notify_text(value: str, limit: int = 90) -> str:
    text = ' '.join(str(value or '').split())
    return text if len(text) <= limit else text[:limit - 1] + '…'


def _append_client_notification(kind: str, title: str, message: str, entity_key: str = '') -> None:
    try:
        from server.notification_events import append_notification_event
        append_notification_event(kind, title, message, entity_key=entity_key)
    except Exception:
        pass


def _receipt_action_label(target_stage: str) -> str:
    if target_stage == '구매팀입고':
        return '입고처리'
    if target_stage == '품질검수':
        return '출고처리'
    if target_stage == '생산팀입고':
        return '생산팀 입고처리'
    return str(target_stage or '처리')


def _receipt_item_notify_message(item: ReceiptItem, action_label: str, actor: str, suffix: str = '') -> str:
    order = getattr(item, 'order', None)
    pr = getattr(order, 'purchase_request', None) if order else None
    req_no = (getattr(pr, 'request_no', '') or getattr(order, 'order_no', '') or '').strip() if order else ''
    vendor = str(getattr(order, 'vendor_name', '') or '').strip() if order else ''
    title_text = _short_notify_text((getattr(pr, 'title_full', '') or getattr(pr, 'project_name', '') or '') if pr else '')
    qty = str(getattr(item, 'quantity', '') or '').strip()
    unit = str(getattr(item, 'unit', '') or '').strip()
    qty_text = f" ({qty}{unit})" if qty or unit else ''
    head = ' / '.join([x for x in [req_no, vendor] if x]) or '입고현황'
    lines = [head, f"{action_label}: {getattr(item, 'item_name', '') or '품목'}{qty_text}"]
    if actor:
        lines.append(f"처리자: {actor}")
    if suffix:
        lines.append(suffix)
    if title_text:
        lines.append(title_text)
    return '\n'.join(lines)


def _receipt_bulk_notify_message(items: list[ReceiptItem], action_label: str, actor: str) -> str:
    reqs, vendors, names = [], [], []
    for it in items or []:
        order = getattr(it, 'order', None)
        pr = getattr(order, 'purchase_request', None) if order else None
        req_no = (getattr(pr, 'request_no', '') or getattr(order, 'order_no', '') or '').strip() if order else ''
        vendor = str(getattr(order, 'vendor_name', '') or '').strip() if order else ''
        name = str(getattr(it, 'item_name', '') or '').strip()
        if req_no and req_no not in reqs:
            reqs.append(req_no)
        if vendor and vendor not in vendors:
            vendors.append(vendor)
        if name and name not in names:
            names.append(name)
    lines = [f"{action_label} 완료: {len(items or [])}개 품목"]
    if reqs:
        more = f" 외 {len(reqs)-3}건" if len(reqs) > 3 else ''
        lines.append('문서: ' + ', '.join(reqs[:3]) + more)
    if vendors:
        more = f" 외 {len(vendors)-3}개" if len(vendors) > 3 else ''
        lines.append('업체: ' + ', '.join(vendors[:3]) + more)
    if names:
        lines.append('대표품목: ' + _short_notify_text(names[0], 70))
    if actor:
        lines.append(f"처리자: {actor}")
    return '\n'.join(lines)

class ScanRequest(BaseModel):
    qr_code: str
    role: str = ''
    scanned_by: str

class InspectionReportRequest(BaseModel):
    order_no: str
    report_data: Optional[Dict[str, Any]] = None

class TransferItemsRequest(BaseModel):
    item_ids: List[int]
    target_stage: str
    user_name: Optional[str] = ''

class DeleteReceiptItemsRequest(BaseModel):
    item_ids: List[int]

class QualityControlExportRequest(BaseModel):
    item_ids: List[int]

class InspectionPhotoDeleteRequest(BaseModel):
    item_id: int
    filename: str
    category: Optional[str] = ''

class ReceiptItemUpdate(BaseModel):
    material_code: Optional[str] = None
    item_name: Optional[str] = None
    maker: Optional[str] = None
    spec: Optional[str] = None
    item_group: Optional[str] = None
    quantity: Optional[str] = None
    unit: Optional[str] = None
    axis_type: Optional[str] = None
    reason: Optional[str] = None
    order_round: Optional[str] = None
    delivery_date: Optional[str] = None
    purchase_recv_at: Optional[str] = None
    quality_recv_at: Optional[str] = None
    manufacture_recv_at: Optional[str] = None
    requested_by: Optional[str] = None
    order_completed_by: Optional[str] = None
    purchase_recv_by: Optional[str] = None
    quality_recv_by: Optional[str] = None
    note: Optional[str] = None


def _split_photo_paths(value: str) -> list[str]:
    if not value: return []
    return [x for x in str(value).split(';') if x]

def _append_photo_path(old_value: str, new_path: str) -> str:
    paths = _split_photo_paths(old_value)
    if new_path: paths.append(new_path)
    return ';'.join(paths)

def _photo_exists_count(*values: str) -> int:
    return sum(1 for v in values for fp in _split_photo_paths(v) if fp and os.path.exists(fp))

def _fmt_dt_value(value) -> str:
    if not value:
        return ""
    try:
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d %H:%M")
        s = str(value)
        return s[:16] if len(s) >= 16 else s
    except Exception:
        return str(value or "")

def _photo_latest_at(*values: str) -> str:
    latest = None
    for v in values:
        for fp in _split_photo_paths(v):
            if fp and os.path.exists(fp):
                try:
                    ts = os.path.getmtime(fp)
                    if latest is None or ts > latest:
                        latest = ts
                except Exception:
                    pass
    if latest is None:
        return ""
    return datetime.fromtimestamp(latest).strftime("%Y-%m-%d %H:%M")

def _display_or_x(value) -> str:
    return value if value else "X"


def _parse_optional_datetime(value):
    txt = str(value or '').strip()
    if not txt or txt.upper() == 'X':
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(txt, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(txt.replace('T', ' ')[:19])
    except Exception:
        return None


def _optional_current_user_for_inventory(
    authorization: Optional[str] = Header(None),
    user_name: str = Query('', alias='user_name'),
    db: Session = Depends(get_db),
):
    """검수조사서 미리보기/작성용 느슨한 사용자 확인.

    기존에는 토큰이 만료되면 입고현황은 보이는데 검수조사서 미리보기만
    '로그인이 필요합니다'로 막혔다. 토큰이 유효하면 토큰 사용, 만료/누락 시
    화면에서 넘긴 user_name으로 직원 정보를 보정한다.
    """
    token = ''
    if authorization and authorization.lower().startswith('bearer '):
        token = authorization.split(' ', 1)[1].strip()
    if token:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            username = payload.get('sub') or ''
            user = db.query(User).filter(User.username == username).first()
            if user:
                return user
        except Exception:
            pass
    uname = str(user_name or '').strip()
    if uname:
        user = db.query(User).filter((User.username == uname) | (User.name == uname)).first()
        if user:
            return user
    user = db.query(User).filter(User.username == 'admin').first() or db.query(User).first()
    if user:
        return user
    return User(username='system', name='시스템', role='admin', department='관리', position='')


def _extract_qr_code(raw: str) -> str:
    """QR 스캔값 정규화: QR 표시문자, URL, query, path 모두 지원."""
    val = unquote(str(raw or '').strip())
    if not val:
        return ''

    m = re.search(r'(?i)\bQR\s*[:=]\s*([A-F0-9]{8,32})\b', val)
    if m:
        return m.group(1).upper()

    try:
        parsed = urlparse(val)
        if parsed.scheme and parsed.netloc:
            qs = parse_qs(parsed.query)
            for key in ('qr', 'qr_code', 'code'):
                if qs.get(key):
                    val = qs[key][0]
                    break
            else:
                if parsed.path:
                    val = parsed.path.rstrip('/').split('/')[-1]
        elif '?' in val:
            qs = parse_qs(val.split('?', 1)[1])
            for key in ('qr', 'qr_code', 'code'):
                if qs.get(key):
                    val = qs[key][0]
                    break
        elif val.startswith('/m/') or val.startswith('/item/') or val.startswith('/scan/'):
            val = val.rstrip('/').split('/')[-1]
    except Exception:
        pass

    if '&' in val:
        val = val.split('&', 1)[0]
    if '?' in val:
        val = val.split('?', 1)[0]

    cleaned = re.sub(r'[^A-Za-z0-9_-]', '', val.strip())
    m = re.search(r'(?i)([A-F0-9]{12})', cleaned)
    if m:
        return m.group(1).upper()
    return cleaned.upper()

def _find_item_by_qr(db: Session, qr_raw: str):
    """QR값으로 품목 조회. URL/표시문자/경로/파일명까지 최대한 복구해서 찾는다."""
    qr = _extract_qr_code(qr_raw)
    if not qr:
        return None

    try:
        item = db.query(ReceiptItem).filter(func.upper(func.trim(ReceiptItem.qr_code)) == qr.upper()).first()
        if item:
            return item
    except Exception:
        db.rollback()

    try:
        candidates = list(dict.fromkeys([qr, qr.upper(), qr.lower(), str(qr_raw or '').strip()]))
        item = db.query(ReceiptItem).filter(ReceiptItem.qr_code.in_(candidates)).first()
        if item:
            return item
    except Exception:
        db.rollback()

    try:
        item = db.query(ReceiptItem).filter(ReceiptItem.qr_code_path.like(f'%{qr}%')).first()
        if item:
            item.qr_code = qr
            db.add(item); db.commit(); db.refresh(item)
            return item
    except Exception:
        db.rollback()

    try:
        import glob
        search_dirs = [os.path.join(BASE_DIR, 'qr_codes'), os.path.join(os.getcwd(), 'qr_codes')]
        for qdir in search_dirs:
            if not os.path.isdir(qdir):
                continue
            for fp in glob.glob(os.path.join(qdir, f'*{qr}*.png')):
                name = os.path.basename(fp)
                m = re.match(r'QR_(.+)_' + re.escape(qr) + r'\.png$', name, re.I)
                if not m:
                    continue
                order_no = m.group(1)
                order = db.query(PurchaseOrder).filter(PurchaseOrder.order_no == order_no).first()
                if not order:
                    continue
                for it in order.items:
                    if it.qr_code_path and os.path.basename(str(it.qr_code_path)) == name:
                        it.qr_code = qr
                        it.qr_code_path = fp
                        db.add(it); db.commit(); db.refresh(it)
                        return it
                if len(order.items) == 1:
                    it = order.items[0]
                    it.qr_code = qr
                    it.qr_code_path = fp
                    db.add(it); db.commit(); db.refresh(it)
                    return it
    except Exception:
        db.rollback()

    return None

def _photo_records_from_value(value: str, category: str) -> list[dict]:
    records = []
    for fp in _split_photo_paths(value):
        if not fp:
            continue
        exists = os.path.exists(fp)
        records.append({
            'category': category,
            'path': fp,
            'filename': os.path.basename(fp),
            'exists': exists,
            'url': f'/inventory/photo_file/{os.path.basename(fp)}' if exists else '',
        })
    return records


def _inspection_photo_records(item: ReceiptItem) -> list[dict]:
    # 검수조사서용 사진: 품질검수 사진 우선, 없으면 구매/생산 단계 사진도 검수자료로 사용
    records = []
    records.extend(_photo_records_from_value(item.quality_photo, '품질검수'))
    records.extend(_photo_records_from_value(item.purchase_photo, '구매팀입고'))
    records.extend(_photo_records_from_value(item.manufacture_photo, '생산팀입고'))
    seen, out = set(), []
    for r in records:
        key = r.get('path') or r.get('filename')
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def _inspection_photo_paths(item: ReceiptItem) -> list[str]:
    return [r['path'] for r in _inspection_photo_records(item) if r.get('exists') and r.get('path')]


def _photo_field_for_category(category: str) -> str | None:
    cat = str(category or '').strip()
    if cat == '구매팀입고':
        return 'purchase_photo'
    if cat == '품질검수':
        return 'quality_photo'
    if cat == '생산팀입고':
        return 'manufacture_photo'
    return None


def _remove_photo_path_value(value: str, target_path: str) -> tuple[str, bool]:
    target_base = os.path.basename(str(target_path or ''))
    kept = []
    removed = False
    for fp in _split_photo_paths(value):
        if fp == target_path or os.path.basename(str(fp or '')) == target_base:
            removed = True
            continue
        kept.append(fp)
    return ';'.join(kept), removed


def _is_photo_still_referenced(db: Session, filename: str) -> bool:
    safe = os.path.basename(str(filename or ''))
    if not safe:
        return True
    like = f'%{safe}%'
    return db.query(ReceiptItem).filter(
        (ReceiptItem.purchase_photo.like(like)) |
        (ReceiptItem.quality_photo.like(like)) |
        (ReceiptItem.manufacture_photo.like(like))
    ).first() is not None


def _inspection_photo_missing_items(order: PurchaseOrder) -> list[str]:
    missing = []
    for idx, item in enumerate(order.items or [], 1):
        if not _inspection_photo_paths(item):
            missing.append(f"{idx}. {item.item_name or '(품명 없음)'} / {item.spec or ''}".strip())
    return missing

def _lookup_vendor_info(vendor_name: str) -> dict:
    # LOOKUP_VENDOR_MERGE_BIZNO_FIX_20260509
    name = str(vendor_name or '').strip()
    if not name:
        return {}
    try:
        def _n(s: str) -> str:
            return _normalize_vendor_name(str(s or '').replace('레이져', '레이저'))
        target = _n(name)
        if not target:
            return {}
        matches = []
        partial = []
        for v in load_vendor_list():
            vn = str(v.get('vendor_name') or v.get('name') or '')
            cand = _n(vn)
            if not cand:
                continue
            if cand == target:
                matches.append(v)
            elif target in cand or cand in target:
                partial.append(v)
        chosen = matches or partial
        if not chosen:
            return {}
        merged = {}
        for v in chosen:
            for k, val in v.items():
                if val is not None and str(val).strip():
                    if not merged.get(k):
                        merged[k] = val
            biz = _vendor_biz_no_value(v)
            if biz and not merged.get('biz_no'):
                merged['biz_no'] = biz
                merged['business_no'] = biz
                merged['registration_no'] = biz
        return merged
    except Exception:
        return {}

def _inspection_file_url(path: str) -> str:
    return f"/inventory/inspection_report_file/{os.path.basename(path)}" if path else ''

def _normalize_stage(stage: str) -> str:
    return '생산팀입고' if stage == '제조입고' else (stage or '미입고')

def _next_stage(current: str) -> str:
    current = _normalize_stage(current)
    return {'미입고':'구매팀입고','구매팀입고':'품질검수','품질검수':'생산팀입고','생산팀입고':'완료'}.get(current,'완료')

def _group_stage_from_user(db: Session, scanned_by: str):
    name = str(scanned_by or '').strip()
    if not name: return None, None
    user = db.query(User).filter((User.username == name) | (User.name == name)).first()
    if not user: return None, None
    dept = str(user.department or '').strip()
    if '구매그룹' in dept or '개발그룹' in dept: return 'purchase', '구매팀입고'
    if '품질그룹' in dept: return 'quality', '품질검수'
    if '제조그룹' in dept or '생산그룹' in dept: return 'manufacture', '생산팀입고'
    return None, None

def _stage_rank(stage: str) -> int:
    return {'미입고':0,'구매팀입고':1,'품질검수':2,'생산팀입고':3,'완료':3}.get(_normalize_stage(stage),0)

def _item_is_purchase_received(item: ReceiptItem) -> bool:
    """구매팀 입고 이상이면 대시보드상 입고 처리된 품목으로 판단한다."""
    if not item:
        return False
    if item.purchase_recv_at or item.purchase_recv_by:
        return True
    if item.quality_recv_at or item.quality_recv_by:
        return True
    if item.manufacture_recv_at or item.manufacture_recv_by:
        return True
    return _stage_rank(item.stage or '') >= 1

def _sync_order_receipt_status(order: PurchaseOrder | None):
    """발주 건의 모든 품목이 구매팀 입고 이상이면 입고완료로 표시한다.
    일부라도 미입고면 기존 완료 표시는 해제하여 대시보드에서 다시 추적되게 한다.
    """
    if not order:
        return
    items = list(order.items or [])
    if not items:
        return
    all_received = all(_item_is_purchase_received(it) for it in items)
    if all_received:
        order.status = '입고완료'
    elif str(order.status or '').strip() in {'입고완료', '완료'}:
        order.status = '발주서전송완료' if order.email_sent else '발주'

def _sync_request_receipt_status(db: Session, request_id):
    if not request_id:
        return
    orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == request_id).all()
    for o in orders:
        _sync_order_receipt_status(o)

def _recalc_stage(item: ReceiptItem) -> str:
    done=[]
    if item.purchase_recv_at or item.purchase_recv_by: done.append('구매팀입고')
    if item.quality_recv_at or item.quality_recv_by: done.append('품질검수')
    if item.manufacture_recv_at or item.manufacture_recv_by: done.append('생산팀입고')
    return max(done, key=_stage_rank) if done else '미입고'

def _mark_group_receipt(item: ReceiptItem, stage: str, scanned_by: str, now: datetime, photo_path: str | None = None):
    if stage == '구매팀입고':
        item.purchase_recv_at = item.purchase_recv_at or now
        item.purchase_recv_by = item.purchase_recv_by or scanned_by
        if photo_path: item.purchase_photo = _append_photo_path(item.purchase_photo, photo_path)
        label = '구매팀 입고 완료'
    elif stage == '품질검수':
        item.quality_recv_at = item.quality_recv_at or now
        item.quality_recv_by = item.quality_recv_by or scanned_by
        if photo_path: item.quality_photo = _append_photo_path(item.quality_photo, photo_path)
        label = '출고 완료'
    elif stage == '생산팀입고':
        item.manufacture_recv_at = item.manufacture_recv_at or now
        item.manufacture_recv_by = item.manufacture_recv_by or scanned_by
        if photo_path: item.manufacture_photo = _append_photo_path(item.manufacture_photo, photo_path)
        label = '생산팀 입고 완료'
    else:
        raise HTTPException(status_code=400, detail='처리 가능한 그룹이 아닙니다')
    item.stage = _recalc_stage(item)
    return label


@router.get('/scan')
def scan_item_query(qr: str = Query('', alias='qr'), qr_code: str = Query('', alias='qr_code'), code: str = Query('', alias='code'), db: Session = Depends(get_db)):
    """모바일 프록시가 QR을 query string으로 넘길 때 사용하는 안전 조회 API."""
    return scan_item(qr or qr_code or code, db)

@router.get('/scan/{qr_code:path}')
def scan_item(qr_code: str, db: Session = Depends(get_db)):
    item = _find_item_by_qr(db, qr_code)
    if not item:
        raise HTTPException(status_code=404, detail=_qr_not_found_detail(db, qr_code if 'qr_code' in locals() else getattr(data, 'qr_code', '')))
    order = item.order
    stage = _recalc_stage(item)
    if stage != item.stage:
        item.stage = stage
        db.commit()
    return {
        'id': item.id,
        'qr_code': item.qr_code,
        'item_name': item.item_name,
        'spec': item.spec,
        'quantity': item.quantity,
        'unit': item.unit,
        'stage': stage,
        'order_no': order.order_no if order else '',
        'vendor': order.vendor_name if order else '',
        'delivery_date': order.delivery_date.strftime('%Y-%m-%d') if order and order.delivery_date else '',
        'purchase_photo': item.purchase_photo,
        'quality_photo': item.quality_photo,
        'manufacture_photo': item.manufacture_photo,
        'inspection_photo_count': len(_inspection_photo_paths(item)),
        'photo_count': _photo_exists_count(item.purchase_photo, item.quality_photo, item.manufacture_photo),
        'next_stage': _next_stage(stage),
    }

@router.post('/scan')
def process_scan(data: ScanRequest, db: Session = Depends(get_db)):
    item = _find_item_by_qr(db, data.qr_code)
    if not item:
        raise HTTPException(status_code=404, detail=_qr_not_found_detail(db, qr_code if 'qr_code' in locals() else getattr(data, 'qr_code', '')))
    by = str(data.scanned_by or '').strip()
    role, target_stage = _group_stage_from_user(db, by)
    if not target_stage:
        return {'success': False, 'message': f'사용자({by})의 그룹을 확인할 수 없습니다. 조직도 계정을 확인하세요.'}
    now = datetime.now()
    label = _mark_group_receipt(item, target_stage, by, now)
    order = item.order
    if order:
        _sync_request_receipt_status(db, order.request_id)
        _sync_order_receipt_status(order)
    db.commit()
    action_label = _receipt_action_label(target_stage)
    _append_client_notification(
        'mobile_qr_receipt',
        'DevERP QR 입고/출고 알림',
        _receipt_item_notify_message(item, action_label, by, '모바일 QR 처리'),
        entity_key=str(getattr(item, 'id', '') or getattr(item, 'qr_code', '') or ''),
    )
    return {'success': True, 'item_name': item.item_name, 'stage': item.stage, 'role': role, 'message': f'✅ {item.item_name} {label}'}

@router.post('/scan_with_photo')
async def scan_with_photo(qr_code: str = Form(...), scanned_by: str = Form(...),
                          photos: list[UploadFile] = File(default=[]),
                          photo: UploadFile | None = File(default=None),
                          db: Session = Depends(get_db)):
    item = _find_item_by_qr(db, qr_code)
    if not item:
        raise HTTPException(status_code=404, detail=_qr_not_found_detail(db, qr_code if 'qr_code' in locals() else getattr(data, 'qr_code', '')))
    by = str(scanned_by or '').strip()
    role, target_stage = _group_stage_from_user(db, by)
    if not target_stage:
        return {'success': False, 'message': f'사용자({by})의 그룹을 확인할 수 없습니다. 조직도 계정을 확인하세요.'}

    upload_list = []
    if photos:
        upload_list.extend([p for p in photos if p])
    if photo:
        upload_list.append(photo)

    saved = []
    safe_qr = _extract_qr_code(item.qr_code or qr_code) or 'QR'
    for idx, up in enumerate(upload_list, 1):
        ext = os.path.splitext(up.filename or '.jpg')[1] or '.jpg'
        fname = f"{safe_qr}_{role}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{idx}{ext}"
        fpath = os.path.join(PHOTO_DIR, fname)
        with open(fpath, 'wb') as f:
            f.write(await up.read())
        saved.append(fpath)

    now = datetime.now()
    label = _mark_group_receipt(item, target_stage, by, now)
    for fpath in saved:
        if target_stage == '구매팀입고':
            item.purchase_photo = _append_photo_path(item.purchase_photo, fpath)
        elif target_stage == '품질검수':
            item.quality_photo = _append_photo_path(item.quality_photo, fpath)
        elif target_stage == '생산팀입고':
            item.manufacture_photo = _append_photo_path(item.manufacture_photo, fpath)

    order = item.order
    if order:
        _sync_request_receipt_status(db, order.request_id)
        _sync_order_receipt_status(order)
    db.commit()
    action_label = _receipt_action_label(target_stage)
    _append_client_notification(
        'mobile_qr_receipt_photo',
        'DevERP QR 입고/출고 알림',
        _receipt_item_notify_message(item, action_label, by, f"모바일 QR 처리 / 사진 {len(saved)}장" if saved else '모바일 QR 처리'),
        entity_key=str(getattr(item, 'id', '') or getattr(item, 'qr_code', '') or ''),
    )
    return {
        'success': True,
        'item_name': item.item_name,
        'stage': item.stage,
        'photo_count': len(saved),
        'inspection_photo_count': len(_inspection_photo_paths(item)),
        'message': f'✅ {item.item_name} {label}' + (f' (사진 {len(saved)}장 저장됨)' if saved else '')
    }

def _delete_file_if_exists(path: str) -> bool:
    """Best-effort file cleanup helper used by admin delete actions."""
    try:
        if path and os.path.exists(str(path)):
            os.remove(str(path))
            return True
    except Exception:
        pass
    return False


def _delete_receipt_item_files(item: ReceiptItem) -> None:
    """Delete files that belong to a single receipt item only."""
    if not item:
        return
    for photo_value in [item.purchase_photo, item.quality_photo, item.manufacture_photo]:
        for fp in _split_photo_paths(photo_value):
            _delete_file_if_exists(fp)
    _delete_file_if_exists(getattr(item, 'qr_code_path', '') or '')


def _delete_empty_order_if_needed(db: Session, order: PurchaseOrder) -> bool:
    """품목이 0개가 된 업체/발주 건은 실제로 삭제한다.

    삭제 품목 재생성 방지 tombstone은 구매의뢰서에 먼저 이동한다.
    그래서 빈 업체 행을 DB에 남기지 않아도 자동 보정 로직이 삭제 품목을 다시 만들지 않는다.
    """
    if not order or not getattr(order, 'id', None):
        return False
    try:
        remain = db.query(ReceiptItem).filter(ReceiptItem.order_id == order.id).count()
    except Exception:
        remain = len(getattr(order, 'items', []) or [])
    if remain > 0:
        return False
    try:
        from server.routers.purchase import _move_order_deleted_receipt_tombstones_to_request
        _move_order_deleted_receipt_tombstones_to_request(order)
    except Exception:
        pass
    try:
        db.delete(order)
        return True
    except Exception:
        return False


def _sync_request_after_item_delete(db: Session, request_id: int) -> None:
    """Keep purchase request status only when the whole request has no remaining orders."""
    if not request_id:
        return
    pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == request_id).first()
    if not pr:
        return
    remaining_orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == request_id).count()
    if remaining_orders <= 0:
        pr.status = '작성완료'
        pr.bizbox_uploaded = 0
    else:
        # 기존 발주/입고 데이터가 남아 있으면 구매의뢰서는 유지하고 상태만 재동기화한다.
        try:
            _sync_request_receipt_status(db, request_id)
        except Exception:
            pass


@router.post('/receipt_items/delete')
def delete_receipt_items(data: DeleteReceiptItemsRequest, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    if getattr(current_user, 'role', '') != 'admin':
        raise HTTPException(status_code=403, detail='관리자만 입고현황 품목을 삭제할 수 있습니다.')
    ids = []
    seen_ids = set()
    for x in (data.item_ids or []):
        try:
            item_id = int(x)
        except Exception:
            continue
        if item_id > 0 and item_id not in seen_ids:
            ids.append(item_id)
            seen_ids.add(item_id)
    if not ids:
        return {'success': False, 'message': '삭제할 입고 품목을 선택하세요.'}

    items = db.query(ReceiptItem).filter(ReceiptItem.id.in_(ids)).all()
    if not items:
        return {'success': False, 'message': '선택한 입고 품목을 찾을 수 없습니다.'}

    affected_orders: dict[int, PurchaseOrder] = {}
    affected_request_ids = set()
    deleted_items = 0
    deleted_names = []
    for item in items:
        order = item.order
        if order and getattr(order, 'id', None):
            affected_orders[order.id] = order
            if getattr(order, 'request_id', None):
                affected_request_ids.add(order.request_id)
        name = (getattr(item, 'item_name', '') or getattr(item, 'spec', '') or getattr(item, 'qr_code', '') or '').strip()
        if name:
            deleted_names.append(name)
        # 자동 보정 로직이 다음 입고현황 조회 때 삭제 품목을 다시 만들지 않도록
        # 삭제 전 발주 건에 tombstone 키를 남긴다.
        if order:
            try:
                from server.routers.purchase import _remember_deleted_receipt_item_for_order
                _remember_deleted_receipt_item_for_order(order, item)
            except Exception:
                pass
        _delete_receipt_item_files(item)
        db.delete(item)
        deleted_items += 1

    # 선택 품목 삭제를 먼저 반영한 뒤, 품목이 0개가 된 업체/발주 건은 실제 삭제한다.
    # tombstone은 구매의뢰서에 남겨 자동 보정으로 다시 살아나지 않게 한다.
    db.flush()
    deleted_empty_orders = 0
    for order in list(affected_orders.values()):
        try:
            if _delete_empty_order_if_needed(db, order):
                deleted_empty_orders += 1
        except Exception:
            pass
    db.flush()

    for rid in list(affected_request_ids):
        _sync_request_after_item_delete(db, rid)

    db.commit()
    try:
        from server.routers.purchase import _sync_material_budget_master
        _sync_material_budget_master(db)
    except Exception:
        pass

    sample = ', '.join(deleted_names[:3])
    more = f' 외 {max(0, len(deleted_names)-3)}개' if len(deleted_names) > 3 else ''
    extra = f'\n삭제 품목: {sample}{more}' if sample else ''
    empty_order_msg = ''
    return {
        'success': True,
        'message': f'선택한 입고현황 품목 {deleted_items}개만 삭제했습니다. 품목이 0개가 된 업체/발주 건 {deleted_empty_orders}건도 DB에서 삭제했습니다. 삭제 품목은 자동 보정으로 다시 생성되지 않게 기록했고, 재료비관리 사용금액을 남은 품목 기준으로 다시 계산했습니다.{extra}'
    }

@router.post('/receipt_item/{item_id}/update')
def update_receipt_item(item_id: int, data: ReceiptItemUpdate, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    """관리자용 입고현황 셀 수정."""
    if getattr(current_user, 'role', '') != 'admin':
        raise HTTPException(status_code=403, detail='관리자만 입고현황을 직접 수정할 수 있습니다.')
    item = db.query(ReceiptItem).filter(ReceiptItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail='품목을 찾을 수 없습니다.')

    payload = data.dict(exclude_unset=True)
    order = item.order
    pr = order.purchase_request if order else None

    text_fields = ['material_code', 'item_name', 'maker', 'spec', 'item_group', 'axis_type', 'order_round', 'note']
    for field in text_fields:
        if field in payload:
            setattr(item, field, str(payload.get(field) or '').strip())

    if 'quantity' in payload:
        raw = str(payload.get('quantity') or '').strip()
        # '3 EA'처럼 표시된 값을 수정한 경우 숫자와 단위를 같이 보정한다.
        m = re.match(r'\s*([-+]?\d+(?:\.\d+)?)\s*([A-Za-z가-힣]*)', raw)
        if m:
            try:
                item.quantity = int(float(m.group(1)))
            except Exception:
                pass
            if m.group(2):
                item.unit = m.group(2).strip()
    if 'unit' in payload:
        item.unit = str(payload.get('unit') or '').strip() or 'EA'

    if 'reason' in payload and pr:
        pr.reason = str(payload.get('reason') or '').strip()
    if 'delivery_date' in payload and order:
        order.delivery_date = _parse_optional_datetime(payload.get('delivery_date'))
    if 'purchase_recv_at' in payload:
        item.purchase_recv_at = _parse_optional_datetime(payload.get('purchase_recv_at'))
    if 'quality_recv_at' in payload:
        item.quality_recv_at = _parse_optional_datetime(payload.get('quality_recv_at'))
    if 'manufacture_recv_at' in payload:
        item.manufacture_recv_at = _parse_optional_datetime(payload.get('manufacture_recv_at'))
    if 'requested_by' in payload and pr:
        pr.requested_by = str(payload.get('requested_by') or '').strip()
    if 'order_completed_by' in payload and order:
        order.order_completed_by = str(payload.get('order_completed_by') or '').strip()
    if 'purchase_recv_by' in payload:
        item.purchase_recv_by = str(payload.get('purchase_recv_by') or '').strip()
    if 'quality_recv_by' in payload:
        item.quality_recv_by = str(payload.get('quality_recv_by') or '').strip()

    item.stage = _recalc_stage(item)
    if order:
        _sync_request_receipt_status(db, order.request_id)
        _sync_order_receipt_status(order)
    db.commit()
    db.refresh(item)
    return {'success': True, 'message': '수정되었습니다.', 'item_id': item.id, 'stage': item.stage}


@router.get('/receipt_list')
def get_receipt_list(db: Session = Depends(get_db)):
    # 발주관리/대시보드는 보이는데 입고현황만 비는 구 DB/중간 오류 케이스를 방지한다.
    # PurchaseOrder는 존재하지만 ReceiptItem이 누락된 경우, 구매의뢰의 items_json 기준으로 즉시 보정한다.
    try:
        from server.routers.purchase import _repair_missing_vendor_orders_for_request, _cleanup_empty_receipt_orders, _is_receipt_order, _classify_order_types_for_request
        seen_pr = set()
        for po in db.query(PurchaseOrder).all():
            pr = getattr(po, 'purchase_request', None)
            if pr and getattr(pr, 'id', None) and pr.id not in seen_pr:
                seen_pr.add(pr.id)
                _classify_order_types_for_request(db, pr)
                _repair_missing_vendor_orders_for_request(db, pr)
        # v48/v49에서 남겨진 품목 0개 업체/발주 건을 실제 삭제한다.
        # 삭제 품목 tombstone은 구매의뢰서에 보존되어 자동 보정으로 다시 살아나지 않는다.
        cleaned_empty_orders = _cleanup_empty_receipt_orders(db)
        if seen_pr or cleaned_empty_orders:
            db.commit()
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print(f"[WARN] receipt_list repair skipped: {e}")

    orders = [o for o in db.query(PurchaseOrder).order_by(PurchaseOrder.order_date.desc(), PurchaseOrder.id.desc()).all() if _is_receipt_order(o)]

    def vendor_display(names):
        out = []
        for name in names:
            n = str(name or '').strip() or '미정업체'
            if n not in out:
                out.append(n)
        if not out:
            return ''
        return out[0] if len(out) == 1 else f"{out[0]} 외 {len(out)-1}개"

    grouped = []
    seen = {}
    for order in orders:
        key = f"request:{order.request_id}" if getattr(order, 'request_id', None) else f"order:{order.id}"
        if key not in seen:
            seen[key] = []
            grouped.append(seen[key])
        seen[key].append(order)
    for group in grouped:
        group.sort(key=lambda x: x.id or 0)

    result = []
    for group in grouped:
        base = group[0]
        pr = base.purchase_request
        all_items = []
        vendor_groups = []
        any_in = False
        any_out = False
        first_in = None
        first_out = None
        latest_photo_at = ''
        inspection_photo_total = 0
        inspection_missing = []
        statuses = []
        first_receipt_url = ''

        for o in group:
            order_items = []
            for item in o.items:
                stage = _recalc_stage(item)
                if stage != item.stage:
                    item.stage = stage
                purchase_at = _fmt_dt_value(item.purchase_recv_at)
                quality_at = _fmt_dt_value(item.quality_recv_at)
                photo_at = _photo_latest_at(item.purchase_photo, item.quality_photo, item.manufacture_photo)
                inspection_count = len(_inspection_photo_paths(item))
                inspection_photo_total += inspection_count
                if inspection_count <= 0:
                    inspection_missing.append(item.item_name or f'품목ID {item.id}')
                if photo_at and (not latest_photo_at or photo_at > latest_photo_at):
                    latest_photo_at = photo_at
                if purchase_at:
                    any_in = True
                    if not first_in or item.purchase_recv_at < first_in:
                        first_in = item.purchase_recv_at
                if quality_at:
                    any_out = True
                    if not first_out or item.quality_recv_at < first_out:
                        first_out = item.quality_recv_at
                row = {
                    'id': item.id,
                    'order_id': o.id,
                    'order_no': o.order_no,
                    'vendor_name': o.vendor_name or '',
                    'project_code': pr.project_code if pr else '',
                    'actual_project': pr.project_name if pr else '',
                    'project_name': (pr.title_full or pr.project_name) if pr else '',
                    'material_code': getattr(item, 'material_code', '') or '',
                    'item_name': item.item_name,
                    'maker': getattr(item, 'maker', '') or '',
                    'spec': item.spec,
                    'item_group': getattr(item, 'item_group', '') or '',
                    'quantity': item.quantity,
                    'unit': item.unit,
                    'axis_type': getattr(item, 'axis_type', '') or '',
                    'reason': pr.reason if pr else '',
                    'order_round': getattr(item, 'order_round', '') or '',
                    'note': getattr(item, 'note', '') or '',
                    'qr_code': item.qr_code,
                    'stage': stage,
                    'purchase_recv_at': purchase_at,
                    'purchase_recv_by': item.purchase_recv_by or '',
                    'quality_recv_at': quality_at,
                    'quality_recv_by': item.quality_recv_by or '',
                    'manufacture_recv_at': _fmt_dt_value(item.manufacture_recv_at),
                    'purchase_photo': item.purchase_photo or '',
                    'quality_photo': item.quality_photo or '',
                    'manufacture_photo': item.manufacture_photo or '',
                    'photo_count': _photo_exists_count(item.purchase_photo, item.quality_photo, item.manufacture_photo),
                    'inspection_photo_count': inspection_count,
                    'inspection_photo_ready': inspection_count > 0,
                    'photo_at': photo_at,
                }
                order_items.append(row)
                all_items.append(row)
            report_path = getattr(o, 'inspection_report_pdf_path', '') or ''
            report_exists = bool(report_path and os.path.exists(report_path))
            receipt_path = _receipt_list_pdf_path(o) if '_receipt_list_pdf_path' in globals() else ''
            receipt_exists = bool(receipt_path and os.path.exists(receipt_path))
            receipt_url = f"/inventory/receipt_list_file/{os.path.basename(receipt_path)}" if receipt_exists else ''
            if receipt_url and not first_receipt_url:
                first_receipt_url = receipt_url
            # 관리자가 품목을 모두 삭제한 빈 업체 발주 건은 DB에는 tombstone 보존용으로 남기되,
            # 입고현황 화면의 업체 목록에는 표시하지 않는다.
            if not order_items:
                continue
            if o.status and o.status not in statuses:
                statuses.append(o.status)
            vendor_inbound_by = _join_people_names([it.get('purchase_recv_by') for it in order_items])
            vendor_outbound_by = _join_people_names([it.get('quality_recv_by') for it in order_items])
            vendor_inspection_photo_total = sum(int(it.get('inspection_photo_count') or 0) for it in order_items)
            vendor_inspection_missing = [it.get('item_name') or f"품목ID {it.get('id')}" for it in order_items if not it.get('inspection_photo_ready')]
            vendor_groups.append({
                'order_id': o.id,
                'order_no': o.order_no,
                'requested_by': (getattr(pr, 'requested_by', '') or pr.requester) if pr else '',
                'requester': pr.requester if pr else '',
                'order_completed_by': getattr(o, 'order_completed_by', '') or '',
                'inbound_by': vendor_inbound_by,
                'outbound_by': vendor_outbound_by,
                'vendor': o.vendor_name,
                'vendor_name': o.vendor_name,
                'delivery_date': o.delivery_date.strftime('%Y-%m-%d') if o.delivery_date else '',
                'status': o.status,
                'inspection_photo_total': vendor_inspection_photo_total,
                'inspection_photo_ready': len(vendor_inspection_missing) == 0 and len(order_items) > 0,
                'inspection_photo_missing': vendor_inspection_missing,
                'inspection_report_exists': report_exists,
                'inspection_report_url': _inspection_file_url(report_path) if report_exists else '',
                'receipt_list_exists': receipt_exists,
                'receipt_list_url': receipt_url,
                'items': order_items,
            })

        # 표시할 품목이 하나도 없는 구매의뢰/발주 묶음은 입고현황 목록에서 숨긴다.
        # 빈 PurchaseOrder 자체는 자동 보정 재생성 방지용 tombstone을 보존하기 위해 DB에만 남긴다.
        if not all_items:
            continue

        visible_order_ids = [v.get('order_id') for v in vendor_groups if v.get('order_id')]
        visible_vendor_names = [v.get('vendor_name') or v.get('vendor') or '' for v in vendor_groups]
        visible_base_order_no = vendor_groups[0].get('order_no') if vendor_groups else base.order_no

        result.append({
            'order_no': pr.request_no if pr and pr.request_no else visible_base_order_no,
            'raw_order_no': visible_base_order_no,
            'order_ids': visible_order_ids,
            'request_no': pr.request_no if pr else '',
            'actual_project': pr.project_name if pr else '',
            'project_code': pr.project_code if pr else '',
            'project_name': (pr.title_full or pr.project_name) if pr else '',
            'vendor': vendor_display(visible_vendor_names),
            'vendor_name': vendor_display(visible_vendor_names),
            'vendor_count': len(vendor_groups),
            'vendor_groups': vendor_groups,
            'delivery_date': base.delivery_date.strftime('%Y-%m-%d') if base.delivery_date else '',
            'actual_received_at': _fmt_dt_value(first_in),
            'actual_outbound_at': _fmt_dt_value(first_out),
            'photo_at': latest_photo_at,
            'inspection_photo_total': inspection_photo_total,
            'inspection_photo_ready': len(inspection_missing) == 0 and len(all_items) > 0,
            'inspection_photo_missing': inspection_missing,
            'status': ' / '.join(statuses) if len(statuses) > 1 else (statuses[0] if statuses else ''),
            'any_inbound': any_in,
            'any_outbound': any_out,
            'inspection_report_exists': any(v.get('inspection_report_exists') for v in vendor_groups),
            'inspection_report_url': next((v.get('inspection_report_url') for v in vendor_groups if v.get('inspection_report_url')), ''),
            'inspection_report_path': '',
            'receipt_list_exists': bool(first_receipt_url),
            'receipt_list_url': first_receipt_url,
            'items': all_items,
        })
    db.commit()
    return result

def _inspection_group_orders_for_order(db: Session, order: PurchaseOrder) -> list[PurchaseOrder]:
    """검수조사서 기준: 구매의뢰서(request_id) + 업체별 묶음."""
    if not order:
        return []
    target_vendor = _normalize_vendor_name(order.vendor_name or '')
    orders = []
    if getattr(order, 'request_id', None):
        candidates = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == order.request_id).order_by(PurchaseOrder.id.asc()).all()
        for o in candidates:
            if _normalize_vendor_name(o.vendor_name or '') == target_vendor:
                orders.append(o)
    if not orders:
        orders = [order]
    return orders


def _build_inspection_report_payload(order: PurchaseOrder, db: Session, current_user: User) -> dict:
    group_orders = _inspection_group_orders_for_order(db, order)
    if not group_orders:
        raise HTTPException(status_code=404, detail='검수조사서 작성 대상 발주 건을 찾을 수 없습니다.')

    first_order = group_orders[0]
    pr = first_order.purchase_request
    vendor_name = first_order.vendor_name or ''
    vendor_info = _lookup_vendor_info(vendor_name)

    photos = []
    missing_items = []
    items = []
    supply_amount = 0
    order_dates = []
    delivery_dates = []

    for o in group_orders:
        if o.order_date:
            order_dates.append(o.order_date)
        if o.delivery_date:
            delivery_dates.append(o.delivery_date)
        for i in o.items:
            item_photos = _inspection_photo_paths(i)
            if not item_photos:
                missing_items.append(f"{i.item_name or '(품명 없음)'} / {i.spec or ''}".strip())
            photos.extend(item_photos)
            qty = i.quantity or 0
            unit_price = i.unit_price or 0
            supply_amount += int(unit_price * qty)
            items.append({
                'item_name': i.item_name or '',
                'spec': i.spec or '',
                'order_qty': qty,
                'recv_qty': qty,
                'remain_qty': 0,
                'unit': i.unit or 'EA',
                'note': '',
            })

    # 중복 사진 제거
    seen = set()
    unique_photos = []
    for pth in photos:
        key = os.path.abspath(str(pth or ''))
        if key and key not in seen and os.path.exists(key):
            seen.add(key)
            unique_photos.append(key)

    dept = str(current_user.department or '')
    inspector_dept = '개발그룹' if '개발그룹' in dept else (dept.split('/')[-1] if dept else '개발그룹')
    order_date = min(order_dates).strftime('%Y-%m-%d') if order_dates else ''
    delivery_date = max(delivery_dates).strftime('%Y-%m-%d') if delivery_dates else datetime.now().strftime('%Y-%m-%d')

    display_vendor_name = vendor_info.get('vendor_name') or vendor_info.get('name') or vendor_name
    report_data = {
        'project_name': (pr.title_full or pr.project_name) if pr else '',
        'request_no': pr.request_no if pr else (first_order.order_no or ''),
        'project_code': pr.project_code if pr else '',
        'vendor_biz_no': _vendor_biz_no_value(vendor_info),
        'vendor_name': display_vendor_name,
        'vendor_ceo': vendor_info.get('ceo') or vendor_info.get('representative') or '',
        'vendor_address': vendor_info.get('address') or '',
        'vendor_contact_name': vendor_info.get('contact_name') or '',
        'vendor_phone': vendor_info.get('phone') or '',
        'vendor_email': vendor_info.get('email') or '',
        'buyer_biz_no': '124-87-31775',
        'buyer_name': '이노로보틱스 주식회사',
        'buyer_ceo': '이현철',
        'buyer_address': '경기도 화성시 정남면 정남산단로 80',
        'supply_amount': supply_amount,
        'vat_amount': int(supply_amount * 0.1),
        'order_date': order_date,
        'delivery_date': delivery_date,
        'inspection_date': datetime.now().strftime('%Y-%m-%d'),
        'inspection_place': '이노로보틱스㈜ 1층 자재보관실(전실)',
        'check_quantity': True,
        'check_spec': True,
        'check_parts': True,
        'check_condition': True,
        'inspection_result': '합격',
        'inspection_opinion': '',
        'items': items,
        'inspector_dept': inspector_dept,
        'inspector_rank': current_user.position or '',
        'inspector_name': current_user.name or current_user.username,
        'item_purpose': (pr.title_full or pr.project_name) if pr else '',
    }
    return {
        'report_data': report_data,
        'photos': unique_photos,
        'missing_items': missing_items,
        'orders': group_orders,
        'group_key': f"{report_data['request_no']}__{display_vendor_name}",
    }



def _quality_fmt_date(value) -> str:
    if not value:
        return ''
    try:
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        txt = str(value or '').strip()
        return txt[:10] if len(txt) >= 10 else txt
    except Exception:
        return str(value or '')


def _quality_progress_value(item: ReceiptItem, order: PurchaseOrder | None) -> str:
    """품질관리 LIST용 진행상태.

    요청 기준: 구매팀 입고 처리된 품목은 '입고', 발주 진행/미입고 상태는 모두 '미입고'로 표기한다.
    """
    stage = str(getattr(item, 'stage', '') or '')
    order_status = str(getattr(order, 'status', '') if order else '')
    if item.purchase_recv_at or '구매팀 입고' in stage or '입고완료' in stage or '입고완료' in order_status:
        return '입고'
    return '미입고'


def _quality_infer_item_group_text(text: str) -> str:
    t = re.sub(r'\s+', '', str(text or '')).lower()
    if not t:
        return ''
    rules = [
        ('판금품', ['판금', 'sheetmetal', 'sheet-metal', 'sheetmetal']),
        ('선반품', ['선반', 'turning', 'lathe', 'cnc선반']),
        ('밀링품', ['밀링', 'milling', 'mct']),
        ('가공품', ['가공', 'machining', 'machinepart', '가공견적']),
        ('구매품', ['구매품', '구매']),
        ('전장품', ['전장', '전장품', '전장부품', '케이블', 'cable']),
        ('프레임', ['프레임', 'frame']),
        ('석정반', ['석정반', 'granite']),
    ]
    for group, keys in rules:
        for k in keys:
            kk = re.sub(r'\s+', '', str(k or '')).lower()
            if kk and kk in t:
                return group
    return ''


def _quality_item_group_value(item: ReceiptItem, pr=None) -> str:
    if getattr(item, 'item_group', ''):
        return item.item_group
    ctx = ' '.join([
        str(getattr(item, 'item_name', '') or ''),
        str(getattr(item, 'spec', '') or ''),
        str(getattr(item, 'maker', '') or ''),
        str(getattr(item, 'note', '') or ''),
        str(getattr(pr, 'sub_category', '') or ''),
        str(getattr(pr, 'item_type', '') or ''),
        str(getattr(pr, 'title_full', '') or ''),
    ])
    return _quality_infer_item_group_text(ctx)


@router.post('/quality_control_export')
def export_quality_control_xlsx(data: QualityControlExportRequest, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    """입고현황에서 선택한 품목/구매의뢰서를 품질관리용 진행 LIST 엑셀로 내보낸다.

    첨부 예제 파일의 주요 컬럼 순서/명칭(B4:U4)을 기준으로 생성한다.
    DB 저장은 하지 않고 생성 파일만 다운로드한다.
    """
    ids = []
    seen = set()
    for raw in data.item_ids or []:
        try:
            iid = int(raw)
        except Exception:
            continue
        if iid > 0 and iid not in seen:
            ids.append(iid); seen.add(iid)
    if not ids:
        raise HTTPException(status_code=400, detail='품질관리 파일로 내보낼 품목을 선택하세요.')

    items = db.query(ReceiptItem).filter(ReceiptItem.id.in_(ids)).all()
    item_map = {i.id: i for i in items}
    ordered_items = [item_map[i] for i in ids if i in item_map]
    if not ordered_items:
        raise HTTPException(status_code=404, detail='선택한 품목을 찾을 수 없습니다.')

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Excel 생성 모듈을 불러오지 못했습니다: {e}')

    headers = [
        '문서번호', 'PJT 코드', 'PJT 명', '자재코드', '자재명', '품목군', '요청량', '축구분', '발생사유', '발주차수',
        '진행', '납기', '입고', '등급', '검사량', '완료일', '업체명', '검토', '비고_Issue 및 기타 사항', '건당'
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '품질관리'
    ws.append(headers)

    for item in ordered_items:
        order = item.order
        pr = order.purchase_request if order else None
        request_no = pr.request_no if pr and pr.request_no else (order.order_no if order else '')
        pjt_name = ''
        if pr:
            pjt_name = pr.project_name or pr.title_full or ''
        if not pjt_name and pr:
            pjt_name = pr.title_full or ''
        qty = item.quantity if item.quantity is not None else ''
        row = [
            request_no,
            (pr.project_code if pr else '') or '',
            pjt_name,
            item.spec or item.material_code or '',
            item.item_name or '',
            _quality_item_group_value(item, pr),
            qty,
            item.axis_type or '',
            getattr(item, 'reason', '') or '',
            item.order_round or '',
            _quality_progress_value(item, order),
            _quality_fmt_date(order.delivery_date if order else None),
            _quality_fmt_date(item.purchase_recv_at),
            '',
            '',
            _quality_fmt_date(item.quality_recv_at or item.manufacture_recv_at),
            (order.vendor_name if order else '') or item.maker or '',
            '',
            '',  # 비고_Issue 및 기타 사항: 품질관리 내보내기에서는 비워둔다.
            '',
        ]
        ws.append(row)

    header_fill = PatternFill('solid', fgColor='E7EAF0')
    thin = Side(style='thin', color='D9DEE8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    widths = [14, 12, 34, 22, 28, 12, 10, 10, 18, 12, 14, 12, 12, 8, 10, 12, 18, 12, 34, 10]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = True

    created = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'품질관리_진행_LIST_{created}.xlsx'
    path = os.path.join(QUALITY_DIR, filename)
    wb.save(path)
    return FileResponse(
        path,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@router.get('/item_photos/{item_id}')
def get_item_photos(item_id: int, db: Session = Depends(get_db)):
    item = db.query(ReceiptItem).filter(ReceiptItem.id == item_id).first()
    if not item:
        return {'success': False, 'message': '품목을 찾을 수 없습니다.'}
    return {
        'success': True,
        'item_id': item.id,
        'item_name': item.item_name,
        'spec': item.spec,
        'photos': _inspection_photo_records(item),
    }


@router.get('/inspection_photos/{order_no}')
def get_inspection_photos(order_no: str, db: Session = Depends(get_db)):
    """검수조사서 작성창에서 업체별 발주 건의 입고사진을 목록화한다."""
    order = db.query(PurchaseOrder).filter(PurchaseOrder.order_no == order_no).first()
    if not order:
        return {'success': False, 'message': '발주 건을 찾을 수 없습니다.'}
    group_orders = _inspection_group_orders_for_order(db, order) or [order]
    rows = []
    missing = []
    total = 0
    for o in group_orders:
        for item in o.items or []:
            photos = _inspection_photo_records(item)
            existing = [r for r in photos if r.get('exists')]
            if not existing:
                missing.append(f"{item.item_name or '(품명 없음)'} / {item.spec or ''}".strip())
            item_row = {
                'order_no': o.order_no or '',
                'vendor_name': o.vendor_name or '',
                'item_id': item.id,
                'item_name': item.item_name or '',
                'spec': item.spec or '',
                'quantity': item.quantity or 0,
                'unit': item.unit or 'EA',
                'photos': photos,
            }
            total += len(existing)
            rows.append(item_row)
    return {
        'success': True,
        'order_no': order_no,
        'request_no': order.purchase_request.request_no if order.purchase_request else order.order_no,
        'vendor_name': order.vendor_name or '',
        'photo_count': total,
        'missing_items': missing,
        'items': rows,
    }


@router.delete('/inspection_photo')
def delete_inspection_photo(data: InspectionPhotoDeleteRequest, db: Session = Depends(get_db), current_user: User = Depends(_optional_current_user_for_inventory)):
    """검수조사서용 입고사진을 품목 DB 연결에서 제거하고, 더 이상 참조되지 않으면 파일도 삭제한다."""
    item = db.query(ReceiptItem).filter(ReceiptItem.id == data.item_id).first()
    if not item:
        return {'success': False, 'message': '품목을 찾을 수 없습니다.'}
    safe_name = os.path.basename(str(data.filename or '').strip())
    if not safe_name:
        return {'success': False, 'message': '삭제할 사진 파일명이 없습니다.'}

    records = _inspection_photo_records(item)
    target = None
    for r in records:
        if os.path.basename(str(r.get('path') or r.get('filename') or '')) == safe_name:
            if data.category and str(r.get('category') or '') != str(data.category):
                continue
            target = r
            break
    if not target:
        return {'success': False, 'message': '해당 품목에 연결된 사진을 찾을 수 없습니다.'}

    fields = []
    category_field = _photo_field_for_category(target.get('category') or data.category or '')
    if category_field:
        fields.append(category_field)
    else:
        fields = ['purchase_photo', 'quality_photo', 'manufacture_photo']

    removed = False
    target_path = target.get('path') or os.path.join(PHOTO_DIR, safe_name)
    for field in fields:
        new_value, did_remove = _remove_photo_path_value(getattr(item, field, '') or '', target_path)
        if did_remove:
            setattr(item, field, new_value)
            removed = True

    if not removed:
        return {'success': False, 'message': '사진 연결 정보를 삭제하지 못했습니다.'}

    db.commit()

    deleted_file = False
    try:
        photo_path = os.path.join(PHOTO_DIR, safe_name)
        # DB 연결을 먼저 제거한 뒤 전체 DB에서 같은 파일명이 더 남아있지 않을 때만 실제 파일 삭제
        if os.path.exists(photo_path) and not _is_photo_still_referenced(db, safe_name):
            os.remove(photo_path)
            deleted_file = True
    except Exception:
        deleted_file = False

    return {
        'success': True,
        'message': '사진을 삭제했습니다.' + (' 파일도 삭제되었습니다.' if deleted_file else ' 다른 항목에서 참조 중이면 파일은 유지됩니다.'),
        'item_id': item.id,
        'filename': safe_name,
        'deleted_file': deleted_file,
        'deleted_by': current_user.name or current_user.username,
    }


@router.post('/inspection_photo')
async def upload_inspection_photo(
    item_id: int = Form(...),
    category: str = Form('품질검수'),
    replace: str = Form('0'),
    photos: list[UploadFile] = File(default=[]),
    photo: UploadFile | None = File(default=None),
    db: Session = Depends(get_db),
    user_name: str = Form(''),
    current_user: User = Depends(_optional_current_user_for_inventory),
):
    """검수조사서 작성창에서 품목별 입고사진을 추가한다.

    토큰이 만료된 브라우저에서도 현재 로그인 사용자명(user_name) 또는 관리자 fallback으로 처리한다.
    """
    item = db.query(ReceiptItem).filter(ReceiptItem.id == item_id).first()
    if not item:
        return {'success': False, 'message': '품목을 찾을 수 없습니다.'}

    upload_list = []
    if photos:
        upload_list.extend([p for p in photos if p])
    if photo:
        upload_list.append(photo)
    if not upload_list:
        return {'success': False, 'message': '업로드할 사진 파일을 선택하세요.'}

    field = _photo_field_for_category(category or '품질검수') or 'quality_photo'
    do_replace = str(replace or '').strip().lower() in ('1', 'true', 'yes', 'y', 'replace')
    old_paths = _split_photo_paths(getattr(item, field, '') or '') if do_replace else []
    if do_replace:
        setattr(item, field, '')

    saved = []
    safe_item = re.sub(r'[^A-Za-z0-9_-]+', '_', str(item.qr_code or item.id or 'ITEM')).strip('_') or str(item.id)
    for idx, up in enumerate(upload_list, 1):
        ext = os.path.splitext(up.filename or '.jpg')[1] or '.jpg'
        if not re.match(r'^\.[A-Za-z0-9]{1,8}$', ext):
            ext = '.jpg'
        fname = f"INSPECTION_{safe_item}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{idx}{ext}"
        fpath = os.path.join(PHOTO_DIR, fname)
        with open(fpath, 'wb') as f:
            f.write(await up.read())
        saved.append(fpath)
        setattr(item, field, _append_photo_path(getattr(item, field, '') or '', fpath))

    # 사진만 교체해도 검수자료로 집계되도록 품질검수 시각/처리자를 보정한다.
    now = datetime.now()
    if field == 'quality_photo' and not item.quality_recv_at:
        item.quality_recv_at = now
        item.quality_recv_by = current_user.name or current_user.username
    elif field == 'purchase_photo' and not item.purchase_recv_at:
        item.purchase_recv_at = now
        item.purchase_recv_by = current_user.name or current_user.username
    elif field == 'manufacture_photo' and not item.manufacture_recv_at:
        item.manufacture_recv_at = now
        item.manufacture_recv_by = current_user.name or current_user.username

    order = item.order
    if order:
        _sync_request_receipt_status(db, order.request_id)
        _sync_order_receipt_status(order)

    db.commit()

    deleted_old = 0
    if do_replace:
        for old in old_paths:
            try:
                old_name = os.path.basename(str(old or ''))
                old_path = os.path.join(PHOTO_DIR, old_name)
                if old_name and os.path.exists(old_path) and not _is_photo_still_referenced(db, old_name):
                    os.remove(old_path)
                    deleted_old += 1
            except Exception:
                pass

    return {
        'success': True,
        'message': f"사진 {len(saved)}장 {'교체' if do_replace else '추가'} 완료" + (f" / 기존 사진 {deleted_old}장 삭제" if deleted_old else ''),
        'item_id': item.id,
        'saved_count': len(saved),
        'replace': do_replace,
        'photos': _inspection_photo_records(item),
    }


@router.get('/photo_file/{filename}')
def get_photo_file(filename: str):
    safe = os.path.basename(filename)
    path = os.path.join(PHOTO_DIR, safe)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail='사진 파일이 없습니다.')
    media_type = mimetypes.guess_type(path)[0] or 'application/octet-stream'
    return FileResponse(path, filename=safe, media_type=media_type)

@router.get('/inspection_report_preview/{order_no}')
def inspection_report_preview(order_no: str, db: Session = Depends(get_db), current_user: User = Depends(_optional_current_user_for_inventory)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.order_no == order_no).first()
    if not order:
        return {'success': False, 'message': '발주 건을 찾을 수 없습니다.'}
    payload = _build_inspection_report_payload(order, db, current_user)
    return {
        'success': True,
        'order_no': order_no,
        'group_key': payload.get('group_key', ''),
        'report_data': payload['report_data'],
        'photo_count': len(payload['photos']),
        'missing_photo_items': payload.get('missing_items', []),
        'photo_ready': len(payload.get('photos', [])) > 0,
    }

@router.post('/inspection_report')
def create_inspection_report(data: InspectionReportRequest, db: Session = Depends(get_db), current_user: User = Depends(_optional_current_user_for_inventory)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.order_no == data.order_no).first()
    if not order:
        return {'success': False, 'message': '발주 건을 찾을 수 없습니다.'}
    payload = _build_inspection_report_payload(order, db, current_user)
    photos = payload['photos']
    missing_items = payload.get('missing_items', [])
    # 작성 조건: 선택한 견적/업체별 발주 건에 검수조사서용 사진이 1장 이상 있으면 생성 가능.
    # 일부 품목 사진이 누락되어도 누락 목록은 화면에 안내만 하고 XLSX 생성을 막지 않는다.
    if not photos:
        return {'success': False, 'message': '검수조사서용 사진이 1장 이상 있어야 XLSX를 생성할 수 있습니다.'}

    report_data = payload['report_data']
    if data.report_data:
        report_data.update({k: v for k, v in data.report_data.items() if k != 'items'})
        if isinstance(data.report_data.get('items'), list):
            report_data['items'] = data.report_data.get('items')

    created_at = datetime.now().strftime('%Y%m%d_%H%M%S')
    req_no = str(report_data.get('request_no') or order.order_no or '')
    vendor = str(report_data.get('vendor_name') or order.vendor_name or '')
    safe_name = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in f'{req_no}_{vendor}')[:120]
    xlsx_path = os.path.join(REPORT_DIR, f'{safe_name}_{created_at}_검수조사서.xlsx')

    try:
        generate_inspection_report_template_xlsx(report_data, photos, xlsx_path)
        for o in payload.get('orders') or [order]:
            o.inspection_report_pdf_path = xlsx_path
            o.inspection_report_created_at = datetime.now()
        db.commit()
    except Exception as e:
        db.rollback()
        return {'success': False, 'message': f'검수조사서 생성 중 오류가 발생했습니다: {e}'}

    return {
        'success': True,
        'order_no': order.order_no,
        'group_key': payload.get('group_key', ''),
        'file_path': xlsx_path,
        'folder_path': REPORT_DIR,
        'download_url': f'/inventory/inspection_report_file/{os.path.basename(xlsx_path)}',
        'message': '검수조사서 XLSX 생성 완료'
    }

@router.get('/inspection_report_file/{filename}')
def get_inspection_report_file(filename: str):
    path = os.path.join(REPORT_DIR, os.path.basename(filename))
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail='검수조사서 파일이 없습니다.')
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xlsx':
        media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif ext == '.pdf':
        media_type = 'application/pdf'
    else:
        media_type = mimetypes.guess_type(path)[0] or 'application/octet-stream'
    return FileResponse(path, filename=os.path.basename(path), media_type=media_type)

@router.get('/qr_image/{qr_code:path}')
def get_qr_image(qr_code: str, db: Session = Depends(get_db)):
    item = _find_item_by_qr(db, qr_code)
    if not item:
        raise HTTPException(status_code=404, detail=_qr_not_found_detail(db, qr_code))

    path = (item.qr_code_path or '').strip()
    if path and os.path.exists(path):
        try:
            from server.qr_generator import ensure_qr_image_has_item_name
            ensure_qr_image_has_item_name(path, item)
        except Exception:
            pass
        return FileResponse(path, media_type='image/png')

    # 과거 데이터처럼 DB에는 QR 값만 있고 이미지 파일이 없는 경우, 현재 ngrok URL 기준으로 즉시 재생성한다.
    try:
        from server.qr_generator import generate_qr_for_item
        order = item.order
        if not order:
            raise RuntimeError('연결된 발주 데이터가 없습니다.')
        path = generate_qr_for_item(item, order, db)
        return FileResponse(path, media_type='image/png')
    except Exception as e:
        raise HTTPException(status_code=404, detail=f'QR 이미지 파일이 없습니다: {e}')


# ── 출고 수동 처리 + 구매의뢰서별 입고 목록 PDF ──────────────────────
from reportlab.pdfgen import canvas as _canvas
from reportlab.lib.pagesizes import A4 as _A4, landscape as _landscape
from reportlab.lib.units import mm as _mm
from reportlab.pdfbase import pdfmetrics as _pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont as _TTFont

def _receipt_list_font():
    for fp in ['C:/Windows/Fonts/malgun.ttf','/usr/share/fonts/truetype/nanum/NanumGothic.ttf']:
        if os.path.exists(fp):
            try: _pdfmetrics.registerFont(_TTFont('KorReceiptList', fp)); return 'KorReceiptList'
            except Exception: pass
    return 'Helvetica'

def _safe_pdf_name(s: str) -> str:
    return ''.join('_' if ch in '<>:"/\\|?*' else ch for ch in str(s or '')).strip() or 'list'

def _receipt_list_pdf_path(order: PurchaseOrder) -> str:
    pr=order.purchase_request; req_no=(pr.request_no if pr else order.order_no) or order.order_no; title=(pr.title_full if pr else '') or (pr.project_name if pr else '') or ''
    return os.path.join(LIST_DIR, f"입고목록_{_safe_pdf_name(req_no)}_{_safe_pdf_name(title)[:60]}.pdf")

def _make_receipt_list_pdf(order: PurchaseOrder) -> str:
    path=_receipt_list_pdf_path(order); font=_receipt_list_font(); page_size=_landscape(_A4); c=_canvas.Canvas(path, pagesize=page_size)
    w,h=page_size; pr=order.purchase_request; margin=10*_mm; row_h=7*_mm
    headers=['문서번호','PJT 코드','프로젝트명','입고','출고','발주번호','자재명','규격','요청량','입고요청일자','실 입고일자','검수사진','비고']
    widths=[24,24,52,12,12,28,42,30,18,24,28,18,32]; scale=(w-margin*2)/(sum(widths)*_mm); widths=[v*_mm*scale for v in widths]
    def _short(v,n):
        s=str(v or ''); return s if len(s)<=n else s[:n-1]+'…'
    def _draw_header(first=True):
        y=h-margin
        if first:
            c.setFont(font,15); c.drawString(margin,y,'구매의뢰서별 입고 목록'); y-=9*_mm
            c.setFont(font,9); c.drawString(margin,y,f"문서번호: {(pr.request_no if pr else order.order_no) or ''}"); y-=5*_mm
            c.drawString(margin,y,f"프로젝트명: {(pr.title_full if pr else '') or (pr.project_name if pr else '') or ''}"); y-=5*_mm
            c.drawString(margin,y,f"업체: {order.vendor_name or ''}"); y-=8*_mm
        x=margin; c.setFont(font,7)
        for hd,cw in zip(headers,widths): c.rect(x,y-row_h,cw,row_h,stroke=1,fill=0); c.drawString(x+1.2*_mm,y-4.5*_mm,hd); x+=cw
        return y-row_h
    y=_draw_header(True)
    for it in order.items:
        if y < margin+row_h: c.showPage(); y=_draw_header(False)
        purchase_at=_fmt_dt_value(it.purchase_recv_at); quality_at=_fmt_dt_value(it.quality_recv_at); photo_at=_photo_latest_at(it.purchase_photo,it.quality_photo,it.manufacture_photo)
        vals=[(pr.request_no if pr else order.order_no) or '', pr.project_code if pr else '', (pr.title_full or pr.project_name) if pr else '', _display_or_x(purchase_at), _display_or_x(quality_at), order.order_no or '', it.item_name or '', it.spec or '', f"{it.quantity or 0} {it.unit or 'EA'}", order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else '', _display_or_x(purchase_at), _display_or_x(photo_at), '']
        x=margin; c.setFont(font,6.2)
        for val,cw in zip(vals,widths): c.rect(x,y-row_h,cw,row_h,stroke=1,fill=0); c.drawString(x+1.0*_mm,y-4.5*_mm,_short(val,max(4,int(cw/(2.2*_mm))))); x+=cw
        y-=row_h
    c.save(); return path

@router.post('/transfer_items')
def transfer_items(data: TransferItemsRequest, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    raw_stage = str(data.target_stage or '').strip()
    if '구매' in raw_stage or ('입고' in raw_stage and '생산' not in raw_stage and '제조' not in raw_stage):
        target = '구매팀입고'
    elif '품질' in raw_stage or '출고' in raw_stage:
        target = '품질검수'
    else:
        target = '생산팀입고'
    by=data.user_name or current_user.name or current_user.username; now=datetime.now()
    items=db.query(ReceiptItem).filter(ReceiptItem.id.in_(data.item_ids or [])).all()
    if not items: return {'success':False,'message':'선택된 품목이 없습니다.'}
    orders=set(); request_ids=set()
    for it in items:
        _mark_group_receipt(it,target,by,now)
        if it.order:
            orders.add(it.order.id)
            if it.order.request_id:
                request_ids.add(it.order.request_id)
    for rid in request_ids:
        _sync_request_receipt_status(db, rid)
    for oid in list(orders):
        order=db.query(PurchaseOrder).filter(PurchaseOrder.id==oid).first()
        _sync_order_receipt_status(order)
    db.commit(); pdfs=[]
    action_label = _receipt_action_label(target)
    notify_title = 'DevERP 출고 알림' if target == '품질검수' else ('DevERP 생산팀 입고 알림' if target == '생산팀입고' else 'DevERP 입고 알림')
    if len(items) == 1:
        _append_client_notification(
            f'receipt_{target}',
            notify_title,
            _receipt_item_notify_message(items[0], action_label, by, '웹 체크박스 처리'),
            entity_key=str(getattr(items[0], 'id', '')),
        )
    else:
        _append_client_notification(
            f'receipt_{target}_bulk',
            notify_title,
            _receipt_bulk_notify_message(items, action_label, by),
            entity_key='bulk_' + '_'.join(str(getattr(it, 'id', '')) for it in items[:10]),
        )
    # 출고/품질 처리 때만 입고목록 PDF를 만든다. 단순 입고처리는 QR/사진 없이 상태만 갱신한다.
    if target == '품질검수':
        for oid in orders:
            order=db.query(PurchaseOrder).filter(PurchaseOrder.id==oid).first()
            if order:
                try: pdfs.append(_make_receipt_list_pdf(order))
                except Exception: pass
    return {'success':True,'count':len(items),'target_stage':target,'pdf_urls':[f'/inventory/receipt_list_file/{os.path.basename(p)}' for p in pdfs]}

@router.get('/receipt_list_file/{filename}')
def get_receipt_list_file(filename: str):
    path=os.path.join(LIST_DIR, os.path.basename(filename))
    if not os.path.exists(path): raise HTTPException(status_code=404, detail='입고 목록 PDF가 없습니다.')
    return FileResponse(path, filename=os.path.basename(path), media_type='application/pdf')


# INSPECTION_STABLE_FIX_20260509
def _vendor_biz_no_value(vendor_info: dict) -> str:
    '업체관리 데이터의 여러 컬럼명에서 사업자등록번호를 최대한 찾아온다.'
    if not vendor_info:
        return ''
    keys = [
        'biz_no', 'business_no', 'registration_no',
        '사업자등록번호', '사업자 등록번호', '사업자번호', '사업자 번호',
        '사업자등록', '사업자', 'BUSINESS_NO', 'Business No', 'BusinessNo',
        'tax_no', 'tax_id', 'corp_no', 'corporate_no',
    ]
    for k in keys:
        v = vendor_info.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    for k, v in vendor_info.items():
        kk = str(k).replace(' ', '').replace('_', '').lower()
        if v is not None and str(v).strip() and (
            '사업자' in kk or '등록번호' in kk or 'biz' in kk or 'business' in kk or 'registration' in kk
        ):
            return str(v).strip()
    return ''


def _json_error(message: str, detail: str = '') -> dict:
    msg = str(message or '처리 중 오류가 발생했습니다.')
    if detail:
        msg = f"{msg}: {detail}"
    return {'success': False, 'message': msg}

