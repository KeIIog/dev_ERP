
from fastapi import APIRouter, Depends, BackgroundTasks, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field
from datetime import datetime
from typing import Optional, List
import sys, os, json, shutil, zipfile, tempfile, subprocess
from pathlib import Path
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from server.database import get_db, PurchaseRequest, PurchaseOrder, ReceiptItem, User
from server.qr_generator import generate_qr_for_item, ensure_qr_image_has_item_name
from server.purchase_order_generator import generate_purchase_order_pdf
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from server.email_sender import build_order_mail_subject, build_order_mail_body, build_order_mail_body_html, create_email_draft_file
import shared.config as cfg
from server.routers.users import _get_current_user

router = APIRouter()


def _join_people_names(values):
    out = []
    for v in values or []:
        n = str(v or '').strip()
        if n and n not in out:
            out.append(n)
    if not out:
        return ''
    return out[0] if len(out) == 1 else f"{out[0]} 외 {len(out)-1}명"


def _short_notify_text(value: str, limit: int = 90) -> str:
    text = ' '.join(str(value or '').split())
    return text if len(text) <= limit else text[:limit - 1] + '…'


def _append_client_notification(kind: str, title: str, message: str, entity_key: str = '') -> None:
    try:
        from server.notification_events import append_notification_event
        append_notification_event(kind, title, message, entity_key=entity_key)
    except Exception:
        pass



def _estimate_parser_log(message: str) -> None:
    """견적서 자동인식 상태를 콘솔과 logs/estimate_parser.log에 남긴다."""
    try:
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        line = f"[{ts}] {message}"
        print(f"[ESTIMATE_PARSER] {message}", flush=True)
        os.makedirs('logs', exist_ok=True)
        with open(os.path.join('logs', 'estimate_parser.log'), 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except Exception:
        pass


ATTACH_DIR = './uploaded_estimates'
ORDER_DIR = './generated/purchase_orders'
DRAFT_DIR = './generated/email_drafts'
PACKAGE_DIR = './generated/order_packages'
os.makedirs(ATTACH_DIR, exist_ok=True)
os.makedirs(ORDER_DIR, exist_ok=True)
os.makedirs(DRAFT_DIR, exist_ok=True)
os.makedirs(PACKAGE_DIR, exist_ok=True)

class OrderEmailPrepareRequest(BaseModel):
    orderer_name: Optional[str] = ''
    orderer_phone: Optional[str] = ''

class OrderBizboxMailRequest(BaseModel):
    bizbox_id: Optional[str] = ''
    bizbox_pw: Optional[str] = ''
    orderer_name: Optional[str] = ''
    orderer_phone: Optional[str] = ''
    # server: 서버 PC에서 Selenium 실행 / client: 웹 접속 PC의 로컬 에이전트에서 실행
    automation_target: Optional[str] = 'server'




def _create_order_qr_zip(order: PurchaseOrder) -> str:
    os.makedirs(DRAFT_DIR, exist_ok=True)
    req_no = order.purchase_request.request_no if order.purchase_request else order.order_no
    today = datetime.now().strftime('%Y%m%d')
    zip_path = os.path.join(DRAFT_DIR, f'{_safe_filename(req_no)}_{today}_QR코드.zip')
    qr_paths = []
    seen = set()
    for item in order.items:
        path = (item.qr_code_path or '').strip()
        if path and os.path.exists(path):
            ensure_qr_image_has_item_name(path, item)
        if path and os.path.exists(path) and path not in seen:
            seen.add(path)
            qr_paths.append(path)
    if not qr_paths:
        return ''
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for path in qr_paths:
            zf.write(path, arcname=os.path.basename(path))
    return zip_path

def _register_kor_font():
    for name, fp in [
        ("Kor", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        ("Kor", "C:/Windows/Fonts/malgun.ttf"),
    ]:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont(name, fp))
                return name
            except Exception:
                pass
    return "Helvetica"





# V31: 발주서/입고 품목 기준 정리
# - 견적서 파싱/기존 DB 값에서 단위 칼럼에 도번·규격이 잘못 들어온 경우 규격 칼럼으로 이동
# - 단위는 발주서/발주관리/입고현황 기준으로 항상 EA 사용
_KNOWN_UNIT_TOKENS = {
    'EA', 'EACH', '개', 'PCS', 'PC', 'SET', 'SETS', 'BOX', 'ROLL', 'M', 'MM', 'CM', 'KG', 'G', 'L', 'ML', '식', '본', '매', '조', '대'
}

def _clean_spec_unit_values(spec, unit):
    spec_txt = str(spec or '').strip()
    unit_txt = str(unit or '').strip()
    if unit_txt and unit_txt.upper() != 'EA':
        token = unit_txt.upper().replace('.', '').replace(' ', '')
        # 일반 단위표현은 버리고 EA로 통일한다. 그 외 값은 규격/도번으로 보고 규격 칼럼에 보존한다.
        if token not in _KNOWN_UNIT_TOKENS and unit_txt not in spec_txt:
            spec_txt = (spec_txt + ' ' + unit_txt).strip() if spec_txt else unit_txt
    return spec_txt, 'EA'

def _normalize_item_spec_unit(item):
    d = dict(item or {})
    spec, unit = _clean_spec_unit_values(d.get('spec', ''), d.get('unit', ''))
    d['spec'] = spec
    d['unit'] = unit
    return d

def _normalize_items_spec_unit(items):
    return [_normalize_item_spec_unit(it) for it in (items or [])]

def _safe_filename(name: str) -> str:
    bad = '<>:"/\\|?*'
    out = ''.join('_' if ch in bad else ch for ch in str(name or ''))
    return out.strip() or 'file'



# 품질관리/견적서 자동인식 보조 필드 처리
# 화면 입력값은 유지하고, DB/품질관리 엑셀용 숨은 값만 보강한다.
def _compact_text(value: str) -> str:
    import re
    return re.sub(r'\s+', '', str(value or '')).lower()


def _infer_item_group_from_text(text: str) -> str:
    """파일명/견적서 문맥/품명/규격에서 품목군을 추정한다."""
    t = _compact_text(text)
    if not t:
        return ''
    rules = [
        ('판금품', ['판금', 'sheetmetal', 'sheet-metal', 'sheet metal', 'sheetmetal']),
        ('선반품', ['선반', 'turning', 'lathe', 'cnc선반']),
        ('밀링품', ['밀링', 'milling', 'mct']),
        ('가공품', ['가공', 'machining', 'machinepart', '가공견적']),
        ('구매품', ['구매품', '구매', 'purchase']),
        ('전장품', ['전장', '전장품', '전장부품', '전장부품류', 'cable', '케이블']),
        ('프레임', ['프레임', 'frame']),
        ('석정반', ['석정반', 'granite']),
    ]
    for group, keys in rules:
        for key in keys:
            if _compact_text(key) in t:
                return group
    return ''


def _src_material_code(src: dict) -> str:
    """품질관리 요청 기준: 견적서의 규격/도번(spec)을 자재코드 보조값으로 우선 저장한다."""
    src = src or {}
    for key in ('material_code', 'code', 'part_no', 'partno', 'drawing_no', '도번', '품번'):
        val = str(src.get(key, '') or '').strip()
        if val:
            return val
    return str(src.get('spec', '') or '').strip()


def _src_item_group(src: dict, context: str = '') -> str:
    """견적서 자동인식/수기 품목의 DB용 품목군 값을 추정한다."""
    src = src or {}
    for key in ('item_group', 'group', '품목군', '구분', 'category', 'sub_category'):
        val = str(src.get(key, '') or '').strip()
        if val:
            return val
    ctx = ' '.join([
        str(context or ''),
        str(src.get('item_name', '') or ''),
        str(src.get('name', '') or ''),
        str(src.get('spec', '') or ''),
        str(src.get('note', '') or ''),
        str(src.get('maker', '') or ''),
        str(src.get('vendor_name', '') or ''),
    ])
    return _infer_item_group_from_text(ctx)


def _enrich_estimate_items_for_quality(items: list[dict], context: str = '') -> list[dict]:
    """견적서 자동인식 결과에 품질관리용 숨은 필드를 안전하게 채운다.

    이 함수가 없으면 /api/purchase/estimate/parse_multiple 이 500을 반환하고,
    브라우저에서는 `Unexpected token 'I', Internal Server Error is not valid JSON`로 보인다.
    """
    enriched = []
    group_from_context = _infer_item_group_from_text(context)
    for raw in items or []:
        it = dict(raw or {})
        if not str(it.get('material_code') or '').strip():
            mc = _src_material_code(it)
            if mc:
                it['material_code'] = mc
        if not str(it.get('item_group') or '').strip():
            it['item_group'] = group_from_context or _src_item_group(it, context)
        enriched.append(it)
    return enriched



def _order_package_filename(order: PurchaseOrder) -> str:
    pr = order.purchase_request
    req_no = (pr.request_no if pr else order.order_no) or order.order_no or ''
    title = (pr.title_full or pr.project_name or order.order_no) if pr else order.order_no
    vendor = order.vendor_name or '미정업체'
    return f"{_safe_filename(req_no)[:40]}_{_safe_filename(title)[:80]}_{_safe_filename(vendor)[:50]}.zip"


def _create_order_package_zip(order: PurchaseOrder) -> str:
    """발주 건 1개 기준 압축파일 생성: 발주서 PDF + QR ZIP + 메일초안 EML."""
    os.makedirs(PACKAGE_DIR, exist_ok=True)
    package_path = os.path.join(PACKAGE_DIR, _order_package_filename(order))
    files = []
    if order.pdf_path and os.path.exists(order.pdf_path):
        files.append(order.pdf_path)
    qr_zip = _create_order_qr_zip(order)
    if qr_zip and os.path.exists(qr_zip):
        files.append(qr_zip)
    if order.email_draft_path and os.path.exists(order.email_draft_path):
        files.append(order.email_draft_path)
    if not files:
        return ''
    with zipfile.ZipFile(package_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        used = set()
        for fp in files:
            name = os.path.basename(fp)
            base, ext = os.path.splitext(name)
            arc = name
            n = 2
            while arc in used:
                arc = f"{base}_{n}{ext}"; n += 1
            used.add(arc)
            zf.write(fp, arcname=arc)
    return package_path


def _norm_vendor_name(v):
    if not isinstance(v, dict):
        return str(v or '').strip()
    return str(v.get('name') or v.get('vendor_name') or v.get('company') or '미정').strip() or '미정'


def _vendor_key(v) -> str:
    import re
    return re.sub(r'[^0-9A-Za-z가-힣]', '', str(v or '').lower())


def _item_vendor_candidates(item: dict) -> list[str]:
    # 견적서 자동인식/수기작성 모두 마지막 칼럼인 비고(제조사)를 업체 구분값으로 사용할 수 있게 한다.
    keys = [
        'vendor_name', 'vendor', 'company', 'supplier', 'maker', 'manufacturer', 'brand',
        'note', 'remark', 'remarks', 'comment', '비고', '제조사', '비고(제조사)',
    ]
    out = []
    for k in keys:
        v = str((item or {}).get(k, '') or '').strip()
        if v and v not in out:
            out.append(v)
    return out


def _item_vendor_match(item: dict, vendor: dict, idx: int) -> bool:
    vi = item.get('vendor_index', None)
    if vi is not None and str(vi) != '' and str(vi) == str(vendor.get('vendor_index', idx)):
        return True
    vname = _norm_vendor_name(vendor)
    vkey = _vendor_key(vname)
    if not vkey:
        return False
    for cand in _item_vendor_candidates(item):
        ckey = _vendor_key(cand)
        if ckey and (ckey == vkey or ckey in vkey or vkey in ckey):
            return True
    return False


def _infer_vendors_from_items(items: list[dict]) -> list[dict]:
    vendors = []
    seen = set()
    for it in items or []:
        for cand in _item_vendor_candidates(it):
            key = _vendor_key(cand)
            # 품명/도번/일반 설명이 업체로 들어가는 것을 줄이기 위해 너무 짧은 값은 제외한다.
            if len(key) < 2 or key in seen:
                continue
            seen.add(key)
            vendors.append({'name': cand, 'vendor_name': cand, 'email': '', 'contact': '', 'vendor_index': len(vendors)})
            break
    return vendors



def _merge_vendors_with_item_vendors(vendors: list[dict], items: list[dict]) -> list[dict]:
    """견적서 자동인식/수기행의 vendor_name·비고(제조사)를 업체 목록에 병합한다.

    화면의 추천업체 표가 기존 1행만 남아 있거나 일부 업체가 누락되어도,
    품목에 들어있는 vendor_index/vendor_name 기준으로 발주관리·입고현황에
    업체별 PurchaseOrder/ReceiptItem이 빠지지 않도록 보정한다.
    """
    merged: list[dict] = []

    def _has_vendor(name: str) -> bool:
        key = _vendor_key(name)
        if not key:
            return False
        for v in merged:
            vk = _vendor_key(_norm_vendor_name(v))
            if vk and (vk == key or vk in key or key in vk):
                return True
        return False

    for idx, v in enumerate(vendors or []):
        if isinstance(v, dict):
            vv = dict(v)
            vv.setdefault('vendor_index', idx)
            vv.setdefault('name', vv.get('vendor_name') or vv.get('company') or vv.get('vendor') or '')
            vv.setdefault('vendor_name', vv.get('name') or '')
            vv.setdefault('reason', vv.get('reason') or '기존거래업체')
        else:
            vv = {'name': str(v or '').strip(), 'vendor_name': str(v or '').strip(), 'vendor_index': idx, 'reason': '기존거래업체'}
        name = _norm_vendor_name(vv)
        if name and name != '미정' and not _has_vendor(name):
            merged.append(vv)

    # 품목 자체에 보존된 업체명/vendor_index를 기준으로 누락 업체를 다시 추가한다.
    for it in items or []:
        for cand in _item_vendor_candidates(it):
            cand = str(cand or '').strip()
            key = _vendor_key(cand)
            if len(key) < 2:
                continue
            if _has_vendor(cand):
                break
            vi = it.get('vendor_index', None)
            try:
                vi = int(vi) if vi is not None and str(vi) != '' else len(merged)
            except Exception:
                vi = len(merged)
            merged.append({
                'name': cand,
                'vendor_name': cand,
                'email': '',
                'contact': '',
                'phone': '',
                'vendor_index': vi,
                'reason': '기존거래업체',
            })
            break

    if not merged:
        merged = [{'name': '미정', 'vendor_name': '미정', 'email': '', 'contact': '', 'vendor_index': 0, 'reason': '기존거래업체'}]
    return merged


def _safe_int_value(v):
    try:
        return int(float(v or 0))
    except Exception:
        return 0


def _safe_float_value(v):
    try:
        return float(str(v or 0).replace(',', ''))
    except Exception:
        return 0.0


def _same_receipt_src_item(db_item, src: dict) -> bool:
    return (
        str(db_item.item_name or '').strip() == str(src.get('item_name', '') or '').strip()
        and str(db_item.spec or '').strip() == str(src.get('spec', '') or '').strip()
    )


def _receipt_delete_key_text(value) -> str:
    return ' '.join(str(value or '').replace('\u3000', ' ').split()).strip().lower()


def _receipt_item_repair_key_values(item_name='', spec='', material_code='') -> str:
    """자동 보정 재생성 방지용 대표 품목 키.

    입고현황에서 관리자가 삭제한 품목은 구매의뢰 원본 items_json에는 계속 남아 있으므로,
    /api/inventory/receipt_list의 보정 로직이 다음 조회 때 다시 만들 수 있다.
    그래서 발주 건별로 삭제 품목 키를 저장하고, 보정 시 같은 키는 다시 생성하지 않는다.
    """
    return '|'.join([
        _receipt_delete_key_text(item_name),
        _receipt_delete_key_text(spec),
        _receipt_delete_key_text(material_code),
    ])


def _receipt_item_repair_key_variants_values(item_name='', spec='', material_code='') -> set[str]:
    """삭제 품목 재생성 방지용 확장 키 묶음.

    운영 DB에는 과거 패치 이력에 따라 견적서의 `규격/도번`이
    receipt_items.spec, material_code, item_name 중 다른 칸으로 들어간 경우가 있다.
    대표키 하나만 저장하면 자동 보정 시 원본 items_json의 키와 달라져서 삭제 품목이
    다시 살아날 수 있으므로, 품명/규격/자재코드 단독 키와 조합 키를 함께 저장한다.
    """
    name = _receipt_delete_key_text(item_name)
    spec_v = _receipt_delete_key_text(spec)
    code = _receipt_delete_key_text(material_code)
    keys = {
        _receipt_item_repair_key_values(name, spec_v, code),
        _receipt_item_repair_key_values(name, spec_v, ''),
        _receipt_item_repair_key_values(name, '', code),
        _receipt_item_repair_key_values('', spec_v, code),
    }
    if name:
        keys.add(f'name::{name}')
    if spec_v:
        keys.add(f'spec::{spec_v}')
    if code:
        keys.add(f'code::{code}')
    return {k for k in keys if k and k != '||'}


def _receipt_item_repair_key_variants_from_src(src: dict) -> set[str]:
    src = src or {}
    return _receipt_item_repair_key_variants_values(
        src.get('item_name', '') or '',
        src.get('spec', '') or '',
        _src_material_code(src),
    )


def _receipt_item_repair_key_from_src(src: dict) -> str:
    return _receipt_item_repair_key_values(
        src.get('item_name', '') or '',
        src.get('spec', '') or '',
        _src_material_code(src),
    )


def _receipt_item_repair_key_variants_from_db_item(item) -> set[str]:
    return _receipt_item_repair_key_variants_values(
        getattr(item, 'item_name', '') or '',
        getattr(item, 'spec', '') or '',
        getattr(item, 'material_code', '') or '',
    )


def _receipt_item_repair_key_from_db_item(item) -> str:
    return _receipt_item_repair_key_values(
        getattr(item, 'item_name', '') or '',
        getattr(item, 'spec', '') or '',
        getattr(item, 'material_code', '') or '',
    )


def _json_deleted_key_set(raw_value) -> set[str]:
    """JSON 문자열/리스트에서 삭제 품목 tombstone 키를 안전하게 읽는다."""
    try:
        data = json.loads(raw_value or '[]') if isinstance(raw_value, str) else (raw_value or [])
    except Exception:
        data = []
    if not isinstance(data, list):
        data = []
    return {str(x or '') for x in data if str(x or '')}


def _load_deleted_receipt_item_keys_for_request(pr: PurchaseRequest) -> set[str]:
    if not pr:
        return set()
    return _json_deleted_key_set(getattr(pr, 'deleted_receipt_item_keys_json', '') or '')


def _save_deleted_receipt_item_keys_for_request(pr: PurchaseRequest, keys: set[str]) -> None:
    if not pr:
        return
    cleaned = sorted({str(k or '') for k in keys if str(k or '')})
    pr.deleted_receipt_item_keys_json = json.dumps(cleaned, ensure_ascii=False) if cleaned else ''


def _load_deleted_receipt_item_keys(order: PurchaseOrder) -> set[str]:
    """발주 건 tombstone + 구매의뢰서 tombstone을 함께 읽는다."""
    if not order:
        return set()
    keys = _json_deleted_key_set(getattr(order, 'deleted_receipt_item_keys_json', '') or '')
    try:
        pr = getattr(order, 'purchase_request', None)
        keys.update(_load_deleted_receipt_item_keys_for_request(pr))
    except Exception:
        pass
    return keys


def _load_deleted_receipt_item_keys_order_only(order: PurchaseOrder) -> set[str]:
    if not order:
        return set()
    return _json_deleted_key_set(getattr(order, 'deleted_receipt_item_keys_json', '') or '')


def _save_deleted_receipt_item_keys(order: PurchaseOrder, keys: set[str]) -> None:
    if not order:
        return
    cleaned = sorted({str(k or '') for k in keys if str(k or '')})
    order.deleted_receipt_item_keys_json = json.dumps(cleaned, ensure_ascii=False) if cleaned else ''


def _remember_deleted_receipt_item_for_request(pr: PurchaseRequest, item) -> str:
    """품목 삭제 tombstone을 구매의뢰서에 보존한다.

    발주 업체의 마지막 품목을 삭제하면 빈 PurchaseOrder 자체는 삭제하는 것이 맞다.
    대신 재조회 시 자동 보정 로직이 같은 품목을 다시 만들지 않도록 tombstone은
    구매의뢰서(purchase_requests)에 저장한다.
    """
    if not pr or not item:
        return ''
    variants = _receipt_item_repair_key_variants_from_db_item(item)
    if not variants:
        return ''
    keys = _load_deleted_receipt_item_keys_for_request(pr)
    keys.update(variants)
    _save_deleted_receipt_item_keys_for_request(pr, keys)
    return sorted(variants)[0]


def _remember_deleted_receipt_item_for_order(order: PurchaseOrder, item) -> str:
    """관리자 입고현황 품목 삭제를 자동 보정 예외로 저장한다.

    v50부터는 빈 업체/발주 건을 DB에 남기지 않고 삭제하므로, 발주 건과 함께
    구매의뢰서에도 tombstone을 저장한다. 기존 DB 호환을 위해 발주 건 컬럼에도
    한 번 더 저장한다.
    """
    if not order or not item:
        return ''
    variants = _receipt_item_repair_key_variants_from_db_item(item)
    if not variants:
        return ''
    keys = _load_deleted_receipt_item_keys_order_only(order)
    keys.update(variants)
    _save_deleted_receipt_item_keys(order, keys)
    try:
        _remember_deleted_receipt_item_for_request(getattr(order, 'purchase_request', None), item)
    except Exception:
        pass
    return sorted(variants)[0]


def _src_deleted_by_request_tombstone(pr: PurchaseRequest, src: dict) -> bool:
    if not pr:
        return False
    deleted_keys = _load_deleted_receipt_item_keys_for_request(pr)
    src_keys = _receipt_item_repair_key_variants_from_src(src)
    return bool(deleted_keys and src_keys and (deleted_keys & src_keys))


def _move_order_deleted_receipt_tombstones_to_request(order: PurchaseOrder) -> int:
    """빈 발주 건 삭제 전에 발주 건 tombstone을 구매의뢰서로 이동한다."""
    if not order:
        return 0
    pr = getattr(order, 'purchase_request', None)
    if not pr:
        return 0
    order_keys = _load_deleted_receipt_item_keys_order_only(order)
    if not order_keys:
        return 0
    req_keys = _load_deleted_receipt_item_keys_for_request(pr)
    before = len(req_keys)
    req_keys.update(order_keys)
    _save_deleted_receipt_item_keys_for_request(pr, req_keys)
    return max(0, len(req_keys) - before)


def _cleanup_empty_receipt_orders(db: Session) -> int:
    """품목이 0개인 빈 업체/발주 건을 실제 DB에서 삭제한다.

    삭제 품목 재생성 방지 정보는 purchase_requests.deleted_receipt_item_keys_json에 보존한다.
    이 함수는 v48/v49에서 tombstone 보존용으로 남긴 빈 발주 건도 정리한다.
    """
    deleted = 0
    try:
        orders = db.query(PurchaseOrder).all()
    except Exception:
        return 0
    for order in list(orders):
        try:
            if len(getattr(order, 'items', []) or []) > 0:
                continue
            _move_order_deleted_receipt_tombstones_to_request(order)
            db.delete(order)
            deleted += 1
        except Exception:
            pass
    if deleted:
        try:
            db.flush()
        except Exception:
            pass
    return deleted


def _unique_order_no_for_request(db: Session, pr: PurchaseRequest, idx: int) -> str:
    base = pr.request_no or f"{datetime.now().strftime('%Y')}-{db.query(PurchaseOrder).count()+1:04d}"
    candidate = base if idx == 0 else f"{base}-{idx+1}"
    if not db.query(PurchaseOrder).filter(PurchaseOrder.order_no == candidate).first():
        return candidate
    n = idx + 1
    while True:
        candidate = f"{base}-{n+1}"
        if not db.query(PurchaseOrder).filter(PurchaseOrder.order_no == candidate).first():
            return candidate
        n += 1


def _ensure_receipt_items_for_order(db: Session, po: PurchaseOrder, src_items: list[dict]) -> tuple[int, int]:
    """발주 건에 품목/QR이 빠져 있으면 보정한다. return=(items_created, qr_created)."""
    created_items = 0
    created_qr = 0
    for src in _normalize_items_spec_unit(src_items or []):
        if not str(src.get('item_name', '') or src.get('spec', '') or '').strip():
            continue
        ri = None
        for old in po.items:
            fixed_old_spec, fixed_old_unit = _clean_spec_unit_values(old.spec or '', old.unit or 'EA')
            if old.spec != fixed_old_spec or old.unit != fixed_old_unit:
                old.spec = fixed_old_spec
                old.unit = fixed_old_unit
            if _same_receipt_src_item(old, src):
                ri = old
                break

        src_qty = _safe_int_value(src.get('quantity', 0))
        src_amount = _safe_float_value(src.get('amount') or src.get('total') or src.get('total_amount') or 0)
        src_unit_price = _safe_float_value(src.get('unit_price', 0))
        if src_amount and not src_unit_price:
            src_unit_price = src_amount / (src_qty or 1)

        if ri is None:
            deleted_keys = _load_deleted_receipt_item_keys(po)
            src_keys = _receipt_item_repair_key_variants_from_src(src)
            if deleted_keys and src_keys and (deleted_keys & src_keys):
                # 관리자가 입고현황에서 의도적으로 삭제한 품목은 자동 보정으로 되살리지 않는다.
                continue
            ri = ReceiptItem(
                order_id=po.id,
                item_name=src.get('item_name', '') or '',
                spec=src.get('spec', '') or '',
                quantity=src_qty,
                unit=src.get('unit') or 'EA',
                unit_price=src_unit_price,
                maker=(src.get('maker') or src.get('vendor_name') or src.get('vendor') or src.get('note') or ''),
                note=(src.get('note') or src.get('maker') or src.get('vendor_name') or ''),
                material_code=_src_material_code(src),
                item_group=_src_item_group(src, po.purchase_request.title_full if po and po.purchase_request else ''),
                axis_type=src.get('axis') or src.get('axis_type') or '',
                order_round=src.get('order_round') or '',
                stage='미입고',
            )
            db.add(ri)
            db.flush()
            created_items += 1
        else:
            if not ri.quantity:
                ri.quantity = src_qty
            if not ri.unit_price:
                ri.unit_price = src_unit_price
            fixed_spec, fixed_unit = _clean_spec_unit_values(ri.spec or src.get('spec', ''), ri.unit or src.get('unit', 'EA'))
            ri.spec = fixed_spec
            ri.unit = fixed_unit or 'EA'

        maker_val = src.get('maker') or src.get('vendor_name') or src.get('vendor') or src.get('note') or ''
        if maker_val and not getattr(ri, 'maker', ''):
            ri.maker = maker_val
        note_val = src.get('note') or src.get('maker') or src.get('vendor_name') or ''
        if note_val and not getattr(ri, 'note', ''):
            ri.note = note_val
        for attr, keys in {
            'material_code': ['material_code', 'code', 'spec'],
            'item_group': ['item_group', 'group'],
            'axis_type': ['axis', 'axis_type'],
            'order_round': ['order_round'],
        }.items():
            if not getattr(ri, attr, ''):
                for key in keys:
                    val = src.get(key)
                    if val:
                        setattr(ri, attr, val)
                        break
        if not getattr(ri, 'material_code', '') and getattr(ri, 'spec', ''):
            ri.material_code = ri.spec
        if not getattr(ri, 'item_group', ''):
            g = _src_item_group(src, po.purchase_request.title_full if po and po.purchase_request else '')
            if g:
                ri.item_group = g
        try:
            if not ri.qr_code or not ri.qr_code_path or not os.path.exists(str(ri.qr_code_path)):
                generate_qr_for_item(ri, po, db)
                created_qr += 1
        except Exception:
            pass
    return created_items, created_qr


def _request_items_and_vendors_for_orders(pr: PurchaseRequest) -> tuple[list[dict], list[dict]]:
    # 입고관리/QR 보정용: 실제입고품 다름이면 실제 입고품 목록과 실제입고품 업체명만 기준으로 한다.
    # 비용처리용 견적 업체(vendors_json)가 입고관리/QR 업체로 섞이면 안 된다.
    items = _receipt_request_items(pr)
    if not items:
        items = [_normalize_item_spec_unit({
            'item_name': pr.item_name or '',
            'spec': pr.spec or '',
            'quantity': pr.quantity or 0,
            'unit': pr.unit or 'EA',
            'unit_price': pr.unit_price or 0,
        })]
    items = _normalize_items_spec_unit([
        it for it in items
        if str(it.get('item_name', '') or it.get('spec', '') or '').strip()
    ])
    if int(getattr(pr, 'actual_received_diff', 0) or 0):
        vendors = _vendors_for_request_items(pr, items, False)
    else:
        try:
            vendors = json.loads(pr.vendors_json or '[]') if pr.vendors_json else []
        except Exception:
            vendors = []
        vendors = _merge_vendors_with_item_vendors(vendors, items)
    return items, vendors


def _repair_missing_vendor_orders_for_request(db: Session, pr: PurchaseRequest) -> dict:
    """기존 DB/패치 과정에서 일부 견적 업체가 발주관리·입고현황으로 안 넘어간 건을 자동 보정."""
    if not pr or not pr.id:
        return {'created_orders': 0, 'created_items': 0, 'created_qr': 0}
    _classify_order_types_for_request(db, pr)
    items, vendors = _request_items_and_vendors_for_orders(pr)
    if not items or not vendors:
        return {'created_orders': 0, 'created_items': 0, 'created_qr': 0}

    # 이전 버전에서 자동 보정 방지용으로 남긴 품목 0개 발주 건은 먼저 정리한다.
    # tombstone은 구매의뢰서 컬럼으로 이동되므로 다시 생성되지 않는다.
    try:
        _cleanup_empty_receipt_orders(db)
    except Exception:
        pass

    orders = [o for o in db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id).order_by(PurchaseOrder.id.asc()).all() if _is_receipt_order(o)]
    created_orders = 0
    created_items = 0
    created_qr = 0

    def _items_for_vendor_local(vendor, idx):
        matched = [it for it in items if _item_vendor_match(it, vendor, idx)]
        if not matched and len(vendors) <= 1:
            matched = items
        return matched

    for idx, vendor in enumerate(vendors):
        vendor_items = _items_for_vendor_local(vendor, idx)
        # 관리자가 입고현황에서 삭제한 품목은 구매의뢰서 tombstone 기준으로 제외한다.
        # 업체의 모든 품목이 삭제된 경우에는 발주 건 자체를 다시 만들지 않는다.
        vendor_items = [it for it in vendor_items if not _src_deleted_by_request_tombstone(pr, it)]
        if not vendor_items:
            continue
        vname = _norm_vendor_name(vendor)
        vkey = _vendor_key(vname)
        po = None
        for existing in orders:
            ekey = _vendor_key(existing.vendor_name or '')
            if ekey and vkey and (ekey == vkey or ekey in vkey or vkey in ekey):
                po = existing
                break
        if po is None:
            po = PurchaseOrder(
                order_no=_unique_order_no_for_request(db, pr, idx),
                request_id=pr.id,
                vendor_name=vname,
                vendor_email=vendor.get('email', '') if isinstance(vendor, dict) else '',
                vendor_contact=vendor.get('contact', '') if isinstance(vendor, dict) else '',
                delivery_date=pr.required_date,
                status='상신완료',
            )
            _set_order_type(po, 'receipt' if int(pr.actual_received_diff or 0) else 'normal')
            db.add(po)
            db.flush()
            orders.append(po)
            created_orders += 1
        ci, cq = _ensure_receipt_items_for_order(db, po, vendor_items)
        created_items += ci
        created_qr += cq

    # 기존 DB에 발주 건만 있고 ReceiptItem이 0개인 경우에는 업체 매칭 실패 여부와 무관하게
    # 첫 발주 건에 전체 품목을 강제로 보정한다. 이 케이스가 입고현황 공백의 주요 원인이다.
    try:
        total_existing_items = sum(len(getattr(o, 'items', []) or []) for o in orders)
    except Exception:
        total_existing_items = 0
    if orders and total_existing_items <= 0 and items:
        repair_items = [it for it in items if not _src_deleted_by_request_tombstone(pr, it)]
        if repair_items:
            ci, cq = _ensure_receipt_items_for_order(db, orders[0], repair_items)
            created_items += ci
            created_qr += cq

    if created_orders or created_items or created_qr:
        db.flush()
    return {'created_orders': created_orders, 'created_items': created_items, 'created_qr': created_qr}


def _repair_missing_vendor_orders_all(db: Session) -> dict:
    total = {'created_orders': 0, 'created_items': 0, 'created_qr': 0}
    prs = db.query(PurchaseRequest).filter(PurchaseRequest.status.in_(['상신완료', '결재중', '발주서전송완료'])).all()
    for pr in prs:
        r = _repair_missing_vendor_orders_for_request(db, pr)
        for k in total:
            total[k] += int(r.get(k, 0) or 0)
    if any(total.values()):
        db.commit()
    return total




def _safe_json_loads_purchase(value, default=None):
    """기존 DB 호환용 JSON 파서. 과거 DB에 일반 문자열/깨진 JSON이 있어도 목록 API가 죽지 않게 한다."""
    if default is None:
        default = []
    if value is None or value == '':
        return default
    try:
        parsed = json.loads(value)
        return parsed if parsed is not None else default
    except Exception:
        return default


def _dt_to_text_purchase(value, fmt='%Y-%m-%d %H:%M'):
    if not value:
        return ''
    try:
        return value.strftime(fmt)
    except Exception:
        return str(value)


def _date_to_text_purchase(value):
    if not value:
        return ''
    try:
        return value.strftime('%Y-%m-%d')
    except Exception:
        return str(value)[:10]


def _fallback_order_detail_row(group):
    """상세 목록 생성 중 일부 구 DB 레코드가 깨져도 발주관리/대시보드가 비지 않게 하는 최소 행."""
    base = group[0]
    try:
        pr = base.purchase_request
    except Exception:
        pr = None
    all_items = []
    vendor_names = []
    for o in group:
        vendor_names.append(str(getattr(o, 'vendor_name', '') or '').strip() or '미정업체')
        try:
            items_iter = list(getattr(o, 'items', []) or [])
        except Exception:
            items_iter = []
        for i in items_iter:
            all_items.append({
                'id': getattr(i, 'id', None),
                'order_id': getattr(o, 'id', None),
                'order_no': getattr(o, 'order_no', '') or '',
                'vendor_name': getattr(o, 'vendor_name', '') or '',
                'item_name': getattr(i, 'item_name', '') or '',
                'spec': getattr(i, 'spec', '') or '',
                'quantity': getattr(i, 'quantity', '') or '',
                'unit': getattr(i, 'unit', '') or 'EA',
                'unit_price': getattr(i, 'unit_price', 0) or 0,
                'maker': getattr(i, 'maker', '') or '',
                'note': getattr(i, 'note', '') or '',
                'material_code': getattr(i, 'material_code', '') or '',
                'item_group': getattr(i, 'item_group', '') or '',
                'axis_type': getattr(i, 'axis_type', '') or '',
                'order_round': getattr(i, 'order_round', '') or '',
                'qr_code': getattr(i, 'qr_code', '') or '',
                'qr_code_path': getattr(i, 'qr_code_path', '') or '',
            })
    try:
        total_items_for_receipt, received_items_for_receipt, all_items_received = _order_group_receipt_counts(group)
    except Exception:
        total_items_for_receipt, received_items_for_receipt, all_items_received = len(all_items), 0, False
    try:
        status = _group_order_status(group)
    except Exception:
        status = str(getattr(base, 'status', '') or '')
    try:
        vendor_display = _vendor_display_name(vendor_names)
    except Exception:
        vendor_display = vendor_names[0] if vendor_names else ''
    return {
        'id': getattr(base, 'id', None),
        'order_ids': [getattr(o, 'id', None) for o in group],
        'request_id': getattr(pr, 'id', None) if pr else getattr(base, 'request_id', None),
        'request_no': getattr(pr, 'request_no', '') if pr else '',
        'requested_by': (getattr(pr, 'requested_by', '') or getattr(pr, 'requester', '')) if pr else '',
        'requester': getattr(pr, 'requester', '') if pr else '',
        'order_completed_by': '',
        'order_completed_at': '',
        'tax_docs_completed_by': '',
        'tax_docs_completed_at': '',
        'order_no': (getattr(pr, 'request_no', '') if pr else '') or getattr(base, 'order_no', ''),
        'raw_order_no': getattr(base, 'order_no', '') or '',
        'project_code': getattr(pr, 'project_code', '') if pr else '',
        'project_name': getattr(pr, 'project_name', '') if pr else '',
        'title_full': getattr(pr, 'title_full', '') if pr else '',
        'inbound_by': _join_people_names([it.get('purchase_recv_by') for it in all_items]),
        'outbound_by': _join_people_names([it.get('quality_recv_by') for it in all_items]),
        'vendor_name': vendor_display,
        'vendor_names': vendor_names,
        'vendor_count': len(group),
        'vendor_email': getattr(base, 'vendor_email', '') or '',
        'delivery_date': _date_to_text_purchase(getattr(base, 'delivery_date', None)),
        'status': status,
        'attach_files': [],
        'reason': getattr(pr, 'reason', '') if pr else '',
        'actual_received_diff': bool(getattr(pr, 'actual_received_diff', 0)) if pr else False,
        'actual_items': [],
        'vendor_groups': [],
        'items': all_items,
        'item_count': len(all_items),
        'qr_count': sum(1 for it in all_items if it.get('qr_code')),
        'receipt_total_count': total_items_for_receipt,
        'receipt_received_count': received_items_for_receipt,
        'receipt_pending_count': max(0, total_items_for_receipt - received_items_for_receipt),
        'all_items_received': all_items_received,
        'representative_qr': next((it.qr_code for o in receipt_group for it in (o.items or []) if getattr(it, 'qr_code', None)), ''),
        'detail_warning': 'fallback_row_used',
    }

def _estimate_vendor_from_file(filename: str) -> str:
    base = os.path.splitext(os.path.basename(filename or ''))[0]
    return base[:50] or '미정업체'






def _xls_direct_text(value) -> str:
    return '' if value is None else str(value).strip()


def _looks_like_html_xls_for_purchase(raw: bytes) -> bool:
    head = raw[:8192].lower()
    return b'<html' in head or b'<table' in head or b'<tr' in head


def _decode_legacy_xls_html_for_purchase(raw: bytes) -> str:
    for enc in ('utf-8-sig', 'cp949', 'euc-kr', 'utf-16', 'latin1'):
        try:
            txt = raw.decode(enc)
            if '<' in txt and '>' in txt:
                return txt
        except Exception:
            pass
    return raw.decode('latin1', errors='ignore')


def _read_html_xls_rows_for_purchase(path: str):
    """Excel COM 없이 HTML table 형식의 .xls 견적서를 직접 읽는다."""
    try:
        raw = Path(path).read_bytes()
        if not _looks_like_html_xls_for_purchase(raw):
            return []
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(_decode_legacy_xls_html_for_purchase(raw), 'html.parser')
        sheets = []
        for ti, table in enumerate(soup.find_all('table'), start=1):
            rows = []
            for tr in table.find_all('tr'):
                cells = []
                for cell in tr.find_all(['th', 'td']):
                    cells.append(_xls_direct_text(cell.get_text(' ', strip=True)))
                while cells and not cells[-1]:
                    cells.pop()
                if any(cells):
                    rows.append(cells)
            if rows:
                sheets.append((f'HTML_TABLE_{ti}', rows))
        return sheets
    except Exception:
        return []


def _xls_cell_text_for_purchase(cell, book) -> str:
    try:
        import xlrd
        if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
            return ''
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                from datetime import datetime
                vals = xlrd.xldate_as_tuple(cell.value, book.datemode)
                if vals[3:] == (0, 0, 0):
                    return datetime(*vals).strftime('%Y-%m-%d')
                return datetime(*vals).strftime('%Y-%m-%d %H:%M')
            except Exception:
                return _xls_direct_text(cell.value)
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            try:
                v = float(cell.value)
                return str(int(v)) if v.is_integer() else ('%f' % v).rstrip('0').rstrip('.')
            except Exception:
                return _xls_direct_text(cell.value)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return 'TRUE' if bool(cell.value) else 'FALSE'
        return _xls_direct_text(cell.value)
    except Exception:
        return _xls_direct_text(getattr(cell, 'value', ''))


def _read_xlrd_xls_rows_for_purchase(path: str):
    """Excel COM/LibreOffice 없이 BIFF .xls를 직접 읽는다."""
    try:
        import xlrd
        book = xlrd.open_workbook(path, on_demand=True)
        sheets = []
        for sh in book.sheets():
            rows = []
            max_r = min(sh.nrows, 300)
            max_c = min(sh.ncols, 80)
            for r in range(max_r):
                row = []
                for c in range(max_c):
                    row.append(_xls_cell_text_for_purchase(sh.cell(r, c), book))
                while row and not row[-1]:
                    row.pop()
                if any(row):
                    rows.append(row)
            if rows:
                sheets.append((str(sh.name), rows))
        try:
            book.release_resources()
        except Exception:
            pass
        return sheets
    except Exception:
        return []


def _read_legacy_xls_rows_direct_for_purchase(path: str):
    """구형 .xls를 변환하지 않고 직접 읽는다.

    처리 순서: HTML-xls -> BIFF-xls(xlrd). 실패 시 빈 목록 반환.
    """
    sheets = _read_html_xls_rows_for_purchase(path)
    if sheets:
        return sheets, ''
    sheets = _read_xlrd_xls_rows_for_purchase(path)
    if sheets:
        return sheets, ''
    return [], '직접 .xls 읽기 실패 또는 xlrd 미설치'

def _find_soffice_exe_for_purchase():
    for name in ('soffice', 'libreoffice'):
        p = shutil.which(name)
        if p:
            return p
    if os.name == 'nt':
        for p in [
            r'C:\Program Files\LibreOffice\program\soffice.exe',
            r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
        ]:
            if os.path.exists(p):
                return p
    return ''


def _convert_legacy_xls_for_purchase(path: str):
    """server-side robust parser용 구형 .xls -> 임시 .xlsx 변환."""
    path = os.path.abspath(path)
    tmp_dir = tempfile.mkdtemp(prefix='deverp_xls_')
    out_path = os.path.join(tmp_dir, os.path.splitext(os.path.basename(path))[0] + '.xlsx')
    errors = []
    if os.name == 'nt':
        try:
            import pythoncom
            import win32com.client
            pythoncom.CoInitialize()
            excel = win32com.client.DispatchEx('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(path)
            wb.SaveAs(out_path, FileFormat=51)
            wb.Close(False)
            excel.Quit()
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
            if os.path.exists(out_path) and os.path.getsize(out_path) > 0:
                return out_path, tmp_dir, ''
        except Exception as e:
            errors.append(f'Excel COM 변환 실패: {e}')
            try:
                excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
    soffice = _find_soffice_exe_for_purchase()
    if soffice:
        try:
            cp = subprocess.run([soffice, '--headless', '--convert-to', 'xlsx', '--outdir', tmp_dir, path], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=60)
            candidates = sorted(Path(tmp_dir).glob('*.xlsx'), key=lambda q: q.stat().st_mtime, reverse=True)
            if candidates and candidates[0].exists() and candidates[0].stat().st_size > 0:
                return str(candidates[0]), tmp_dir, ''
            errors.append((cp.stdout or '') + ' ' + (cp.stderr or ''))
        except Exception as e:
            errors.append(f'LibreOffice 변환 실패: {e}')
    shutil.rmtree(tmp_dir, ignore_errors=True)
    return '', '', '; '.join(x for x in errors if x).strip() or '구형 .xls 변환 도구를 찾지 못했습니다.'

def _robust_xlsx_parse(path: str) -> dict:
    """MODEL/품번/도번/규격 구조 견적서 보조 인식.

    .xls는 Excel COM/LibreOffice가 없어도 HTML-xls 또는 xlrd 기반 BIFF-xls로 먼저 직접 읽는다.
    직접 읽기에 실패한 경우에만 Excel COM 또는 LibreOffice 임시 xlsx 변환을 시도한다.
    """
    tmp_dir = ''
    try:
        import openpyxl, re
        read_path = path
        sheet_sources = []

        if os.path.splitext(path)[1].lower() == '.xls':
            direct_sheets, _direct_err = _read_legacy_xls_rows_direct_for_purchase(path)
            if direct_sheets:
                sheet_sources = direct_sheets
                read_path = ''
            else:
                converted, tmp_dir, conv_err = _convert_legacy_xls_for_purchase(path)
                if converted:
                    read_path = converted
                else:
                    return {'items': [], 'vendor_name': '', 'error': conv_err or _direct_err}

        if read_path:
            wb = openpyxl.load_workbook(read_path, data_only=True)
            for ws in wb.worksheets:
                merged = {}
                for mr in ws.merged_cells.ranges:
                    val = ws.cell(mr.min_row, mr.min_col).value
                    for rr in range(mr.min_row, mr.max_row + 1):
                        for cc in range(mr.min_col, mr.max_col + 1):
                            merged[(rr, cc)] = val
                rows = []
                for rr in range(1, min(ws.max_row, 120) + 1):
                    row = []
                    for cc in range(1, min(ws.max_column, 40) + 1):
                        v = merged.get((rr, cc), ws.cell(rr, cc).value)
                        row.append('' if v is None else str(v).strip())
                    if any(row):
                        rows.append(row)
                if rows:
                    sheet_sources.append((str(ws.title), rows))

        all_items = []
        vendor = ''
        for _sheet_name, rows in sheet_sources:
            if not vendor:
                joined = ' '.join(' '.join(r) for r in rows[:12])
                m = re.search(r'(?:공\s*급\s*자|상\s*호|업\s*체\s*명)\s*([가-힣A-Za-z0-9㈜()\s]{2,30})', joined)
                if m:
                    vendor = m.group(1).strip()
                if not vendor:
                    m2 = re.search(r'([가-힣A-Za-z0-9._-]{2,}(?:주식회사|㈜|\(주\)))', joined)
                    if m2:
                        vendor = m2.group(1).strip()
                if vendor:
                    toks = [x for x in re.split(r'\s+', vendor) if x]
                    for tok in toks:
                        if ('㈜' in tok or '(주)' in tok or '주식회사' in tok) and len(tok) >= 3:
                            vendor = tok
                    for stop in ['대표이사','대표','사업자','주소','TEL','전화','FAX']:
                        if stop in vendor:
                            vendor = vendor.split(stop)[0].strip()

            def norm(x):
                return re.sub(r'\s+', '', str(x or '')).lower()

            header_i = -1
            cols = {}
            for i, row in enumerate(rows[:50]):
                m = {}
                for j, cell in enumerate(row):
                    n = norm(cell)
                    if not n:
                        continue
                    if n in ('model','modle','모델') or 'model' in n or 'modle' in n or n == '품명' or '품목' in n or n == '명칭':
                        m.setdefault('item', j)
                    if n == '규격' or '규격' in n:
                        m.setdefault('spec_reg', j)
                    if n in ('품번','도번') or '품번' in n or '도번' in n:
                        m.setdefault('spec_draw', j)
                    if '수량' in n or n in ('qty', "q'ty", 'quantity'):
                        m.setdefault('qty', j)
                    if n == '단가' or '단가' in n or n in ('price','unitprice'):
                        m.setdefault('price', j)
                    if n == '금액' or '금액' in n or 'amount' in n:
                        m.setdefault('amount', j)
                    if '단위' in n or n == 'unit':
                        m.setdefault('unit', j)
                    if '재질' in n or '표면' in n or '비고' in n or 'maker' in n or '제조사' in n:
                        m.setdefault('maker', j)
                if 'item' in m and ('qty' in m or 'price' in m or 'amount' in m):
                    cols = m
                    header_i = i
                    break
            if header_i < 0:
                continue

            spec_idx = cols.get('spec_reg', cols.get('spec_draw'))

            def num(v):
                ss = re.sub(r'[^0-9.]', '', str(v or ''))
                try:
                    return float(ss) if ss else 0
                except Exception:
                    return 0

            def c(row, key):
                idx = cols.get(key)
                return row[idx].strip() if idx is not None and idx < len(row) else ''

            draw_vals = []
            for row in rows[header_i+1:]:
                if spec_idx is not None and spec_idx < len(row):
                    draw_vals.append(row[spec_idx].strip())
            numeric_specs = [v for v in draw_vals if v and re.fullmatch(r'\d+', v)]
            use_spec_idx = spec_idx
            if cols.get('spec_reg') is None and numeric_specs and len(numeric_specs) >= max(2, len([v for v in draw_vals if v])//2):
                use_spec_idx = None

            for row in rows[header_i+1:]:
                text = ' '.join(row)
                if any(k in text for k in ['소계','합계','이하여백','이 하 여 백','견적금액']):
                    continue
                item = c(row, 'item')
                if not item or item.lower() in ('model','modle','모델','품명'):
                    continue
                spec = row[use_spec_idx].strip() if use_spec_idx is not None and use_spec_idx < len(row) else ''
                maker = c(row, 'maker')
                qty = int(num(c(row, 'qty')) or 1)
                price = int(num(c(row, 'price')) or 0)
                amount = int(num(c(row, 'amount')) or 0)
                if price == 0 and amount and qty:
                    price = amount // qty
                if not amount and price and qty:
                    amount = price * qty
                # 가격처럼 큰 숫자가 규격으로 들어오면 제거
                if re.fullmatch(r'[0-9,]+', spec or '') and int(spec.replace(',','') or 0) > 9999:
                    spec = ''
                all_items.append(_normalize_item_spec_unit({
                    'item_name': item,
                    'spec': spec,
                    'unit_price': price,
                    'quantity': qty,
                    'unit': c(row,'unit') or 'EA',
                    'amount': amount,
                    'maker': maker,
                }))
        return {'items': all_items, 'vendor_name': vendor}
    except Exception as e:
        return {'items': [], 'vendor_name': '', 'error': str(e)}
    finally:
        if tmp_dir:
            shutil.rmtree(tmp_dir, ignore_errors=True)

def _actual_item_vendor_name(item: dict, idx: int = 0) -> str:
    return str(item.get('vendor_name') or item.get('vendor') or item.get('maker') or item.get('note') or '업체 미지정').strip() or '업체 미지정'


def _item_amount_value(item: dict) -> float:
    try:
        amount = float(str(item.get('amount') or 0).replace(',', ''))
        if amount:
            return amount
    except Exception:
        pass
    try:
        return float(str(item.get('unit_price') or 0).replace(',', '')) * float(str(item.get('quantity') or 0).replace(',', ''))
    except Exception:
        return 0.0


def _render_actual_item_table_pdf(pr: PurchaseRequest, items: list[dict]) -> str:
    req_dir = os.path.join(ATTACH_DIR, pr.request_no)
    os.makedirs(req_dir, exist_ok=True)
    pdf_path = os.path.join(req_dir, f'{pr.request_no}_실제입고품목표.pdf')
    font_name = _register_kor_font()
    c = canvas.Canvas(pdf_path, pagesize=A4)
    w, h = A4
    y = h - 18 * mm

    def draw(txt, x, yy, size=10):
        c.setFont(font_name, size)
        c.drawString(x, yy, str(txt))

    def fmt_money(v):
        try:
            return f"{int(round(float(v or 0))):,}"
        except Exception:
            return '0'

    draw('실제 입고 품목표', 18*mm, y, 16); y -= 10*mm
    draw(f'구매의뢰번호: {pr.request_no}', 18*mm, y, 10); y -= 6*mm
    draw(f'제목: {pr.title_full or ""}', 18*mm, y, 10); y -= 10*mm

    headers = ['No', '업체', '품명', '규격', '수량', '단위', '단가', '금액']
    col_x = [18, 30, 56, 96, 138, 153, 168, 190]
    row_h = 7*mm

    def new_page():
        nonlocal y
        c.showPage(); y = h - 18*mm
        draw('실제 입고 품목표', 18*mm, y, 16); y -= 10*mm
        draw_header()

    def draw_header():
        nonlocal y
        for i, hd in enumerate(headers):
            draw(hd, col_x[i]*mm, y, 8)
        y -= 4*mm
        c.line(18*mm, y, (w - 18*mm), y)
        y -= 4*mm

    draw_header()
    norm_items = [it for it in _normalize_items_spec_unit(items or []) if str(it.get('item_name') or it.get('spec') or '').strip()]
    groups = []
    cur = None
    for item in norm_items:
        vname = _actual_item_vendor_name(item)
        if not cur or cur['name'] != vname:
            cur = {'name': vname, 'items': [], 'total': 0.0}
            groups.append(cur)
        cur['items'].append(item)
        cur['total'] += _item_amount_value(item)

    row_no = 1
    for group in groups:
        for item in group['items']:
            if y < 22*mm:
                new_page()
            amount = _item_amount_value(item)
            vals = [
                row_no,
                group['name'],
                item.get('item_name', ''),
                item.get('spec', ''),
                item.get('quantity', ''),
                item.get('unit', 'EA') or 'EA',
                fmt_money(item.get('unit_price', 0)),
                fmt_money(amount),
            ]
            for i, val in enumerate(vals):
                draw(val, col_x[i]*mm, y, 8)
            row_no += 1
            y -= row_h
        if y < 22*mm:
            new_page()
        draw(f"합계금액({group['name']})", 30*mm, y, 9)
        draw(fmt_money(group['total']), 190*mm, y, 9)
        y -= row_h
    c.save()
    return pdf_path

def _effective_request_items(pr: PurchaseRequest):
    if int(pr.actual_received_diff or 0):
        try:
            items = json.loads(pr.actual_items_json or '[]') if pr.actual_items_json else []
            if items:
                return _normalize_items_spec_unit(items)
        except Exception:
            pass
    try:
        return _normalize_items_spec_unit(json.loads(pr.items_json or '[]') if pr.items_json else [])
    except Exception:
        return []




def _quote_request_items(pr: PurchaseRequest) -> list[dict]:
    """발주관리/비용처리 기준 품목: 견적서 자동인식/사용자 수정 품목표(items_json)."""
    if not pr:
        return []
    try:
        items = json.loads(pr.items_json or '[]') if pr.items_json else []
    except Exception:
        items = []
    if not items:
        items = [{
            'item_name': pr.item_name or '',
            'spec': pr.spec or '',
            'quantity': pr.quantity or 0,
            'unit': pr.unit or 'EA',
            'unit_price': pr.unit_price or 0,
        }]
    return _normalize_items_spec_unit([
        it for it in items
        if str((it or {}).get('item_name', '') or (it or {}).get('spec', '') or '').strip()
    ])


def _receipt_request_items(pr: PurchaseRequest) -> list[dict]:
    """입고관리/QR 기준 품목: 실제 입고품 다름이면 actual_items_json, 아니면 견적 품목."""
    if not pr:
        return []
    if int(getattr(pr, 'actual_received_diff', 0) or 0):
        try:
            items = json.loads(pr.actual_items_json or '[]') if pr.actual_items_json else []
            items = _normalize_items_spec_unit([
                it for it in items
                if str((it or {}).get('item_name', '') or (it or {}).get('spec', '') or '').strip()
            ])
            if items:
                return items
        except Exception:
            pass
    return _quote_request_items(pr)


def _vendors_for_request_items(pr: PurchaseRequest, items: list[dict], use_request_vendors: bool = True) -> list[dict]:
    vendors = []
    if use_request_vendors:
        try:
            vendors = json.loads(pr.vendors_json or '[]') if pr and pr.vendors_json else []
        except Exception:
            vendors = []
    return _merge_vendors_with_item_vendors(vendors, items or [])


def _order_type_value(order: PurchaseOrder) -> str:
    raw = str(getattr(order, 'order_type', '') or '').strip().lower()
    return raw or 'normal'


def _is_cost_order(order: PurchaseOrder) -> bool:
    return _order_type_value(order) in {'normal', 'cost', 'order', 'purchase'}


def _is_receipt_order(order: PurchaseOrder) -> bool:
    return _order_type_value(order) in {'normal', 'receipt', 'inbound'}


def _set_order_type(order: PurchaseOrder, typ: str) -> None:
    try:
        order.order_type = typ
    except Exception:
        pass


def _classify_order_types_for_request(db: Session, pr: PurchaseRequest) -> None:
    """구버전 DB 호환. 실제입고품 다름 건의 기존 무타입 발주 건은 입고관리용으로 분류한다."""
    if not pr or not getattr(pr, 'id', None):
        return
    try:
        orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id).all()
    except Exception:
        return
    if int(getattr(pr, 'actual_received_diff', 0) or 0):
        for po in orders:
            if not str(getattr(po, 'order_type', '') or '').strip() or _order_type_value(po) == 'normal':
                # v2.2.1 이전에는 실제입고품 다름에서 PurchaseOrder/ReceiptItem이 입고관리용으로 생성되었다.
                _set_order_type(po, 'receipt')
    else:
        for po in orders:
            if not str(getattr(po, 'order_type', '') or '').strip():
                _set_order_type(po, 'normal')


def _items_for_vendor_from(items: list[dict], vendors: list[dict], vendor, idx: int) -> list[dict]:
    matched = [it for it in (items or []) if _item_vendor_match(it, vendor, idx)]
    if not matched and len(vendors or []) <= 1:
        matched = items or []
    return matched


def _unique_typed_order_no_for_request(db: Session, pr: PurchaseRequest, idx: int, typ: str) -> str:
    base = pr.request_no or f"{datetime.now().strftime('%Y')}-{db.query(PurchaseOrder).count()+1:04d}"
    if typ == 'cost':
        candidates = [base if idx == 0 else f"{base}-{idx+1}", f"{base}-C{idx+1}"]
    elif typ == 'receipt':
        candidates = [f"{base}-R{idx+1}", f"{base}-I{idx+1}"]
    else:
        candidates = [base if idx == 0 else f"{base}-{idx+1}"]
    for candidate in candidates:
        if not db.query(PurchaseOrder).filter(PurchaseOrder.order_no == candidate).first():
            return candidate
    n = idx + 1
    while True:
        candidate = f"{base}-{typ[:1].upper()}{n+1}"
        if not db.query(PurchaseOrder).filter(PurchaseOrder.order_no == candidate).first():
            return candidate
        n += 1


def _find_order_for_vendor(orders: list[PurchaseOrder], vendor_name: str, desired_type: str = '') -> PurchaseOrder | None:
    vkey = _vendor_key(vendor_name)
    for existing in orders or []:
        if desired_type and _order_type_value(existing) != desired_type:
            continue
        ekey = _vendor_key(existing.vendor_name or '')
        if ekey and vkey and (ekey == vkey or ekey in vkey or vkey in ekey):
            return existing
    return None


def _ensure_order_shell(db: Session, pr: PurchaseRequest, vendor, idx: int, typ: str) -> PurchaseOrder:
    _classify_order_types_for_request(db, pr)
    orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id).order_by(PurchaseOrder.id.asc()).all()
    vname = _norm_vendor_name(vendor)
    existing = _find_order_for_vendor(orders, vname, typ)
    if existing:
        existing.vendor_name = vname or existing.vendor_name
        if isinstance(vendor, dict):
            existing.vendor_email = vendor.get('email', '') or existing.vendor_email
            existing.vendor_contact = vendor.get('contact', '') or existing.vendor_contact
        existing.delivery_date = pr.required_date
        existing.status = existing.status or '상신완료'
        _set_order_type(existing, typ)
        return existing
    po = PurchaseOrder(
        order_no=_unique_typed_order_no_for_request(db, pr, idx, typ),
        request_id=pr.id,
        vendor_name=vname,
        vendor_email=vendor.get('email', '') if isinstance(vendor, dict) else '',
        vendor_contact=vendor.get('contact', '') if isinstance(vendor, dict) else '',
        delivery_date=pr.required_date,
        status='상신완료',
    )
    _set_order_type(po, typ)
    db.add(po)
    db.flush()
    return po


def _ensure_items_for_order_no_qr(db: Session, po: PurchaseOrder, src_items: list[dict]) -> tuple[int, int]:
    """비용처리/발주서 전송용 품목. QR/입고처리에는 사용하지 않는다."""
    created = 0
    fixed = 0
    for src in _normalize_items_spec_unit(src_items or []):
        if not str(src.get('item_name', '') or src.get('spec', '') or '').strip():
            continue
        found = None
        for old in po.items:
            if _same_receipt_src_item(old, src):
                found = old
                break
        qty = _safe_int_value(src.get('quantity', 0))
        amount = _safe_float_value(src.get('amount') or src.get('total') or src.get('total_amount') or 0)
        unit_price = _safe_float_value(src.get('unit_price', 0))
        if amount and not unit_price:
            unit_price = amount / (qty or 1)
        if found is None:
            found = ReceiptItem(
                order_id=po.id,
                item_name=src.get('item_name', '') or '',
                spec=src.get('spec', '') or '',
                quantity=qty,
                unit=src.get('unit') or 'EA',
                unit_price=unit_price,
                maker=(src.get('maker') or src.get('vendor_name') or src.get('vendor') or src.get('note') or ''),
                note=(src.get('note') or src.get('maker') or src.get('vendor_name') or ''),
                material_code=_src_material_code(src),
                item_group=_src_item_group(src, po.purchase_request.title_full if po and po.purchase_request else ''),
                axis_type=src.get('axis') or src.get('axis_type') or '',
                order_round=src.get('order_round') or '',
                stage='발주관리용',
            )
            db.add(found)
            db.flush()
            created += 1
        else:
            fixed += 1
            found.quantity = found.quantity or qty
            found.unit = found.unit or (src.get('unit') or 'EA')
            found.unit_price = found.unit_price or unit_price
            # cost order는 QR을 생성하지 않는다.
            found.stage = found.stage or '발주관리용'
    return created, fixed


def _ensure_cost_orders_for_request(db: Session, pr: PurchaseRequest) -> dict:
    """실제입고품 다름 건에서 발주관리/비용처리용 발주 건을 견적 품목 기준으로 보정 생성."""
    if not pr:
        return {'created_orders': 0, 'created_items': 0}
    _classify_order_types_for_request(db, pr)
    quote_items = _quote_request_items(pr)
    if not quote_items:
        return {'created_orders': 0, 'created_items': 0}
    vendors = _vendors_for_request_items(pr, quote_items, True)
    if not vendors:
        vendors = [{'name': '미정', 'email': '', 'contact': ''}]
    before_orders = len(db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id, PurchaseOrder.order_type == 'cost').all())
    created_items = 0
    for idx, vendor in enumerate(vendors):
        vendor_items = _items_for_vendor_from(quote_items, vendors, vendor, idx)
        if not vendor_items and len(vendors) > 1:
            continue
        po = _ensure_order_shell(db, pr, vendor, idx, 'cost' if int(pr.actual_received_diff or 0) else 'normal')
        ci, _ = _ensure_items_for_order_no_qr(db, po, vendor_items or quote_items)
        created_items += ci
    after_orders = len(db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id, PurchaseOrder.order_type == ('cost' if int(pr.actual_received_diff or 0) else 'normal')).all())
    return {'created_orders': max(0, after_orders - before_orders), 'created_items': created_items}

def _build_actual_item_comment(pr: PurchaseRequest, items: list[dict]) -> str:
    names = ', '.join(str(i.get('item_name','')).strip() for i in (items or []) if str(i.get('item_name','')).strip())
    names = names[:120] + ('...' if len(names) > 120 else '')
    return f"실제 입고 품 다름 체크 건으로 실제 입고 품목표를 첨부합니다. QR 및 입고 기준 품목은 첨부 품목표 기준입니다. 품목: {names}"



class PurchaseRequestCreate(BaseModel):
    project_code: str
    project_year: Optional[str] = ""
    category: str
    sub_category: Optional[str] = ""
    item_type: Optional[str] = ""
    budget_type: Optional[str] = ""
    title_main: str
    title_full: Optional[str] = ""
    project_name: str
    item_name: str
    spec: Optional[str] = ""
    quantity: int
    unit: str = "EA"
    unit_price: float = 0
    reason: str
    requester: str
    department: str
    required_date: str
    purpose: Optional[str] = ""
    purpose_detail: Optional[str] = ""
    items: Optional[list] = []
    actual_received_diff: Optional[bool] = False
    actual_items: Optional[list] = []
    vendors: Optional[list] = []
    items_json: Optional[str] = ""
    vendors_json: Optional[str] = ""
    attach_files: Optional[list] = []
    approval_message: Optional[str] = ''
    bizbox_id: Optional[str] = None
    bizbox_pw: Optional[str] = None
    # server: 서버 PC에서 Selenium 실행 / client: 웹 접속 PC의 로컬 에이전트에서 실행
    automation_target: Optional[str] = 'server'

class QRTestCreate(BaseModel):
    text: str = 'DEVERP-QR-TEST'

class InspectionLike(BaseModel):
    order_no: str

class MarkSubmittedBody(BaseModel):
    custom_request_no: Optional[str] = None   # 사용자가 입력한 구매의뢰서 번호
    # 상신완료 직전에 화면에서 수정/수기 추가한 품목·업체를 DB에 다시 반영한다.
    # 기존에는 결재상신 후 품목을 수정/추가하면 currentRequestId가 이미 있어서
    # 서버의 items_json/vendors_json이 갱신되지 않아 QR/발주/입고현황 품목이 누락될 수 있었다.
    items: Optional[list] = None
    vendors: Optional[list] = None
    actual_received_diff: Optional[bool] = None
    actual_items: Optional[list] = None
    attach_files: Optional[list] = None
    approval_message: Optional[str] = ''


class VendorDirectorySave(BaseModel):
    vendors: List[dict] = Field(default_factory=list)


class MaterialBudgetSave(BaseModel):
    rows: List[dict] = Field(default_factory=list)


class DashboardDetailItemUpdate(BaseModel):
    id: Optional[int] = None
    item_name: Optional[str] = None
    spec: Optional[str] = None
    quantity: Optional[str] = None
    unit: Optional[str] = None


class DashboardDetailUpdate(BaseModel):
    request_no: Optional[str] = None
    vendor_name: Optional[str] = None
    title_full: Optional[str] = None
    project_name: Optional[str] = None
    delivery_date: Optional[str] = None
    requested_by: Optional[str] = None
    order_completed_by: Optional[str] = None
    inbound_by: Optional[str] = None
    outbound_by: Optional[str] = None
    items: List[DashboardDetailItemUpdate] = Field(default_factory=list)


def _parse_dashboard_date(value: Optional[str]):
    raw = str(value or '').strip()
    if not raw or raw.upper() in {'X', '-'}:
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
        try:
            return datetime.strptime(raw, fmt)
        except Exception:
            pass
    raise HTTPException(status_code=400, detail='입고요청일은 YYYY-MM-DD 형식으로 입력하세요.')


def _parse_qty_unit_text(value: Optional[str], default_unit: str = 'EA') -> tuple[int, str]:
    raw = str(value or '').strip()
    if not raw:
        return 0, default_unit or 'EA'
    import re as _re
    m = _re.match(r'\s*([-+]?\d+(?:\.\d+)?)\s*([^\d\s].*)?$', raw)
    if m:
        qty = int(float(m.group(1)))
        unit = (m.group(2) or default_unit or 'EA').strip()
        return qty, unit or 'EA'
    try:
        return int(float(raw)), default_unit or 'EA'
    except Exception:
        return 0, default_unit or 'EA'


def _retire_reusable_request_no_conflict(db: Session, conflict: PurchaseRequest, desired_no: str) -> tuple[bool, str]:
    """
    같은 구매의뢰서 번호가 과거 TEST/삭제/입고현황 삭제 건에 남아 있으면 재사용 가능하게 보관 처리한다.
    단, 해당 기존 건에 발주/입고 품목이 아직 살아 있으면 실제 중복 가능성이 있으므로 차단한다.
    """
    if not conflict:
        return True, ''
    try:
        orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == conflict.id).all()
        active_items = sum(len(getattr(o, 'items', []) or []) for o in orders)
        if orders or active_items:
            return False, f'이미 존재하는 구매의뢰서 번호입니다: {desired_no}'

        archived_no = f"{desired_no}__archived_{conflict.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        conflict.request_no = archived_no
        conflict.status = '삭제'
        conflict.bizbox_uploaded = 0
        db.flush()
        return True, f'기존 삭제/테스트 이력 번호를 보관 처리했습니다: {archived_no}'
    except Exception as e:
        return False, f'기존 번호 확인 중 오류가 발생했습니다: {e}'




def _copy_vendor_from_maker_if_missing(items: list[dict]) -> list[dict]:
    """수기 작성 품목에서 업체명이 비어 있으면 비고/제조사 값을 업체 후보로 보존한다."""
    out = []
    for it in items or []:
        if not isinstance(it, dict):
            continue
        row = dict(it)
        maker = str(row.get('maker') or row.get('note') or '').strip()
        if maker and not str(row.get('vendor_name') or row.get('vendor') or '').strip():
            row['vendor_name'] = maker
        out.append(row)
    return out

def _apply_mark_submitted_screen_payload(pr: PurchaseRequest, body: Optional[MarkSubmittedBody]) -> None:
    """상신완료 직전 화면의 최신 품목/업체/실입고 품목을 DB에 반영한다."""
    if not body:
        return
    if body.items is not None:
        items = _normalize_items_spec_unit(body.items or [])
        # 비고(제조사)/maker가 있으면 수기행도 업체 구분에 사용할 수 있도록 vendor_name 후보로 복사한다.
        items = _copy_vendor_from_maker_if_missing(items)
        pr.items_json = json.dumps(items, ensure_ascii=False)
        if items:
            first = items[0]
            pr.item_name = first.get('item_name') or pr.item_name
            pr.spec = first.get('spec') or pr.spec
            try:
                pr.quantity = int(float(first.get('quantity') or pr.quantity or 0))
            except Exception:
                pass
            pr.unit = first.get('unit') or pr.unit or 'EA'
            try:
                pr.unit_price = float(first.get('unit_price') or pr.unit_price or 0)
            except Exception:
                pass
    if body.vendors is not None:
        vendors = body.vendors or []
        pr.vendors_json = json.dumps(vendors, ensure_ascii=False)
    if body.actual_received_diff is not None:
        pr.actual_received_diff = 1 if body.actual_received_diff else 0
    if body.actual_items is not None:
        actual_items = _normalize_items_spec_unit(body.actual_items or [])
        actual_items = _copy_vendor_from_maker_if_missing(actual_items)
        pr.actual_items_json = json.dumps(actual_items, ensure_ascii=False)
    if body.attach_files is not None:
        pr.attach_files = json.dumps(body.attach_files or [], ensure_ascii=False)


def _normalize_items_vendors(data: PurchaseRequestCreate):
    items = data.items or []
    vendors = data.vendors or []
    if not items and data.items_json:
        try:
            items = json.loads(data.items_json)
        except Exception:
            items = []
    if not vendors and data.vendors_json:
        try:
            vendors = json.loads(data.vendors_json)
        except Exception:
            vendors = []
    return _copy_vendor_from_maker_if_missing(_normalize_items_spec_unit(items)), vendors


def _title_full(data: PurchaseRequestCreate) -> str:
    parts = [p for p in [data.project_code, data.category, data.sub_category, data.item_type] if p]
    prefix = ' '.join(f'[{p}]' for p in parts)
    return f"{prefix} {data.title_main} 구매의 건".strip()



def _estimate_item_quality_score(items: list[dict]) -> int:
    score = 0
    for it in items or []:
        if str(it.get('item_name') or '').strip():
            score += 3
        if str(it.get('spec') or '').strip():
            score += 2
        try:
            if float(it.get('quantity') or 0) > 0:
                score += 1
        except Exception:
            pass
        try:
            if float(it.get('unit_price') or 0) > 0:
                score += 1
        except Exception:
            pass
        try:
            if float(it.get('amount') or 0) > 0:
                score += 1
        except Exception:
            pass
    return score


async def _parse_estimate_files_impl(
    files: list[UploadFile],
):
    """웹 견적서 첨부/자동 인식. 여러 파일 선택 가능, 파일 1개=업체 1개로 반환."""
    os.makedirs(ATTACH_DIR, exist_ok=True)
    req_dir = os.path.join(ATTACH_DIR, '_web_estimate_uploads')
    os.makedirs(req_dir, exist_ok=True)
    estimates, saved_files, errors = [], [], []
    _estimate_parser_log(f"parse_multiple 시작: files={len(files or [])}, parser_version=v2.1.7")
    try:
        from client.estimate_parser import parse_estimate
    except Exception:
        parse_estimate = None
    for idx, up in enumerate(files or []):
        filename = up.filename or f'estimate_{idx+1}.bin'
        safe = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{_safe_filename(filename)}"
        path = os.path.join(req_dir, safe)
        with open(path, 'wb') as f:
            shutil.copyfileobj(up.file, f)
        saved_files.append(path)
        result = {'items': [], 'vendor_name': ''}
        err = ''
        if parse_estimate:
            try:
                _estimate_parser_log(f"기존 파서 시작: {filename}")
                result = parse_estimate(path) or {'items': [], 'vendor_name': ''}
                err = result.get('error') or ''
                _estimate_parser_log(f"기존 파서 완료: {filename}, items={len(result.get('items') or [])}, vendor={result.get('vendor_name') or ''}, err={err or ''}")
            except Exception as e:
                err = str(e)
                _estimate_parser_log(f"기존 파서 오류: {filename}, err={err}")
        # Excel은 보조 인식기로 규격/도번 보정
        ext = os.path.splitext(filename)[1].lower()
        if ext in ('.xlsx', '.xlsm', '.xls'):
            robust = _robust_xlsx_parse(path)
            if robust.get('items'):
                # 기존 인식 결과보다 보조 인식 결과가 더 완전하면 교체
                old_items = result.get('items') or []
                old_specs = sum(1 for it in old_items if str(it.get('spec','')).strip())
                new_specs = sum(1 for it in robust.get('items',[]) if str(it.get('spec','')).strip())
                if not old_items or len(robust['items']) >= len(old_items) or new_specs > old_specs:
                    result['items'] = robust['items']
                if robust.get('vendor_name') and not result.get('vendor_name'):
                    result['vendor_name'] = robust.get('vendor_name')
            if robust.get('error') and not err:
                err = robust.get('error')

        parser_score = _estimate_item_quality_score(result.get('items') or [])
        _estimate_parser_log(f"파서 결과: {filename}, items={len(result.get('items') or [])}, score={parser_score}, vendor={result.get('vendor_name') or ''}, err={err or ''}")

        vendor_name = result.get('vendor_name') or _estimate_vendor_from_file(filename)
        items = result.get('items') or []
        items = _normalize_items_spec_unit(items)
        items = _enrich_estimate_items_for_quality(items, f'{filename} {vendor_name}')
        for it in items:
            # 자동 인식 견적서도 비고(제조사) 칼럼에 업체명이 들어가게 한다.
            # 이후 상신완료 시 이 값으로 업체별 발주/QR/입고 품목을 분리한다.
            it['vendor_name'] = vendor_name
            it['vendor_index'] = idx
            if not str(it.get('material_code') or '').strip() and str(it.get('spec') or '').strip():
                it['material_code'] = str(it.get('spec') or '').strip()
            if not str(it.get('item_group') or '').strip():
                g = _src_item_group(it, f'{filename} {vendor_name}')
                if g:
                    it['item_group'] = g
            # 화면 품목표의 비고(제조사) 칸은 자동으로 채우지 않는다.
            # 업체 구분은 vendor_name/vendor_index로 처리한다.
            it['maker'] = str(it.get('maker') or '').strip()
            it['note'] = str(it.get('note') or '').strip()
            if 'amount' not in it:
                try:
                    it['amount'] = int(float(it.get('unit_price') or 0) * float(it.get('quantity') or 0))
                except Exception:
                    it['amount'] = 0
        est_payload = {
            'filename': filename,
            'saved_path': path,
            'vendor_name': vendor_name,
            'vendor_index': idx,
            'items': items,
            'error': err,
        }
        estimates.append(est_payload)
        if err:
            errors.append(f'{filename}: {err}')
    return {
        'success': True,
        'saved_files': saved_files,
        'estimates': estimates,
        'items': [it for est in estimates for it in est.get('items', [])],
        'vendors': [{'name': e['vendor_name'], 'vendor_name': e['vendor_name'], 'vendor_index': e['vendor_index']} for e in estimates],
        'errors': errors,
    }


@router.post('/estimate/parse')
async def parse_estimate_files(
    files: list[UploadFile] = File(...),
):
    return await _parse_estimate_files_impl(files)


@router.post('/estimate/parse_multiple')
async def parse_estimate_files_multiple(
    files: list[UploadFile] = File(...),
):
    """웹 견적서 자동인식 호환 API. 여러 견적서를 선택하면 1개 파일=1개 업체 기준으로 반환한다."""
    parsed = await _parse_estimate_files_impl(files)
    estimates = parsed.get('estimates') or []
    result_files = []
    for est in estimates:
        result_files.append({
            'filename': est.get('filename',''),
            'saved_path': est.get('saved_path',''),
            'vendor_name': est.get('vendor_name','') or est.get('vendor','') or '미정업체',
            'vendor_info': {'vendor_name': est.get('vendor_name','') or '미정업체'},
            'items': est.get('items') or [],
            'count': len(est.get('items') or []),
            'error': est.get('error',''),
        })
    ok_count = sum(len(f.get('items') or []) for f in result_files)
    msg = f'{ok_count}개 품목 인식 완료' if ok_count else '인식 가능한 품목이 없습니다.'
    return {
        'success': ok_count > 0,
        'files': result_files,
        'saved_files': parsed.get('saved_files') or [],
        'errors': parsed.get('errors') or [],
        'message': msg,
    }


def _public_api_url(path: str) -> str:
    """클라이언트 로컬 에이전트가 서버 파일을 내려받을 때 사용할 URL을 만든다."""
    path = str(path or '')
    if path.startswith('http://') or path.startswith('https://'):
        return path
    if not path.startswith('/'):
        path = '/' + path
    return f"{cfg.API_SERVER_BASE}{path}"


def _purchase_attachment_downloads(pr: PurchaseRequest) -> list:
    """구매의뢰 첨부파일을 브라우저/로컬 에이전트가 접근 가능한 다운로드 URL로 변환한다."""
    try:
        files = json.loads(pr.attach_files or '[]') if pr.attach_files else []
    except Exception:
        files = []
    out = []
    for idx, fp in enumerate(files):
        if fp and os.path.exists(str(fp)):
            rel = f'/api/purchase/request/{pr.id}/attachment_file/{idx}'
            out.append({
                'url': rel,
                'download_url': rel,
                'absolute_url': _public_api_url(rel),
                'filename': os.path.basename(str(fp)),
            })
    return out


def _purchase_request_data_for_bizbox(pr: PurchaseRequest, data: Optional[dict] = None) -> dict:
    """DB에 저장된 최신 구매의뢰 정보를 Bizbox 자동입력 payload로 재구성한다."""
    data = dict(data or {})
    try:
        data['request_no'] = pr.request_no
        data['project_code'] = pr.project_code or data.get('project_code', '')
        data['project_name'] = pr.project_name or data.get('project_name', '')
        data['title_full'] = pr.title_full or data.get('title_full', '')
        data['item_name'] = pr.item_name or data.get('item_name', '')
        data['spec'], data['unit'] = _clean_spec_unit_values(pr.spec or data.get('spec', ''), pr.unit or data.get('unit', 'EA'))
        data['quantity'] = pr.quantity or data.get('quantity', 0)
        data['unit_price'] = pr.unit_price or data.get('unit_price', 0)
        data['reason'] = pr.reason or data.get('reason', '')
        data['purpose'] = pr.reason or data.get('purpose', '') or data.get('reason', '')
        data['purpose_detail'] = pr.reason or data.get('purpose_detail', '') or data.get('reason', '')
        data['required_date'] = pr.required_date.strftime('%Y-%m-%d') if pr.required_date else data.get('required_date', '')
        data['items'] = _normalize_items_spec_unit(json.loads(pr.items_json or '[]') if pr.items_json else data.get('items', []))
        data['actual_items'] = _normalize_items_spec_unit(json.loads(pr.actual_items_json or '[]') if pr.actual_items_json else data.get('actual_items', []))
        data['vendors'] = json.loads(pr.vendors_json or '[]') if pr.vendors_json else data.get('vendors', [])
        data['attach_files'] = json.loads(pr.attach_files or '[]') if pr.attach_files else data.get('attach_files', [])
    except Exception:
        pass
    return data


def _build_purchase_bizbox_job(pr: PurchaseRequest, data: Optional[dict] = None) -> dict:
    """웹 접속 PC의 로컬 에이전트로 넘길 구매의뢰 자동입력 작업."""
    request_data = _purchase_request_data_for_bizbox(pr, data)
    return {
        'type': 'purchase_request',
        'request_id': pr.id,
        'request_no': pr.request_no,
        'server_base': cfg.API_SERVER_BASE,
        'request_data': request_data,
        'attachment_urls': _purchase_attachment_downloads(pr),
    }

@router.get('/request/{request_id}/bizbox_job')
def get_purchase_request_bizbox_job(request_id: int, db: Session = Depends(get_db)):
    """이미 등록된 구매의뢰의 클라이언트 자동상신 작업 정보를 다시 만든다.
    브라우저 타임아웃/중단 후 재시도할 때 사용한다.
    """
    pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == request_id).first()
    if not pr:
        raise HTTPException(404, '구매의뢰를 찾을 수 없습니다.')
    return _build_purchase_bizbox_job(pr)

@router.post('/request')
def create_purchase_request(data: PurchaseRequestCreate, background: BackgroundTasks, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    today = datetime.now().strftime('%Y%m%d')
    count = db.query(PurchaseRequest).count() + 1
    request_no = f'REQ-{today}-{count:04d}'
    items, vendors = _normalize_items_vendors(data)
    actual_items = _copy_vendor_from_maker_if_missing(_normalize_items_spec_unit(data.actual_items or []))
    base_spec, base_unit = _clean_spec_unit_values(data.spec, data.unit)
    pr = PurchaseRequest(
        request_no=request_no,
        project_code=data.project_code,
        project_year=data.project_year,
        category=data.category,
        sub_category=data.sub_category,
        item_type=data.item_type,
        budget_type=data.budget_type or data.item_type,
        title_full=data.title_full or _title_full(data),
        project_name=data.project_name,
        item_name=data.item_name,
        spec=base_spec,
        quantity=data.quantity,
        unit=base_unit,
        unit_price=data.unit_price,
        reason=(data.purpose_detail or data.purpose or data.reason),
        requester=data.requester,
        requested_by=current_user.name or current_user.username,
        department=data.department,
        required_date=datetime.strptime(data.required_date, '%Y-%m-%d'),
        items_json=json.dumps(items, ensure_ascii=False),
        vendors_json=json.dumps(vendors, ensure_ascii=False),
        actual_received_diff=1 if data.actual_received_diff else 0,
        actual_items_json=json.dumps(actual_items, ensure_ascii=False),
        attach_files=json.dumps(data.attach_files or [], ensure_ascii=False),
        status='작성완료',
    )
    db.add(pr)
    db.commit()
    db.refresh(pr)
    if data.actual_received_diff and actual_items:
        attach_files = json.loads(pr.attach_files or '[]') if pr.attach_files else []
        pdf_path = _render_actual_item_table_pdf(pr, actual_items)
        if pdf_path and pdf_path not in attach_files:
            attach_files.append(pdf_path)
            pr.attach_files = json.dumps(attach_files, ensure_ascii=False)
            db.commit()
    _sync_material_budget_master(db)
    target = (data.automation_target or 'server').strip().lower()
    bizbox_job = _build_purchase_bizbox_job(pr, data.dict()) if target == 'client' else None
    if target != 'client' and data.bizbox_id and data.bizbox_pw:
        background.add_task(_bizbox_task, pr.id, data.dict(), data.bizbox_id, data.bizbox_pw)
    resp = {'success': True, 'request_no': request_no, 'id': pr.id, 'automation_target': target}
    if bizbox_job:
        resp['bizbox_job'] = bizbox_job
    return resp


@router.post('/request/{request_id}/attachments')
async def upload_request_attachments(request_id: int, files: list[UploadFile] = File(...), db: Session = Depends(get_db)):
    pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == request_id).first()
    if not pr:
        raise HTTPException(status_code=404, detail='구매의뢰를 찾을 수 없습니다.')
    saved = json.loads(pr.attach_files or '[]') if pr.attach_files else []
    req_dir = os.path.join(ATTACH_DIR, pr.request_no)
    os.makedirs(req_dir, exist_ok=True)
    for up in files:
        filename = up.filename or 'attachment.bin'
        safe_name = f"{datetime.now().strftime('%H%M%S%f')}_{os.path.basename(filename)}"
        path = os.path.join(req_dir, safe_name)
        with open(path, 'wb') as f:
            shutil.copyfileobj(up.file, f)
        saved.append(path)
    pr.attach_files = json.dumps(saved, ensure_ascii=False)
    db.commit()
    return {'success': True, 'saved_files': saved}


def _bizbox_task(pr_id, data, uid, upw):
    from server.bizbox_selenium import auto_upload_purchase_request
    db = next(get_db())
    try:
        pr_obj = db.query(PurchaseRequest).filter(PurchaseRequest.id == pr_id).first()

        # DB에 저장된 최신 구매의뢰 데이터를 Bizbox 자동입력에 다시 반영
        # WEB/EXE 모두 제목·구매목적·상세내용·품목·추천업체·첨부가 누락되지 않게 한다.
        if pr_obj:
            try:
                data['request_no'] = pr_obj.request_no
                data['project_code'] = pr_obj.project_code or data.get('project_code', '')
                data['project_name'] = pr_obj.project_name or data.get('project_name', '')
                data['title_full'] = pr_obj.title_full or data.get('title_full', '')
                data['item_name'] = pr_obj.item_name or data.get('item_name', '')
                data['spec'] = pr_obj.spec or data.get('spec', '')
                data['quantity'] = pr_obj.quantity or data.get('quantity', 0)
                data['unit'] = pr_obj.unit or data.get('unit', 'EA')
                data['unit_price'] = pr_obj.unit_price or data.get('unit_price', 0)
                data['reason'] = pr_obj.reason or data.get('reason', '')
                data['required_date'] = pr_obj.required_date.strftime('%Y-%m-%d') if pr_obj.required_date else data.get('required_date', '')
                data['items'] = json.loads(pr_obj.items_json or '[]') if pr_obj.items_json else data.get('items', [])
                data['vendors'] = json.loads(pr_obj.vendors_json or '[]') if pr_obj.vendors_json else data.get('vendors', [])
                data['attach_files'] = json.loads(pr_obj.attach_files or '[]') if pr_obj.attach_files else data.get('attach_files', [])
            except Exception:
                pass

        result = auto_upload_purchase_request(data, uid, upw)
        pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == pr_id).first()
        if pr and result.get('success'):
            pr.bizbox_uploaded = 1
            pr.bizbox_no = result.get('bizbox_no', '')
            pr.status = '결재중'
            db.commit()
    finally:
        db.close()

@router.post('/request/{request_id}/mark_submitted')
def mark_request_submitted(request_id: int, body: Optional[MarkSubmittedBody] = None, db: Session = Depends(get_db)):
    """
    상신완료 시 발주/입고품목/QR을 반드시 보정 생성한다.

    기존 문제:
      - 이미 PurchaseOrder가 있으면 QR/품목 누락 여부를 확인하지 않고 바로 return
      - 업체가 여러 개일 때 vendor_index 매칭이 어긋나면 모든 업체가 skip되어 0건/0개 가능
      - QR 생성 실패를 조용히 pass해서 화면에는 성공처럼 보이나 QR 0개 가능

    보정:
      - 기존 발주가 있어도 품목/QR 누락분 생성
      - 업체-품목 매칭 실패 시 첫 업체에 전체 품목 fallback
      - 최종 total_orders / total_items / total_qr를 함께 반환
    """
    pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == request_id).first()
    if not pr:
        return {'success': False, 'message': '구매의뢰를 찾을 수 없습니다.'}

    actual_comment = ''

    if body and body.custom_request_no:
        custom_no = body.custom_request_no.strip()
        conflict = db.query(PurchaseRequest).filter(
            PurchaseRequest.request_no == custom_no,
            PurchaseRequest.id != pr.id
        ).first()
        if conflict:
            ok, msg = _retire_reusable_request_no_conflict(db, conflict, custom_no)
            if not ok:
                return {'success': False, 'message': msg}
        pr.request_no = custom_no
        db.flush()

    # 결재상신 이후 화면에서 수기 추가/수정한 품목과 업체를 상신완료 직전에 다시 저장한다.
    _apply_mark_submitted_screen_payload(pr, body)

    # 기준 분리:
    # - 발주관리/발주서전송/비용처리: 견적서 자동인식·사용자 수정 품목표(items_json)
    # - 입고관리/QR: 실제 입고품 다름 체크 시 실제 입고 품목표(actual_items_json)
    quote_items = _quote_request_items(pr)
    receipt_items = _receipt_request_items(pr)
    if not quote_items:
        return {'success': False, 'message': '상신완료 처리할 견적/비용처리 품목이 없습니다.'}
    if not receipt_items:
        return {'success': False, 'message': '상신완료 처리할 입고관리 품목이 없습니다.'}

    quote_vendors = _vendors_for_request_items(pr, quote_items, True)
    receipt_vendors = _vendors_for_request_items(pr, receipt_items, False if int(pr.actual_received_diff or 0) else True)
    if not quote_vendors:
        quote_vendors = [{'name': '미정', 'email': '', 'contact': ''}]
    if not receipt_vendors:
        receipt_vendors = [{'name': '미정', 'email': '', 'contact': ''}]

    # 기존 변수명 호환: 아래 QR/입고 보정 로직은 items/vendors를 입고관리 기준으로 사용한다.
    items = receipt_items
    vendors = receipt_vendors

    # 실제입고품 다름 PDF/코멘트 유지: 반드시 실제 입고품 기준
    if int(pr.actual_received_diff or 0) and receipt_items:
        try:
            pdf_path = _render_actual_item_table_pdf(pr, receipt_items)
            attach_files = json.loads(pr.attach_files or '[]') if pr.attach_files else []
            if pdf_path and pdf_path not in attach_files:
                attach_files.append(pdf_path)
                pr.attach_files = json.dumps(attach_files, ensure_ascii=False)
            actual_comment = _build_actual_item_comment(pr, receipt_items)
        except Exception:
            actual_comment = ''

    _classify_order_types_for_request(db, pr)
    # 실제입고품 다름인 경우 발주관리/비용처리용 발주 건을 원 견적 품목 기준으로 별도 보정한다.
    try:
        _ensure_cost_orders_for_request(db, pr)
    except Exception as e:
        errors = [f'발주관리 비용처리용 품목 보정 실패 - {e}']
    else:
        errors = []

    created_orders = 0
    created_items = 0
    created_qr = 0
    fixed_existing_items = 0
    # errors는 위 비용처리/입고관리 보정 단계에서 누적한다.

    def _safe_int(v):
        try:
            return int(float(v or 0))
        except Exception:
            return 0

    def _safe_float(v):
        try:
            return float(v or 0)
        except Exception:
            return 0.0

    def _same_item(db_item, src):
        return (
            str(db_item.item_name or '').strip() == str(src.get('item_name', '') or '').strip()
            and str(db_item.spec or '').strip() == str(src.get('spec', '') or '').strip()
        )

    def _order_no_for(idx):
        base = pr.request_no or f"{datetime.now().strftime('%Y')}-{db.query(PurchaseOrder).count()+1:04d}"
        order_no = base if idx == 0 else f"{base}-{idx+1}"
        if db.query(PurchaseOrder).filter(PurchaseOrder.order_no == order_no).first():
            order_no = f"{base}-{idx+1}-{datetime.now().strftime('%H%M%S')}"
        return order_no

    def _items_for_vendor(vendor, idx):
        matched = [it for it in items if _item_vendor_match(it, vendor, idx)]
        # 업체가 1개면 전체 품목이 그 업체 품목
        if not matched and len(vendors) <= 1:
            matched = items
        return matched

    def _create_order(vendor, idx, order_items):
        nonlocal created_orders
        po = PurchaseOrder(
            order_no=_order_no_for(idx),
            request_id=pr.id,
            vendor_name=_norm_vendor_name(vendor),
            vendor_email=vendor.get('email', '') if isinstance(vendor, dict) else '',
            vendor_contact=vendor.get('contact', '') if isinstance(vendor, dict) else '',
            delivery_date=pr.required_date,
            status='상신완료',
        )
        _set_order_type(po, 'receipt' if int(pr.actual_received_diff or 0) else 'normal')
        db.add(po)
        db.flush()
        created_orders += 1
        _ensure_items_and_qr(po, order_items)
        return po

    def _ensure_items_and_qr(po, src_items):
        nonlocal created_items, fixed_existing_items, created_qr
        if not src_items:
            src_items = items

        for src in _normalize_items_spec_unit(src_items):
            ri = None
            for old in po.items:
                fixed_old_spec, fixed_old_unit = _clean_spec_unit_values(old.spec or '', old.unit or 'EA')
                if old.spec != fixed_old_spec or old.unit != fixed_old_unit:
                    old.spec = fixed_old_spec
                    old.unit = fixed_old_unit
                if _same_item(old, src):
                    ri = old
                    break

            if ri is None:
                src_qty = _safe_int(src.get('quantity', 0))
                src_amount = _safe_float(src.get('amount') or src.get('total') or src.get('total_amount') or 0)
                src_unit_price = _safe_float(src.get('unit_price', 0))
                if src_amount and not src_unit_price:
                    src_unit_price = src_amount / (src_qty or 1)
                ri = ReceiptItem(
                    order_id=po.id,
                    item_name=src.get('item_name', '') or '',
                    spec=src.get('spec', '') or '',
                    quantity=src_qty,
                    unit=src.get('unit') or 'EA',
                    unit_price=src_unit_price,
                    maker=(src.get('maker') or src.get('vendor_name') or src.get('vendor') or src.get('note') or ''),
                    note=(src.get('note') or src.get('maker') or ''),
                    material_code=_src_material_code(src),
                    item_group=_src_item_group(src, pr.title_full if pr else ''),
                    axis_type=src.get('axis') or src.get('axis_type') or '',
                    order_round=src.get('order_round') or '',
                    stage='미입고',
                )
                db.add(ri)
                db.flush()
                created_items += 1
            else:
                fixed_existing_items += 1
                # 수량/단가 등 빈 값 보정
                if not ri.quantity:
                    ri.quantity = _safe_int(src.get('quantity', 0))
                # 기존 데이터에 규격/도번이 단위 칼럼에 들어간 경우도 즉시 보정
                fixed_spec, fixed_unit = _clean_spec_unit_values(ri.spec or src.get('spec', ''), ri.unit or src.get('unit', 'EA'))
                ri.spec = fixed_spec
                ri.unit = fixed_unit
                if not ri.unit_price:
                    src_amount = _safe_float(src.get('amount') or src.get('total') or src.get('total_amount') or 0)
                    src_unit_price = _safe_float(src.get('unit_price', 0))
                    if src_amount and not src_unit_price:
                        src_unit_price = src_amount / (ri.quantity or _safe_int(src.get('quantity', 0)) or 1)
                    ri.unit_price = src_unit_price

            # 수기 작성/자동스캔으로 입력된 비고(제조사) 및 보조 정보를 발주관리/입고현황에 계속 유지한다.
            maker_val = src.get('maker') or src.get('vendor_name') or src.get('vendor') or src.get('note') or ''
            if maker_val and not getattr(ri, 'maker', ''):
                ri.maker = maker_val
            note_val = src.get('note') or src.get('maker') or ''
            if note_val and not getattr(ri, 'note', ''):
                ri.note = note_val
            for attr, keys in {
                'material_code': ['material_code', 'code', 'spec'],
                'item_group': ['item_group', 'group'],
                'axis_type': ['axis', 'axis_type'],
                'order_round': ['order_round'],
            }.items():
                if not getattr(ri, attr, ''):
                    for key in keys:
                        val = src.get(key)
                        if val:
                            setattr(ri, attr, val)
                            break

            if not getattr(ri, 'material_code', '') and getattr(ri, 'spec', ''):
                ri.material_code = ri.spec
            if not getattr(ri, 'item_group', ''):
                g = _src_item_group(src, pr.title_full if pr else '')
                if g:
                    ri.item_group = g
            try:
                if not ri.qr_code or not ri.qr_code_path or not os.path.exists(str(ri.qr_code_path)):
                    generate_qr_for_item(ri, po, db)
                    created_qr += 1
            except Exception as e:
                errors.append(f"{po.order_no}/{ri.item_name}: QR 생성 실패 - {e}")

    existing_orders = [o for o in db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id).all() if _is_receipt_order(o)]

    if existing_orders:
        # 기존 발주가 있더라도 품목/QR 누락분을 반드시 보정한다.
        # 또한 견적서 자동인식 후 추천업체 표/DB에 일부 업체가 빠져 있었던 기존 건도
        # vendors+items 기준으로 누락 업체 발주 건을 추가 생성한다.
        represented_vendor_keys = set()
        for idx, po in enumerate(existing_orders):
            po.status = '상신완료'
            vendor = None
            matched_idx = idx
            for vi, v in enumerate(vendors):
                if _vendor_key(_norm_vendor_name(v)) and _vendor_key(_norm_vendor_name(v)) == _vendor_key(po.vendor_name or ''):
                    vendor = v
                    matched_idx = vi
                    break
            if vendor is None:
                vendor = vendors[min(idx, len(vendors)-1)] if vendors else {'name': po.vendor_name or '미정'}
            represented_vendor_keys.add(_vendor_key(po.vendor_name or _norm_vendor_name(vendor)))
            src = _items_for_vendor(vendor, matched_idx)
            # 업체가 1개뿐인 기존 데이터에서만 전체 품목 fallback. 다중 업체에서는 잘못된 업체에 전체 품목이 들어가지 않게 한다.
            if not src and len(vendors) <= 1:
                src = items
            if src:
                _ensure_items_and_qr(po, src)

        # 기존 발주 생성 후 새로 인식/추가된 업체가 있으면 발주/입고 품목을 추가 생성
        for idx, vendor in enumerate(vendors):
            vkey = _vendor_key(_norm_vendor_name(vendor))
            if not vkey or vkey in represented_vendor_keys:
                continue
            vendor_items = _items_for_vendor(vendor, idx)
            if not vendor_items:
                continue
            _create_order(vendor, idx, vendor_items)
            represented_vendor_keys.add(vkey)
    else:
        created_any = False

        # 1차: 업체별 매칭 품목으로 생성
        for idx, vendor in enumerate(vendors):
            vendor_items = _items_for_vendor(vendor, idx)
            if not vendor_items and len(vendors) > 1:
                continue
            _create_order(vendor, idx, vendor_items or items)
            created_any = True

        # 2차 fallback: 매칭 실패로 모두 skip된 경우 첫 업체 + 전체 품목으로 강제 생성
        if not created_any:
            vendor = vendors[0] if vendors else {'name': '미정', 'email': '', 'contact': ''}
            _create_order(vendor, 0, items)

    # 안전 보정: 발주 건은 있는데 입고현황 품목이 하나도 없으면 전체 품목을 첫 발주 건에 생성한다.
    # 결재상신/상신완료 도중 브라우저 오류 또는 업체 매칭 오류가 나도 입고현황이 빈 화면이 되지 않게 한다.
    try:
        final_orders_pre = [o for o in db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id).order_by(PurchaseOrder.id.asc()).all() if _is_receipt_order(o)]
        if final_orders_pre and sum(len(getattr(o, 'items', []) or []) for o in final_orders_pre) <= 0 and items:
            _ensure_items_and_qr(final_orders_pre[0], items)
    except Exception as e:
        errors.append(f'입고현황 품목 보정 실패 - {e}')

    pr.status = '상신완료'
    pr.bizbox_uploaded = 1
    db.commit()
    _sync_material_budget_master(db)

    all_orders = [o for o in db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id).all() if _is_receipt_order(o)]
    total_items = sum(len(o.items) for o in all_orders)
    total_qr = sum(1 for o in all_orders for i in o.items if i.qr_code)
    total_orders = len(all_orders)

    return {
        'success': True,
        'created_orders': created_orders,
        'created_items': created_items,
        'created_qr': created_qr,
        'fixed_existing_items': fixed_existing_items,
        'total_orders': total_orders,
        'total_items': total_items,
        'total_qr': total_qr,
        'request_no': pr.request_no,
        'actual_item_comment': actual_comment,
        'errors': errors[:20],
        'display_orders': 1 if total_orders > 0 else 0,
        'vendor_order_count': total_orders,
        'message': f'상신완료 처리되었습니다. 발주관리/입고현황 표시 1건 / 업체 {total_orders}개 / 품목 {total_items}개 / QR {total_qr}개 준비 완료',
    }

@router.get('/requests')
def get_requests(status: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(PurchaseRequest)
    if status:
        q = q.filter(PurchaseRequest.status == status)
    items = q.order_by(PurchaseRequest.request_date.desc()).all()
    return [{
        'id': i.id,
        'request_no': i.request_no,
        'project_code': i.project_code or '',
        'title_full': i.title_full or '',
        'project_name': i.project_name,
        'item_name': i.item_name,
        'quantity': i.quantity,
        'requester': i.requester,
        'required_date': i.required_date.isoformat() if i.required_date else '',
        'status': i.status,
    } for i in items]


@router.get('/request/{request_id}')
def get_request_detail(request_id: int, db: Session = Depends(get_db)):
    pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == request_id).first()
    if not pr:
        raise HTTPException(status_code=404, detail='구매의뢰를 찾을 수 없습니다.')
    return {
        'id': pr.id,
        'request_no': pr.request_no,
        'project_code': pr.project_code or '',
        'title_full': pr.title_full or '',
        'project_name': pr.project_name or '',
        'attach_files': json.loads(pr.attach_files or '[]') if pr.attach_files else [],
        'items': json.loads(pr.items_json or '[]') if pr.items_json else [],
        'actual_received_diff': bool(pr.actual_received_diff),
        'actual_items': json.loads(pr.actual_items_json or '[]') if pr.actual_items_json else [],
        'vendors': json.loads(pr.vendors_json or '[]') if pr.vendors_json else [],
        'status': pr.status,
    }



@router.get('/request/{request_id}/attachment_file/{file_index}')
def get_request_attachment_file(request_id: int, file_index: int, db: Session = Depends(get_db)):
    """로컬 클라이언트 자동화 에이전트가 구매의뢰 첨부파일을 내려받기 위한 내부 API."""
    pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == request_id).first()
    if not pr:
        raise HTTPException(status_code=404, detail='구매의뢰를 찾을 수 없습니다.')
    try:
        files = json.loads(pr.attach_files or '[]') if pr.attach_files else []
    except Exception:
        files = []
    if file_index < 0 or file_index >= len(files):
        raise HTTPException(status_code=404, detail='첨부파일 번호가 올바르지 않습니다.')
    path = str(files[file_index] or '')
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail='첨부파일을 찾을 수 없습니다.')
    return FileResponse(path, filename=os.path.basename(path))





def _receipt_item_is_purchase_received(item: ReceiptItem) -> bool:
    """대시보드 표시용 입고 완료 판단: 구매팀 입고 이상이면 입고 처리된 품목으로 본다."""
    if not item:
        return False
    if getattr(item, 'purchase_recv_at', None) or getattr(item, 'purchase_recv_by', None):
        return True
    if getattr(item, 'quality_recv_at', None) or getattr(item, 'quality_recv_by', None):
        return True
    if getattr(item, 'manufacture_recv_at', None) or getattr(item, 'manufacture_recv_by', None):
        return True
    stage = str(getattr(item, 'stage', '') or '').strip()
    return stage in {'구매팀입고', '품질검수', '생산팀입고', '완료', '입고완료'}


def _order_group_receipt_counts(orders: list[PurchaseOrder]) -> tuple[int, int, bool]:
    items = [it for o in (orders or []) for it in (getattr(o, 'items', None) or [])]
    total = len(items)
    received = sum(1 for it in items if _receipt_item_is_purchase_received(it))
    return total, received, bool(total and received >= total)

def _order_group_key(order: PurchaseOrder) -> str:
    # 발주관리/입고현황 표시 기준: 구매의뢰서 1건 = 화면 1건.
    if getattr(order, 'request_id', None):
        return f"request:{order.request_id}"
    return f"order:{order.id}"


def _group_order_status(orders: list[PurchaseOrder]) -> str:
    statuses = [str(o.status or '').strip() for o in orders if str(o.status or '').strip()]
    if not statuses:
        return ''
    seen = []
    for s in statuses:
        if s not in seen:
            seen.append(s)
    if len(seen) == 1:
        return seen[0]
    return ' / '.join(seen)


def _vendor_display_name(vendor_names: list[str]) -> str:
    names = []
    for name in vendor_names:
        n = str(name or '').strip() or '미정업체'
        if n not in names:
            names.append(n)
    if not names:
        return ''
    if len(names) == 1:
        return names[0]
    return f"{names[0]} 외 {len(names)-1}개"



def _receipt_orders_for_request(db: Session, pr: PurchaseRequest | None, fallback: list[PurchaseOrder] | None = None) -> list[PurchaseOrder]:
    if pr and getattr(pr, 'id', None):
        try:
            _classify_order_types_for_request(db, pr)
            orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == pr.id).order_by(PurchaseOrder.id.asc()).all()
            return [o for o in orders if _is_receipt_order(o)]
        except Exception:
            pass
    return [o for o in (fallback or []) if _is_receipt_order(o)]


def _qr_orders_for_order(db: Session, order: PurchaseOrder) -> list[PurchaseOrder]:
    """QR/입고용 조회는 항상 receipt/normal 입고관리용 발주 건만 사용한다.
    실제 입고품 다름 건에서 cost 발주 건을 클릭해도 QR은 실제 입고 품목표 기준만 보여준다.
    """
    if not order:
        return []
    try:
        if getattr(order, 'request_id', None):
            pr = getattr(order, 'purchase_request', None)
            if pr:
                _classify_order_types_for_request(db, pr)
            related = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == order.request_id).order_by(PurchaseOrder.id.asc()).all()
            receipt_orders = [o for o in related if _is_receipt_order(o)]
            if receipt_orders:
                return receipt_orders
            # 구버전/비정상 데이터 fallback: 실제입고품 다름이 아니면 normal/cost 겸용 발주를 사용한다.
            if not (pr and int(getattr(pr, 'actual_received_diff', 0) or 0)):
                return related
            return []
    except Exception:
        pass
    return [order] if _is_receipt_order(order) else []

def _orders_for_display_groups(orders: list[PurchaseOrder]) -> list[list[PurchaseOrder]]:
    # DB에는 업체별 PurchaseOrder를 유지하되, 화면/API에는 구매의뢰서 기준 1건으로 묶는다.
    # 발주서전송은 기존 업체별 구성 그대로 유지하면서 발주관리/입고현황 중복 표시를 제거한다.
    groups = []
    seen = {}
    for order in orders:
        key = _order_group_key(order)
        if key not in seen:
            seen[key] = []
            groups.append(seen[key])
        seen[key].append(order)
    for group in groups:
        group.sort(key=lambda x: x.id or 0)
    return groups

@router.get('/orders')
def get_orders(status: Optional[str] = None, db: Session = Depends(get_db)):
    _repair_missing_vendor_orders_all(db)
    try:
        for _pr in db.query(PurchaseRequest).filter(PurchaseRequest.status.in_(['상신완료', '결재중', '발주서전송완료'])).all():
            _ensure_cost_orders_for_request(db, _pr)
        db.commit()
    except Exception:
        db.rollback()
    q = db.query(PurchaseOrder)
    if status:
        q = q.filter(PurchaseOrder.status == status)
    orders = [o for o in q.order_by(PurchaseOrder.order_date.desc(), PurchaseOrder.id.desc()).all() if _is_cost_order(o)]
    result = []
    for group in _orders_for_display_groups(orders):
        base = group[0]
        pr = base.purchase_request
        vendor_names = [o.vendor_name for o in group]
        all_items = [it for o in group for it in (o.items or [])]
        receipt_group = _receipt_orders_for_request(db, pr, group)
        total_items_for_receipt, received_items_for_receipt, all_items_received = _order_group_receipt_counts(receipt_group)

        result.append({
            'id': base.id,
            'order_ids': [o.id for o in group],
            'order_no': pr.request_no if pr and pr.request_no else base.order_no,
            'raw_order_no': base.order_no,
            'request_id': pr.id if pr else None,
            'request_no': pr.request_no if pr else '',
            'vendor_name': _vendor_display_name(vendor_names),
            'vendor_names': [str(v or '').strip() or '미정업체' for v in vendor_names],
            'vendor_count': len(group),
            'vendor_email': base.vendor_email,
            'delivery_date': base.delivery_date.isoformat() if base.delivery_date else '',
            'status': _group_order_status(group),
            'tax_docs_completed_by': _join_people_names([getattr(o, 'tax_docs_completed_by', '') for o in group]),
            'tax_docs_completed_at': max([getattr(o, 'tax_docs_completed_at', None) for o in group if getattr(o, 'tax_docs_completed_at', None)] or [None]).strftime('%Y-%m-%d %H:%M') if any(getattr(o, 'tax_docs_completed_at', None) for o in group) else '',
            'email_sent': 1 if any(o.email_sent for o in group) else 0,
            'item_count': len(all_items),
            'qr_count': sum(1 for o in receipt_group for it in (o.items or []) if it.qr_code),
        })
    return result

@router.get('/orders/detail')
def get_orders_detail(db: Session = Depends(get_db)):
    """발주관리/대시보드 공통 상세 목록.

    기존 DB를 새 빌드본에 덮어쓴 경우 일부 JSON/신규 컬럼/레코드가 맞지 않아도
    화면 전체가 빈칸이 되지 않도록 방어적으로 반환한다.
    """
    try:
        _repair_missing_vendor_orders_all(db)
        for _pr in db.query(PurchaseRequest).filter(PurchaseRequest.status.in_(['상신완료', '결재중', '발주서전송완료'])).all():
            _ensure_cost_orders_for_request(db, _pr)
    except Exception as e:
        # 보정 실패가 목록 로딩 실패로 이어지지 않도록 한다.
        try:
            db.rollback()
        except Exception:
            pass
        print(f"[WARN] orders/detail repair skipped: {e}")

    try:
        orders = [o for o in db.query(PurchaseOrder).order_by(PurchaseOrder.order_date.desc(), PurchaseOrder.id.desc()).all() if _is_cost_order(o)]
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        print(f"[ERROR] orders/detail purchase_orders query failed: {e}")
        return []

    result = []
    for group in _orders_for_display_groups(orders):
        try:
            base = group[0]
            pr = base.purchase_request
            vendor_names = [str(o.vendor_name or '').strip() or '미정업체' for o in group]
            vendor_groups = []
            all_items = []
            for o in group:
                group_items = []
                for i in (o.items or []):
                    norm_item = _normalize_item_spec_unit({'spec': i.spec, 'unit': i.unit})
                    if i.spec != norm_item.get('spec') or i.unit != 'EA':
                        i.spec = norm_item.get('spec')
                        i.unit = 'EA'
                    item_row = {
                        'id': i.id,
                        'order_id': o.id,
                        'order_no': o.order_no,
                        'vendor_name': o.vendor_name or '',
                        'item_name': i.item_name,
                        'spec': i.spec,
                        'quantity': i.quantity,
                        'unit': 'EA',
                        'unit_price': i.unit_price,
                        'maker': getattr(i, 'maker', '') or '',
                        'note': getattr(i, 'note', '') or '',
                        'material_code': getattr(i, 'material_code', '') or '',
                        'item_group': getattr(i, 'item_group', '') or '',
                        'axis_type': getattr(i, 'axis_type', '') or '',
                        'order_round': getattr(i, 'order_round', '') or '',
                        'qr_code': i.qr_code,
                        'qr_code_path': i.qr_code_path,
                        'purchase_recv_by': getattr(i, 'purchase_recv_by', '') or '',
                        'quality_recv_by': getattr(i, 'quality_recv_by', '') or '',
                    }
                    group_items.append(item_row)
                    all_items.append(item_row)
                vendor_groups.append({
                    'order_id': o.id,
                    'order_no': o.order_no,
                    'vendor_name': o.vendor_name or '',
                    'vendor_email': o.vendor_email or '',
                    'delivery_date': _date_to_text_purchase(o.delivery_date),
                    'status': o.status or '',
                    'tax_docs_completed_by': getattr(o, 'tax_docs_completed_by', '') or '',
                    'tax_docs_completed_at': _dt_to_text_purchase(getattr(o, 'tax_docs_completed_at', None)),
                    'item_count': len(group_items),
                    'qr_count': sum(1 for it in group_items if it.get('qr_code')),
                    'items': group_items,
                })
            receipt_group = _receipt_orders_for_request(db, pr, group)
            total_items_for_receipt, received_items_for_receipt, all_items_received = _order_group_receipt_counts(receipt_group)
            result.append({
                'id': base.id,
                'order_ids': [o.id for o in group],
                'request_id': pr.id if pr else None,
                'request_no': pr.request_no if pr else '',
                'requested_by': (getattr(pr, 'requested_by', '') or pr.requester) if pr else '',
                'requester': pr.requester if pr else '',
                'order_completed_by': _join_people_names([getattr(o, 'order_completed_by', '') for o in group]),
                'order_completed_at': _dt_to_text_purchase(max([getattr(o, 'order_completed_at', None) for o in group if getattr(o, 'order_completed_at', None)] or [None])),
                'tax_docs_completed_by': _join_people_names([getattr(o, 'tax_docs_completed_by', '') for o in group]),
                'tax_docs_completed_at': _dt_to_text_purchase(max([getattr(o, 'tax_docs_completed_at', None) for o in group if getattr(o, 'tax_docs_completed_at', None)] or [None])),
                'order_no': pr.request_no if pr and pr.request_no else base.order_no,
                'raw_order_no': base.order_no,
                'project_code': pr.project_code if pr else '',
                'project_name': pr.project_name if pr else '',
                'title_full': pr.title_full if pr else '',
                'inbound_by': _join_people_names([getattr(i, 'purchase_recv_by', '') for o in group for i in (o.items or [])]),
                'outbound_by': _join_people_names([getattr(i, 'quality_recv_by', '') for o in group for i in (o.items or [])]),
                'vendor_name': _vendor_display_name(vendor_names),
                'vendor_names': vendor_names,
                'vendor_count': len(group),
                'vendor_email': base.vendor_email or '',
                'delivery_date': _date_to_text_purchase(base.delivery_date),
                'status': _group_order_status(group),
                'attach_files': _safe_json_loads_purchase(pr.attach_files, []) if pr and pr.attach_files else [],
                'reason': pr.reason if pr else '',
                'actual_received_diff': bool(pr.actual_received_diff) if pr else False,
                'actual_items': _safe_json_loads_purchase(pr.actual_items_json, []) if pr and pr.actual_items_json else [],
                'vendor_groups': vendor_groups,
                'items': all_items,
                'item_count': len(all_items),
                'qr_count': sum(1 for o in receipt_group for it in (o.items or []) if getattr(it, 'qr_code', None)),
                'receipt_total_count': total_items_for_receipt,
                'receipt_received_count': received_items_for_receipt,
                'receipt_pending_count': max(0, total_items_for_receipt - received_items_for_receipt),
                'all_items_received': all_items_received,
                'representative_qr': next((it.qr_code for o in receipt_group for it in (o.items or []) if getattr(it, 'qr_code', None)), ''),
            })
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            print(f"[WARN] orders/detail group fallback used: {e}")
            try:
                result.append(_fallback_order_detail_row(group))
            except Exception as inner:
                print(f"[ERROR] orders/detail group skipped: {inner}")
                continue
    try:
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
    return result



@router.post('/order/{order_id}/dashboard_detail_update')
def update_dashboard_order_detail(order_id: int, data: DashboardDetailUpdate, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    """대시보드 발주/납기 상세창 관리자 수정 저장."""
    if getattr(current_user, 'role', '') != 'admin':
        raise HTTPException(status_code=403, detail='관리자만 발주/납기 상세 정보를 수정할 수 있습니다.')
    base = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not base:
        raise HTTPException(status_code=404, detail='발주 건을 찾을 수 없습니다.')
    orders = [o for o in db.query(PurchaseOrder).filter(PurchaseOrder.request_id == base.request_id).order_by(PurchaseOrder.id.asc()).all() if _is_cost_order(o)] if base.request_id else [base]
    if not orders:
        orders = [base]
    pr = base.purchase_request

    new_request_no = str(data.request_no or '').strip()
    if pr and new_request_no and new_request_no != (pr.request_no or ''):
        exists = db.query(PurchaseRequest).filter(PurchaseRequest.request_no == new_request_no, PurchaseRequest.id != pr.id).first()
        if exists:
            raise HTTPException(status_code=400, detail='이미 등록된 구매의뢰서 번호입니다.')
        pr.request_no = new_request_no

    if pr and data.title_full is not None:
        pr.title_full = str(data.title_full or '').strip()
    if pr and data.project_name is not None:
        pr.project_name = str(data.project_name or '').strip()
    if pr and data.requested_by is not None:
        pr.requested_by = str(data.requested_by or '').strip()

    if data.delivery_date is not None:
        parsed_dt = _parse_dashboard_date(data.delivery_date)
        for o in orders:
            o.delivery_date = parsed_dt

    vendor_name = str(data.vendor_name or '').strip() if data.vendor_name is not None else ''
    if vendor_name:
        # 여러 업체가 묶인 행에서 'A 외 n개' 값을 전체 업체명으로 덮어쓰는 사고를 막기 위해
        # 단일 업체 발주 건만 직접 수정한다. 여러 업체는 입고현황/발주서관리의 업체별 행에서 수정한다.
        unique_vendors = {str(o.vendor_name or '').strip() for o in orders}
        if len(orders) == 1 or len(unique_vendors) <= 1:
            for o in orders:
                o.vendor_name = vendor_name

    if data.order_completed_by is not None:
        completed_by = str(data.order_completed_by or '').strip()
        for o in orders:
            o.order_completed_by = completed_by
            if completed_by and not getattr(o, 'order_completed_at', None):
                o.order_completed_at = datetime.now()

    if data.inbound_by is not None:
        inbound_by = str(data.inbound_by or '').strip()
        for o in orders:
            for it in (o.items or []):
                it.purchase_recv_by = inbound_by
    if data.outbound_by is not None:
        outbound_by = str(data.outbound_by or '').strip()
        for o in orders:
            for it in (o.items or []):
                it.quality_recv_by = outbound_by

    item_map = {int(it.id): it for o in orders for it in (o.items or []) if getattr(it, 'id', None)}
    touched_qr_items = []
    for row in data.items or []:
        if not row.id or int(row.id) not in item_map:
            continue
        item = item_map[int(row.id)]
        if row.item_name is not None:
            item.item_name = str(row.item_name or '').strip()
            touched_qr_items.append(item)
        if row.spec is not None:
            item.spec = str(row.spec or '').strip()
            touched_qr_items.append(item)
        if row.quantity is not None:
            qty, unit = _parse_qty_unit_text(row.quantity, item.unit or 'EA')
            item.quantity = qty
            item.unit = unit
        if row.unit is not None:
            item.unit = str(row.unit or '').strip() or 'EA'

    db.commit()
    # 품명/규격 수정 후 기존 QR 이미지도 새 라벨로 저장 반영한다.
    for item in touched_qr_items:
        try:
            if item.qr_code_path and os.path.exists(str(item.qr_code_path)):
                ensure_qr_image_has_item_name(str(item.qr_code_path), item)
        except Exception:
            pass
    return {'success': True, 'message': '발주/납기 상세 정보가 저장되었습니다.'}

def _prepare_single_order_email(order: PurchaseOrder, payload: Optional[OrderEmailPrepareRequest], db: Session) -> dict:
    """발주서 PDF/QR ZIP/메일초안 생성. 웹 UI에서 단건/일괄 공통 사용."""
    pr = order.purchase_request
    if not pr:
        raise HTTPException(status_code=404, detail='원본 구매의뢰를 찾을 수 없습니다.')

    req_items = _quote_request_items(pr)
    req_vendors = _vendors_for_request_items(pr, req_items, True)
    matched_vendor_idx = 0
    for vi, vv in enumerate(req_vendors or []):
        if _vendor_key(_norm_vendor_name(vv)) and _vendor_key(_norm_vendor_name(vv)) == _vendor_key(order.vendor_name or ''):
            matched_vendor_idx = vi
            break
    src_items_for_order = _items_for_vendor_from(req_items, req_vendors, {'name': order.vendor_name or ''}, matched_vendor_idx) or list(order.items or [])

    items = []
    for src in src_items_for_order:
        if isinstance(src, dict):
            items.append(_normalize_item_spec_unit({
                'item_name': src.get('item_name', '') or '',
                'spec': src.get('spec', '') or '',
                'unit': src.get('unit') or 'EA',
                'due_date': order.delivery_date.strftime('%Y/%m/%d') if order.delivery_date else '',
                'quantity': src.get('quantity', 0),
                'unit_price': src.get('unit_price', 0),
                'qr_code_path': '',
                'axis': src.get('axis', '') or src.get('axis_type', '') or '',
                'maker': src.get('maker', '') or src.get('vendor_name', '') or '',
            }))
        else:
            items.append(_normalize_item_spec_unit({
                'item_name': src.item_name,
                'spec': src.spec,
                'unit': src.unit,
                'due_date': order.delivery_date.strftime('%Y/%m/%d') if order.delivery_date else '',
                'quantity': src.quantity,
                'unit_price': src.unit_price,
                'qr_code_path': getattr(src, 'qr_code_path', '') or '',
                'axis': getattr(src, 'axis_type', '') or '',
                'maker': getattr(src, 'maker', '') or '',
            }))

    order_data = {
        'order_no': order.order_no,
        'request_no': pr.request_no,
        'project_code': pr.project_code or '',
        'order_date': order.order_date.strftime('%Y-%m-%d') if order.order_date else datetime.now().strftime('%Y-%m-%d'),
        'delivery_date': order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else '',
        'vendor_name': order.vendor_name,
        'vendor_contact': order.vendor_contact,
        'items': items,
        'request_title': pr.title_full or pr.project_name or '',
        'orderer_name': (payload.orderer_name if payload else '') or '',
        'orderer_phone': (payload.orderer_phone if payload else '') or '',
    }

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    pdf_filename = f"{_safe_filename(pr.request_no or order.order_no)}_{_safe_filename(order.vendor_name)}_{ts}.pdf"
    pdf_path = os.path.join(ORDER_DIR, pdf_filename)
    generate_purchase_order_pdf(order_data, pdf_path)

    attach_files = json.loads(pr.attach_files or '[]') if pr.attach_files else []
    if pdf_path and os.path.exists(pdf_path):
        attach_files.append(pdf_path)

    qr_zip_path = _create_order_qr_zip(order)
    if qr_zip_path:
        attach_files.append(qr_zip_path)

    if int(pr.actual_received_diff or 0):
        try:
            actual_items = json.loads(pr.actual_items_json or '[]') if pr.actual_items_json else _effective_request_items(pr)
            src_pdf = _render_actual_item_table_pdf(pr, actual_items)
            actual_name = f"실 입고품 리스트_{_safe_filename(pr.request_no)}_{_safe_filename(pr.title_full or pr.project_name or '')[:60]}.pdf"
            actual_pdf = os.path.join(ORDER_DIR, actual_name)
            if src_pdf and os.path.exists(src_pdf):
                shutil.copyfile(src_pdf, actual_pdf)
                attach_files.append(actual_pdf)
        except Exception:
            pass

    draft_path = os.path.join(DRAFT_DIR, f'{_safe_filename(order.order_no)}_{_safe_filename(order.vendor_name)}_draft.eml')
    mail_subject = build_order_mail_subject(order_data)
    mail_body_text = build_order_mail_body(order_data)
    mail_body_html = build_order_mail_body_html(order_data)
    create_email_draft_file(
        order.vendor_email or '',
        mail_subject,
        mail_body_text,
        attach_files,
        draft_path,
        body_html=mail_body_html,
    )
    order.pdf_path = pdf_path
    order.email_draft_path = draft_path
    order.email_sent = 1
    order.status = '발주서작성'
    db.flush()
    return {
        'success': True,
        'order_id': order.id,
        'order_no': order.order_no,
        'vendor_name': order.vendor_name,
        'pdf_download_url': f'/api/purchase/order/{order.id}/pdf_file',
        'qr_zip_download_url': f'/api/purchase/order/{order.id}/qr_zip_file' if qr_zip_path else '',
        'email_draft_download_url': f'/api/purchase/order/{order.id}/email_draft_file',
        'draft_download_url': f'/api/purchase/order/{order.id}/email_draft_file',
        'package_download_url': f'/api/purchase/order/{order.id}/package_zip_file',
        'mail_to': order.vendor_email or '',
        'mail_subject': mail_subject,
        'mail_body_html': mail_body_html,
        'pdf_path': pdf_path,
        'qr_zip_path': qr_zip_path or '',
        'email_draft_path': draft_path,
    }

@router.post('/order/{order_id}/prepare_email')
def prepare_order_email(order_id: int, payload: Optional[OrderEmailPrepareRequest] = None, db: Session = Depends(get_db)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail='발주서를 찾을 수 없습니다.')
    result = _prepare_single_order_email(order, payload, db)
    db.commit()
    return result


@router.post('/order/{order_id}/prepare_email_all')
def prepare_order_email_all(order_id: int, payload: Optional[OrderEmailPrepareRequest] = None, db: Session = Depends(get_db)):
    """선택한 발주 건과 같은 구매의뢰서에 속한 모든 업체 발주서를 일괄 생성한다."""
    base = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not base:
        raise HTTPException(status_code=404, detail='발주서를 찾을 수 없습니다.')
    q = db.query(PurchaseOrder)
    if base.request_id:
        orders = q.filter(PurchaseOrder.request_id == base.request_id).order_by(PurchaseOrder.id.asc()).all()
        orders = [o for o in orders if _is_cost_order(o)]
    else:
        orders = [base] if _is_cost_order(base) else []
    results = []
    for order in orders:
        results.append(_prepare_single_order_email(order, payload, db))
    db.commit()
    return {'success': True, 'count': len(results), 'orders': results}




@router.post('/order/{order_id}/send_bizbox_mail_group')
def send_bizbox_mail_group(order_id: int, payload: Optional[OrderBizboxMailRequest] = None, background_tasks: BackgroundTasks = None, db: Session = Depends(get_db)):
    """발주서 전송: 업체별 발주서/QR/메일초안/발주건 ZIP 생성 + Bizbox 메일창 자동작성."""
    base = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not base:
        raise HTTPException(status_code=404, detail='발주서를 찾을 수 없습니다.')
    if base.request_id:
        orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == base.request_id).order_by(PurchaseOrder.id.asc()).all()
        orders = [o for o in orders if _is_cost_order(o)]
    else:
        orders = [base]
    results = []
    mail_jobs = []
    for order in orders:
        r = _prepare_single_order_email(order, OrderEmailPrepareRequest(orderer_name=(payload.orderer_name if payload else ''), orderer_phone=(payload.orderer_phone if payload else '')), db)
        package_path = _create_order_package_zip(order)
        if package_path:
            r['package_path'] = package_path
            r['package_download_url'] = f'/api/purchase/order/{order.id}/package_zip_file'
            r['package_filename'] = os.path.basename(package_path)
        attachments = [p for p in [r.get('pdf_path'), r.get('qr_zip_path')] if p]
        attachment_urls = []
        for key in ['pdf_download_url', 'qr_zip_download_url']:
            if r.get(key):
                attachment_urls.append({
                    'url': r.get(key),
                    'download_url': r.get(key),
                    'absolute_url': _public_api_url(r.get(key)),
                    'filename': os.path.basename(r.get('pdf_path') if key == 'pdf_download_url' else r.get('qr_zip_path') or ''),
                })
        # 서버 자동화는 서버 로컬 파일 경로, 클라이언트 로컬 에이전트는 attachment_urls를 사용한다.
        mail_jobs.append({
            'to': r.get('mail_to') or order.vendor_email or '',
            'vendor_name': order.vendor_name or '',
            'order_no': order.order_no or '',
            'subject': r.get('mail_subject') or '',
            'body_html': r.get('mail_body_html') or '',
            'attachments': attachments,
            'attachment_urls': attachment_urls,
        })
        results.append(r)
    for order in orders:
        if order.status != '발주진행완료':
            order.status = '발주서전송완료'
    db.commit()
    biz_id = (payload.bizbox_id if payload else '') or ''
    biz_pw = (payload.bizbox_pw if payload else '') or ''
    target = ((payload.automation_target if payload else '') or 'server').strip().lower()
    server_started = False
    if target != 'client' and biz_id and biz_pw and background_tasks is not None:
        try:
            from server.bizbox_order_mail import auto_open_order_mail_windows
            background_tasks.add_task(auto_open_order_mail_windows, mail_jobs, biz_id, biz_pw)
            server_started = True
        except Exception as e:
            results.append({'success': False, 'message': f'Bizbox 메일 자동작성 시작 실패: {e}'})
    return {
        'success': True,
        'count': len(results),
        'results': results,
        'mail_jobs': mail_jobs,
        'server_base': cfg.API_SERVER_BASE,
        'automation_target': target,
        'bizbox_started': server_started,
    }

@router.get('/order/{order_id}/package_zip_file')
def get_order_package_zip_file(order_id: int, db: Session = Depends(get_db)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail='발주 건을 찾을 수 없습니다.')
    # 발주서/메일초안이 아직 없으면 먼저 생성
    if not order.pdf_path or not os.path.exists(order.pdf_path) or not order.email_draft_path or not os.path.exists(order.email_draft_path):
        _prepare_single_order_email(order, None, db)
        db.commit()
    package = _create_order_package_zip(order)
    if not package or not os.path.exists(package):
        raise HTTPException(status_code=404, detail='발주 압축파일을 생성하지 못했습니다.')
    return FileResponse(package, filename=os.path.basename(package), media_type='application/zip')

@router.get('/order/{order_id}/email_draft_file')
def get_email_draft_file(order_id: int, db: Session = Depends(get_db)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order or not order.email_draft_path or not os.path.exists(order.email_draft_path):
        raise HTTPException(status_code=404, detail='이메일 초안 파일이 없습니다.')
    return FileResponse(order.email_draft_path, filename=os.path.basename(order.email_draft_path), media_type='message/rfc822')


@router.get('/order/{order_id}/pdf_file')
def get_order_pdf_file(order_id: int, db: Session = Depends(get_db)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order or not order.pdf_path or not os.path.exists(order.pdf_path):
        raise HTTPException(status_code=404, detail='발주 PDF 파일이 없습니다.')
    return FileResponse(order.pdf_path, filename=os.path.basename(order.pdf_path), media_type='application/pdf')


@router.get('/order/{order_id}/qr_zip_file')
def get_order_qr_zip_file(order_id: int, db: Session = Depends(get_db)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail='발주 건을 찾을 수 없습니다.')
    related = _qr_orders_for_order(db, order)
    if len(related) <= 1:
        qr_zip_path = _create_order_qr_zip(related[0]) if related else ''
    else:
        os.makedirs(DRAFT_DIR, exist_ok=True)
        req_no = order.purchase_request.request_no if order.purchase_request else order.order_no
        today = datetime.now().strftime('%Y%m%d')
        qr_zip_path = os.path.join(DRAFT_DIR, f'{_safe_filename(req_no)}_{today}_업체별_QR코드.zip')
        files = []
        seen = set()
        for po in related:
            for item in po.items:
                if (not item.qr_code or not item.qr_code_path or not os.path.exists(str(item.qr_code_path))):
                    try:
                        generate_qr_for_item(item, po, db)
                        db.refresh(item)
                    except Exception:
                        pass
                path = (item.qr_code_path or '').strip()
                if path and os.path.exists(path):
                    ensure_qr_image_has_item_name(path, item)
                if path and os.path.exists(path) and path not in seen:
                    seen.add(path)
                    files.append((po.vendor_name or '미정업체', path))
        if files:
            with zipfile.ZipFile(qr_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                used = set()
                for vendor, path in files:
                    arc = f"{_safe_filename(vendor)}/{os.path.basename(path)}"
                    base, ext = os.path.splitext(arc)
                    n = 2
                    while arc in used:
                        arc = f"{base}_{n}{ext}"; n += 1
                    used.add(arc)
                    zf.write(path, arcname=arc)
        else:
            qr_zip_path = ''
    if not qr_zip_path or not os.path.exists(qr_zip_path):
        raise HTTPException(status_code=404, detail='QR ZIP 파일이 없습니다.')
    return FileResponse(qr_zip_path, filename=os.path.basename(qr_zip_path), media_type='application/zip')


def _html_escape(v: str) -> str:
    return (str(v or '')
        .replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
        .replace('"', '&quot;')
        .replace("'", '&#39;'))

@router.get('/order/{order_id}/qr_page', response_class=HTMLResponse)
def get_order_qr_page(order_id: int, db: Session = Depends(get_db)):
    # 화면 발주관리는 구매의뢰서 1건으로 보이므로 QR 페이지도 같은 구매의뢰서의 업체별 품목을 모두 표시한다.
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order:
        return HTMLResponse('<meta charset="utf-8"><script>alert("발주 건을 찾을 수 없습니다.");window.close();</script>', status_code=404)
    related = _qr_orders_for_order(db, order)
    pr = order.purchase_request
    timestamp = int(datetime.now().timestamp())
    sections = []
    all_cards = []
    total_idx = 1
    for po in related:
        rows = []
        for item in po.items:
            if not item.qr_code or not item.qr_code_path or not os.path.exists(str(item.qr_code_path)):
                try:
                    generate_qr_for_item(item, po, db)
                    db.refresh(item)
                except Exception:
                    pass
            if item.qr_code_path and os.path.exists(str(item.qr_code_path)):
                ensure_qr_image_has_item_name(str(item.qr_code_path), item)
            img_url = f'/api/purchase/order/item/{item.id}/qr_image?ts={timestamp}' if item.qr_code else ''
            open_cell = f'<a class="btn" target="_blank" href="{img_url}">이미지 열기</a>' if img_url else '없음'
            rows.append(f'''
            <tr>
              <td>{total_idx}</td>
              <td>{_html_escape(po.vendor_name or '')}</td>
              <td>{_html_escape(item.item_name or '')}</td>
              <td>{_html_escape(item.spec or '')}</td>
              <td>{_html_escape(str(item.quantity or ''))} {_html_escape(item.unit or 'EA')}</td>
              <td><code>{_html_escape(item.qr_code or '')}</code></td>
              <td>{open_cell}</td>
            </tr>
            ''')
            if img_url:
                all_cards.append(f'''<div class="card"><b>{total_idx}. {_html_escape(po.vendor_name or '')}</b><br><b>{_html_escape(item.item_name or '')}</b><br><span class="small">{_html_escape(item.spec or '')}</span><br><a target="_blank" href="{img_url}"><img src="{img_url}" alt="QR"></a><div class="small">{_html_escape(item.qr_code or '')}</div></div>''')
            total_idx += 1
        body = ''.join(rows) or '<tr><td colspan="7" class="empty">QR 코드가 없습니다.</td></tr>'
        sections.append(f'''
        <h2>{_html_escape(po.vendor_name or '미정업체')} <span class="small">/ 발주번호: {_html_escape(po.order_no or '')} / 품목 {len(po.items or [])}개</span></h2>
        <div class="table-wrap"><table><thead><tr><th>No</th><th>업체</th><th>품명</th><th>규격</th><th>수량</th><th>QR</th><th>열기</th></tr></thead><tbody>{body}</tbody></table></div>
        ''')
    zip_url = f'/api/purchase/order/{order.id}/qr_zip_file'
    title = f"{pr.request_no if pr else order.order_no} / 업체별 QR 코드"
    vendor_label = _vendor_display_name([po.vendor_name for po in related])
    html = f'''<!doctype html>
<html lang="ko"><head><meta charset="utf-8"><title>{_html_escape(title)}</title>
<style>
body{{font-family:'Malgun Gothic',Arial,sans-serif;margin:24px;background:#f5f6fa;color:#111}}
h1{{font-size:24px;margin:0 0 12px}} h2{{font-size:18px;margin:24px 0 8px}} .meta{{margin-bottom:16px;color:#555}}
.table-wrap{{background:#fff;border:1px solid #ddd;border-radius:10px;overflow:auto;margin-bottom:14px}}
table{{border-collapse:collapse;width:100%}} th,td{{border-bottom:1px solid #e5e5e5;padding:10px;text-align:left;font-size:14px}} th{{background:#f0f2f7}}
.btn{{display:inline-block;background:#1463d8;color:white;text-decoration:none;padding:7px 12px;border-radius:6px;font-weight:700}}
.actions{{margin:14px 0;display:flex;gap:8px}} .empty{{text-align:center;padding:30px;color:#777}}
.qr-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:14px;margin-top:18px}}
.card{{background:#fff;border:1px solid #ddd;border-radius:10px;padding:12px;text-align:center}}
.card img{{max-width:180px;max-height:180px}}
.small{{font-size:12px;color:#666;word-break:break-all}}
</style></head><body>
<h1>업체별 · 품목별 QR 코드</h1>
<div class="meta">구매의뢰번호: <b>{_html_escape(pr.request_no if pr else '')}</b> / 업체: <b>{_html_escape(vendor_label)}</b> / 업체수: <b>{len(related)}</b></div>
<div class="actions"><a class="btn" href="{zip_url}">전체 QR ZIP 다운로드</a><a class="btn" href="javascript:location.reload()">새로고침/재생성</a></div>
{''.join(sections)}
<div class="qr-grid">{''.join(all_cards)}</div>
</body></html>'''
    db.commit()
    return HTMLResponse(html)
@router.get('/order/{order_id}/qr_items')
def get_order_qr_items(order_id: int, db: Session = Depends(get_db)):
    # 선택 발주 건과 같은 구매의뢰서의 모든 업체별 QR 품목을 반환.
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail='발주 건을 찾을 수 없습니다.')
    related = _qr_orders_for_order(db, order)
    out = []
    ts = int(datetime.now().timestamp())
    for po in related:
        for item in po.items:
            if (not item.qr_code or not item.qr_code_path or not os.path.exists(str(item.qr_code_path))):
                try:
                    generate_qr_for_item(item, po, db)
                    db.refresh(item)
                except Exception:
                    pass
            if item.qr_code_path and os.path.exists(str(item.qr_code_path)):
                ensure_qr_image_has_item_name(str(item.qr_code_path), item)
            out.append({
                'id': item.id,
                'order_id': po.id,
                'order_no': po.order_no,
                'vendor_name': po.vendor_name or '',
                'item_name': item.item_name or '',
                'spec': item.spec or '',
                'quantity': item.quantity or 0,
                'unit': item.unit or 'EA',
                'qr_code': item.qr_code or '',
                'qr_image_url': f'/api/purchase/order/item/{item.id}/qr_image?ts={ts}' if item.qr_code else '',
            })
    db.commit()
    return {'success': True, 'order_id': order.id, 'request_id': order.request_id, 'items': out, 'count': len(out)}

@router.get('/order/item/{item_id}/qr_image')
def get_order_item_qr_image(item_id: int, db: Session = Depends(get_db)):
    """품목별 QR 이미지 열기. ZIP이 아니라 이미지 파일을 직접 반환한다."""
    item = db.query(ReceiptItem).filter(ReceiptItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail='품목을 찾을 수 없습니다.')
    order = item.order
    if (not item.qr_code_path or not os.path.exists(item.qr_code_path)) and order:
        try:
            generate_qr_for_item(item, order, db)
            db.refresh(item)
        except Exception:
            pass
    if not item.qr_code_path or not os.path.exists(item.qr_code_path):
        raise HTTPException(status_code=404, detail='QR 이미지 파일이 없습니다.')
    ensure_qr_image_has_item_name(item.qr_code_path, item)
    return FileResponse(item.qr_code_path, filename=os.path.basename(item.qr_code_path), media_type='image/png')

@router.post('/order/{order_id}/mark_completed')
def mark_order_completed(order_id: int, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order:
        return {'success': False, 'message': '발주 건을 찾을 수 없습니다.'}
    orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == order.request_id).order_by(PurchaseOrder.id.asc()).all() if order.request_id else [order]
    orders = [o for o in orders if _is_cost_order(o)]

    not_ready = []
    for po in orders:
        send_ready = bool(
            (po.status in ('발주서전송완료', '발주서전송', '발주진행완료', '입고대기', '입고완료', '완료'))
            or (po.email_draft_path and os.path.exists(po.email_draft_path))
            or (po.pdf_path and os.path.exists(po.pdf_path))
        )
        if not send_ready:
            not_ready.append(po.vendor_name or po.order_no or str(po.id))
    if not_ready:
        return {'success': False, 'message': '먼저 발주서전송을 진행해주세요. 미완료 업체: ' + ', '.join(not_ready[:10])}

    actor = current_user.name or current_user.username
    now = datetime.now()
    for po in orders:
        po.status = '발주진행완료'
        po.order_completed_by = actor
        po.order_completed_at = now
    pr = order.purchase_request
    if pr:
        pr.status = '발주진행완료'
    db.commit()
    req_no = pr.request_no if pr else order.order_no
    title_text = _short_notify_text((getattr(pr, 'title_full', '') or getattr(pr, 'project_name', '') or '') if pr else '')
    vendors = _join_people_names([po.vendor_name for po in orders])
    _append_client_notification(
        'purchase_order_completed',
        'DevERP 발주진행완료 알림',
        f"{req_no}\n발주진행완료 처리되었습니다.\n업체: {vendors or len(orders)}\n처리자: {actor}" + (f"\n{title_text}" if title_text else ''),
        entity_key=str(req_no or order_id),
    )
    return {'success': True, 'order_no': req_no, 'status': '발주진행완료', 'count': len(orders), 'message': f'발주진행완료 처리되었습니다. 업체 {len(orders)}개가 함께 처리되었습니다.'}


@router.post('/order/{order_id}/mark_tax_docs_completed')
def mark_tax_docs_completed(order_id: int, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order:
        return {'success': False, 'message': '발주 건을 찾을 수 없습니다.'}
    orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == order.request_id).order_by(PurchaseOrder.id.asc()).all() if order.request_id else [order]
    orders = [o for o in orders if _is_cost_order(o)]
    now = datetime.now()
    actor = current_user.name or current_user.username
    for po in orders:
        po.status = '세금계산서/거래명세서 처리완료'
        po.tax_docs_completed_by = actor
        po.tax_docs_completed_at = now
    pr = order.purchase_request
    if pr:
        pr.status = '세금계산서/거래명세서 처리완료'
    db.commit()
    req_no = pr.request_no if pr else order.order_no
    title_text = _short_notify_text((getattr(pr, 'title_full', '') or getattr(pr, 'project_name', '') or '') if pr else '')
    vendors = _join_people_names([po.vendor_name for po in orders])
    _append_client_notification(
        'purchase_tax_docs_completed',
        'DevERP 세금계산서/거래명세서 알림',
        f"{req_no}\n세금계산서/거래명세서 처리완료.\n업체: {vendors or len(orders)}\n처리자: {actor}" + (f"\n{title_text}" if title_text else ''),
        entity_key=str(req_no or order_id),
    )
    return {
        'success': True,
        'order_no': req_no,
        'status': '세금계산서/거래명세서 처리완료',
        'count': len(orders),
        'message': f'세금계산서/거래명세서 처리완료 상태로 변경되었습니다. 업체 {len(orders)}개가 함께 처리되었습니다.'
    }

@router.post('/qr/test')
def create_test_qr(data: QRTestCreate):
    import qrcode
    from shared.config import QR_SAVE_PATH
    os.makedirs(QR_SAVE_PATH, exist_ok=True)
    file_path = os.path.join(QR_SAVE_PATH, f"TEST_QR_{datetime.now().strftime('%Y%m%d%H%M%S')}.png")
    qrcode.make(data.text).save(file_path)
    return {'success': True, 'file_path': file_path, 'text': data.text}


@router.get('/delivery/deadlines')
def get_deadlines(db: Session = Depends(get_db)):
    today = datetime.now()
    orders = db.query(PurchaseOrder).filter(PurchaseOrder.status.in_(['발주', '입고대기', '발주진행완료', '발주서작성'])).all()
    result = []
    for o in orders:
        if o.delivery_date:
            days = (o.delivery_date - today).days
            result.append({'order_no': o.order_no, 'vendor': o.vendor_name, 'due_date': o.delivery_date.strftime('%Y-%m-%d'), 'days_left': days,
                           'status': '지연' if days < 0 else ('임박' if days <= 3 else '정상')})
    return sorted(result, key=lambda x: x['days_left'])



# ─────────────────────────────────────────────
# MATERIAL_BUDGET_PATCH_20260520
# 정부과제 재료비관리 엑셀 연동
# - 엑셀 탭명 = 정부과제
# - 구매연도 = 연차
# - 구분/구매비용 = 연차별 배정 예산
# - 사용금액/잔액은 구매의뢰 DB 기준으로 자동 계산 후 엑셀에 최신화
# ─────────────────────────────────────────────
def _material_master_candidates() -> list[str]:
    base_dir = getattr(cfg, 'BASE_DIR', os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    exe_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.getcwd()
    meipass = getattr(sys, '_MEIPASS', '')
    names = ['Purchase request_master.xlsx', 'Purchase_request_master.xlsx', 'material_budget_master.xlsx']
    roots = [
        os.path.join(base_dir, 'database'),
        os.path.join(os.getcwd(), 'database'),
        os.path.join(os.getcwd(), '_internal', 'database'),
        os.path.join(exe_dir, 'database'),
        os.path.join(exe_dir, '_internal', 'database'),
    ]
    if meipass:
        roots.append(os.path.join(meipass, 'database'))
    out = []
    for r in roots:
        for n in names:
            p = os.path.join(r, n)
            if p not in out:
                out.append(p)
    return out


def _material_master_path(create: bool = False) -> str:
    candidates = _material_master_candidates()
    for p in candidates:
        if os.path.exists(p):
            return p
    p = candidates[0]
    if create:
        os.makedirs(os.path.dirname(p), exist_ok=True)
    return p


def _budget_clean(v) -> str:
    return str(v or '').replace('\ufeff', '').strip()


def _budget_amount(v) -> float:
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(',', '').replace('원', '').strip()
    try:
        return float(s or 0)
    except Exception:
        return 0.0


def _budget_has_value(v) -> bool:
    if v is None:
        return False
    return str(v).replace(',', '').replace('원', '').strip() != ''


def _budget_year_seq(year_text: str) -> str:
    import re
    m = re.search(r'(\d+)\s*차', str(year_text or ''))
    if not m:
        return ''
    return f"{int(m.group(1)):02d}"


def _budget_project_base(project_code: str) -> str:
    pc = _budget_clean(project_code).upper()
    if '_' in pc:
        return pc.split('_', 1)[0]
    return pc


def _budget_full_project_code(project_base: str, project_year: str) -> str:
    base = _budget_project_base(project_base)
    seq = _budget_year_seq(project_year)
    return f'{base}_{seq}' if base and seq else base


def _load_material_budget_base_rows() -> list[dict]:
    path = _material_master_path(create=False)
    if not os.path.exists(path):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for ws in wb.worksheets:
        current_year = ''
        for r in ws.iter_rows(min_row=3, values_only=True):
            year = _budget_clean(r[0] if len(r) > 0 else '')
            if year:
                current_year = year
            budget_type = _budget_clean(r[1] if len(r) > 1 else '')
            amount = _budget_amount(r[2] if len(r) > 2 else 0)
            # D열(사용금액)은 화면 표시용 자동 계산값일 수 있으므로, F열(수동사용금액)이 있을 때만 수동 보정값으로 본다.
            # 기존 엑셀에 F열이 없는 경우에는 D열을 자동 계산 결과로 유지하고, 재료비관리 화면에서 사용금액을 수정/저장하면 F열에 보정값이 저장된다.
            manual_used = None
            if len(r) > 5 and _budget_has_value(r[5]):
                manual_used = _budget_amount(r[5])
            if not current_year and not budget_type and not amount and manual_used is None:
                continue
            if not budget_type:
                continue
            row = {
                'project_code': _budget_clean(ws.title).upper(),
                'project_year': current_year,
                'budget_type': budget_type,
                'budget_amount': amount,
            }
            if manual_used is not None:
                row['manual_used_amount'] = manual_used
            rows.append(row)
    # 같은 정부과제/연차/구분은 화면에서 한 개의 드롭다운 항목으로만 선택되므로
    # 중복 행은 예산을 합산해 관리한다. 예: 1차(2024) 가공품이 2행이면 합산 잔액으로 표시.
    merged = []
    index = {}
    for r in rows:
        key = _material_row_key(r.get('project_code', ''), r.get('project_year', ''), r.get('budget_type', ''))
        if key in index:
            m = merged[index[key]]
            m['budget_amount'] = float(m.get('budget_amount') or 0) + float(r.get('budget_amount') or 0)
            if 'manual_used_amount' in r:
                m['manual_used_amount'] = float(m.get('manual_used_amount') or 0) + float(r.get('manual_used_amount') or 0)
        else:
            index[key] = len(merged)
            merged.append(dict(r))
    return merged


def _material_row_key(project_code: str, project_year: str, budget_type: str) -> tuple[str, str, str]:
    return (_budget_project_base(project_code), _budget_clean(project_year), _budget_clean(budget_type))


def _material_year_from_code(project_code: str, rows: list[dict]) -> str:
    pc = _budget_clean(project_code).upper()
    if '_' not in pc:
        return ''
    base, seq = pc.rsplit('_', 1)
    for r in rows:
        if _budget_project_base(r.get('project_code', '')) == base and _budget_year_seq(r.get('project_year', '')) == seq:
            return _budget_clean(r.get('project_year', ''))
    return ''


def _material_type_from_title(title: str) -> str:
    import re
    parts = re.findall(r'\[([^\]]+)\]', str(title or ''))
    # [정부과제코드][항목][세세목][구분] 구조에서는 4번째가 구분이다.
    if len(parts) >= 4:
        return _budget_clean(parts[3])
    return _budget_clean(parts[-1]) if parts else ''


def _purchase_request_budget_identity(pr: PurchaseRequest, rows: list[dict]) -> tuple[str, str, str]:
    project_base = _budget_project_base(getattr(pr, 'project_code', '') or '')
    project_year = _budget_clean(getattr(pr, 'project_year', '') or '') or _material_year_from_code(getattr(pr, 'project_code', '') or '', rows)
    budget_type = _budget_clean(getattr(pr, 'budget_type', '') or getattr(pr, 'item_type', '') or '') or _material_type_from_title(getattr(pr, 'title_full', '') or '')
    return _material_row_key(project_base, project_year, budget_type)


def _purchase_request_total_amount(pr: PurchaseRequest) -> float:
    """구매의뢰 화면 품목표 기준 합계.

    예상금액(amount)을 직접 입력한 행은 amount를 우선 사용하고,
    amount가 비어 있으면 예상단가 × 수량으로 계산한다.
    """
    total = 0.0
    try:
        items = json.loads(pr.items_json or '[]') if pr.items_json else []
    except Exception:
        items = []
    for it in items or []:
        if not isinstance(it, dict):
            continue
        amount = _budget_amount(it.get('amount') or it.get('total') or it.get('total_amount'))
        if not amount:
            qty = _budget_amount(it.get('quantity') or 1) or 1
            amount = _budget_amount(it.get('unit_price')) * qty
        total += amount
    if not total:
        total = _budget_amount(pr.unit_price) * (_budget_amount(pr.quantity or 1) or 1)
    return float(total or 0)


def _purchase_order_total_amount(order: PurchaseOrder) -> float:
    """상신완료 후 실제 발주/입고현황에 생성된 품목 기준 합계."""
    total = 0.0
    try:
        for it in getattr(order, 'items', []) or []:
            qty = _budget_amount(getattr(it, 'quantity', 0) or 0)
            unit_price = _budget_amount(getattr(it, 'unit_price', 0) or 0)
            total += unit_price * qty
    except Exception:
        return 0.0
    return float(total or 0)


def _material_usage_map(db: Session, rows: Optional[list[dict]] = None) -> dict[tuple[str, str, str], float]:
    rows = rows if rows is not None else _load_material_budget_base_rows()
    usage = {}
    excluded_pr_statuses = {'', '작성완료', '결재중', '삭제', '취소'}
    excluded_order_statuses = {'', '삭제', '취소'}
    used_request_ids = set()
    active_order_request_ids = set()

    try:
        # 재료비 차감 기준은 [상신완료]로 확정되어 발주서관리/입고현황에 생성된 품목 합계다.
        # 발주서관리 또는 입고현황에서 해당 건을 삭제하면 PurchaseOrder/ReceiptItem이 삭제되고,
        # 같은 합계가 자동으로 사용금액에서 빠져 잔액이 원상복구된다.
        # 단, 품목 삭제 재생성 방지를 위해 품목 0개짜리 PurchaseOrder를 DB에 남기는 경우가 있으므로
        # 활성 발주 건이 하나라도 있는 구매의뢰는 구매의뢰 원본 금액 fallback을 적용하지 않는다.
        for order in db.query(PurchaseOrder).all():
            # 재료비/비용처리는 발주관리 기준(cost/normal) 품목만 집계한다.
            # 실제 입고품 다름 건의 receipt 발주/QR 품목을 같이 더하면 비용이 이중 차감된다.
            if not _is_cost_order(order):
                continue
            pr = getattr(order, 'purchase_request', None)
            if not pr:
                continue
            pr_status = _budget_clean(getattr(pr, 'status', ''))
            order_status = _budget_clean(getattr(order, 'status', ''))
            if pr_status in excluded_pr_statuses or order_status in excluded_order_statuses:
                continue
            if getattr(pr, 'id', None):
                active_order_request_ids.add(pr.id)
            amount = _purchase_order_total_amount(order)
            if amount <= 0:
                continue
            key = _purchase_request_budget_identity(pr, rows)
            if not key[0] or not key[1] or not key[2]:
                continue
            usage[key] = usage.get(key, 0.0) + amount
            if getattr(pr, 'id', None):
                used_request_ids.add(pr.id)

        # 예외 보정: 과거 DB에 상신완료 상태만 있고 발주/입고 품목이 없는 경우에는
        # 구매의뢰 품목표 합계로만 보정한다. 정상 신규 건은 위 PurchaseOrder 기준이 우선이다.
        for pr in db.query(PurchaseRequest).all():
            if getattr(pr, 'id', None) in used_request_ids or getattr(pr, 'id', None) in active_order_request_ids:
                continue
            status = _budget_clean(getattr(pr, 'status', ''))
            if status in excluded_pr_statuses:
                continue
            key = _purchase_request_budget_identity(pr, rows)
            if not key[0] or not key[1] or not key[2]:
                continue
            amount = _purchase_request_total_amount(pr)
            if amount > 0:
                usage[key] = usage.get(key, 0.0) + amount
    except Exception:
        return usage
    return usage

def _material_budget_rows_with_usage(db: Optional[Session] = None) -> list[dict]:
    rows = _load_material_budget_base_rows()
    usage = _material_usage_map(db, rows) if db is not None else {}
    out = []
    for r in rows:
        key = _material_row_key(r.get('project_code', ''), r.get('project_year', ''), r.get('budget_type', ''))
        db_used = float(usage.get(key, 0.0) or 0.0)
        manual_raw = r.get('manual_used_amount', None)
        manual_used = float(manual_raw) if manual_raw is not None and str(manual_raw) != '' else None
        used = manual_used if manual_used is not None else db_used
        budget = float(r.get('budget_amount') or 0.0)
        row = dict(r)
        row['project_code_full'] = _budget_full_project_code(row.get('project_code', ''), row.get('project_year', ''))
        row['used_amount'] = used
        row['db_used_amount'] = db_used
        if manual_used is not None:
            row['manual_used_amount'] = manual_used
        row['remaining_amount'] = budget - used
        out.append(row)
    return out


def _write_material_budget_rows(rows: list[dict], db: Optional[Session] = None, allow_manual_used: bool = False) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    path = _material_master_path(create=True)
    wb = openpyxl.Workbook()
    # 기본 Sheet 제거는 실제 시트 추가 후 수행
    default_ws = wb.active
    grouped = {}
    order = []
    for raw in rows or []:
        project = _budget_project_base(raw.get('project_code') or raw.get('project') or raw.get('sheet') or '')
        year = _budget_clean(raw.get('project_year') or raw.get('year') or '')
        btype = _budget_clean(raw.get('budget_type') or raw.get('type') or raw.get('구분') or '')
        amount = _budget_amount(raw.get('budget_amount') if 'budget_amount' in raw else raw.get('amount'))
        manual_used = None
        if 'manual_used_amount' in raw and raw.get('manual_used_amount') is not None and str(raw.get('manual_used_amount')) != '':
            manual_used = _budget_amount(raw.get('manual_used_amount'))
        elif allow_manual_used and 'used_amount' in raw and raw.get('used_amount') is not None and str(raw.get('used_amount')) != '':
            manual_used = _budget_amount(raw.get('used_amount'))
        if not project or not year or not btype:
            continue
        if project not in grouped:
            grouped[project] = []
            order.append(project)
        row = {'project_code': project, 'project_year': year, 'budget_type': btype, 'budget_amount': amount}
        if manual_used is not None:
            row['manual_used_amount'] = manual_used
        grouped[project].append(row)
    if not order:
        grouped = {'R25GA01': []}
        order = ['R25GA01']
    if default_ws.title == 'Sheet' and order:
        wb.remove(default_ws)
    tmp_rows = [r for p in order for r in grouped.get(p, [])]
    # 미수정 행은 DB 사용액으로 자동 최신화하고, 사용자가 사용금액을 직접 수정한 행은 수동 보정값을 우선한다.
    usage = _material_usage_map(db, tmp_rows) if db is not None else {}
    header_fill = PatternFill('solid', fgColor='E5E7EB')
    sub_fill = PatternFill('solid', fgColor='F9FAFB')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for project in order:
        ws = wb.create_sheet(project[:31])
        ws.append(['구매연도', '구분', '구매비용', '사용금액', '잔액', '수동사용금액'])
        ws.append(['', '', '(원)', '(원)', '(원)', '(원)'])
        last_year = None
        for r in grouped.get(project, []):
            year_cell = r['project_year'] if r['project_year'] != last_year else ''
            last_year = r['project_year']
            key = _material_row_key(project, r['project_year'], r['budget_type'])
            db_used = float(usage.get(key, 0.0) or 0.0)
            manual_raw = r.get('manual_used_amount', None)
            manual_used = float(manual_raw) if manual_raw is not None and str(manual_raw) != '' else None
            used = manual_used if manual_used is not None else db_used
            budget = float(r['budget_amount'] or 0.0)
            ws.append([year_cell, r['budget_type'], budget, used, budget - used, manual_used if manual_used is not None else None])
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[2]:
            cell.fill = sub_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col, width in [('A', 14), ('B', 34), ('C', 16), ('D', 16), ('E', 16), ('F', 16)]:
            ws.column_dimensions[col].width = width
        ws.column_dimensions['F'].hidden = True
        for col in ['C', 'D', 'E', 'F']:
            for cell in ws[col][2:]:
                cell.number_format = '#,##0'
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


def _sync_material_budget_master(db: Optional[Session] = None) -> None:
    try:
        rows = _load_material_budget_base_rows()
        if rows:
            _write_material_budget_rows(rows, db)
    except Exception:
        # 재료비관리 엑셀 동기화 실패가 구매의뢰 저장을 막으면 안 된다.
        pass


def _material_budget_response(db: Optional[Session] = None) -> dict:
    rows = _material_budget_rows_with_usage(db)
    projects = []
    years_by_project = {}
    types_by_project_year = {}
    for r in rows:
        p = _budget_clean(r.get('project_code')).upper()
        y = _budget_clean(r.get('project_year'))
        t = _budget_clean(r.get('budget_type'))
        if p and p not in projects:
            projects.append(p)
        years_by_project.setdefault(p, [])
        if y and y not in years_by_project[p]:
            years_by_project[p].append(y)
        key = f'{p}||{y}'
        types_by_project_year.setdefault(key, [])
        if t and t not in types_by_project_year[key]:
            types_by_project_year[key].append(t)
    return {
        'success': True,
        'path': _material_master_path(create=False),
        'projects': projects,
        'years_by_project': years_by_project,
        'types_by_project_year': types_by_project_year,
        'rows': rows,
    }


@router.get('/material_budget')
def get_material_budget(db: Session = Depends(get_db)):
    _sync_material_budget_master(db)
    return _material_budget_response(db)


@router.get('/material_budget/lookup')
def lookup_material_budget(project_code: str = '', project_year: str = '', budget_type: str = '', db: Session = Depends(get_db)):
    _sync_material_budget_master(db)
    rows = _material_budget_rows_with_usage(db)
    key = _material_row_key(project_code, project_year, budget_type)
    for r in rows:
        if _material_row_key(r.get('project_code', ''), r.get('project_year', ''), r.get('budget_type', '')) == key:
            return {'success': True, 'row': r}
    return {'success': False, 'message': '해당 정부과제/연차/구분의 재료비 예산을 찾을 수 없습니다.', 'row': None}


@router.post('/material_budget/save')
def save_material_budget(data: MaterialBudgetSave, db: Session = Depends(get_db)):
    path = _write_material_budget_rows(data.rows, db, allow_manual_used=True)
    return {**_material_budget_response(db), 'message': f'재료비관리 엑셀 저장 완료: {os.path.basename(path)}'}

@router.get('/vendor_lookup')
def vendor_lookup(name: str):
    try:
        from server.vendor_import import load_vendor_list, get_vendor_email_by_name
        vendor_list = load_vendor_list()
        found = get_vendor_email_by_name(name, vendor_list)
        return {
            'name': found.get('vendor_name', name),
            'category': found.get('category', ''),
            'sub_category': found.get('sub_category', ''),
            'ceo': found.get('ceo', found.get('representative','')),
            'representative': found.get('representative', found.get('ceo','')),
            'biz_no': found.get('biz_no', found.get('business_no', found.get('registration_no',''))),
            'business_no': found.get('business_no', found.get('biz_no','')),
            'registration_no': found.get('registration_no', found.get('biz_no','')),
            'contact_name': found.get('contact_name', ''),
            'contact': found.get('contact', found.get('phone', '')), 
            'phone': found.get('phone', ''),
            'email': found.get('email', ''),
            'fax': found.get('fax', ''),
            'address': found.get('address', ''),
            'note': found.get('note', ''),
        }
    except Exception:
        return {'name': name, 'contact': '', 'email': '', 'fax': ''}


@router.get('/vendors')
def get_vendor_directory(query: str = ''):
    from server.vendor_import import load_vendor_list
    vendors = load_vendor_list()
    q = query.strip().lower()
    if q:
        filtered = []
        for item in vendors:
            hay = ' '.join([
                str(item.get('vendor_name', '')),
                str(item.get('sub_category', '')),
                str(item.get('ceo', '')),
                str(item.get('biz_no', '')),
                str(item.get('contact_name', '')),
                str(item.get('contact', '')), 
                str(item.get('email', '')),
                str(item.get('address', '')),
                str(item.get('maker', '')),
                str(item.get('note', '')), 
            ]).lower()
            if q in hay:
                filtered.append(item)
        vendors = filtered
    vendors.sort(key=lambda x: (x.get('vendor_name', ''), x.get('category', '')))
    return vendors


@router.post('/vendors/save')
def save_vendor_directory(data: VendorDirectorySave):
    from server.vendor_import import save_custom_vendor_entries
    return save_custom_vendor_entries(data.vendors)


@router.delete('/order/{order_id}')
def delete_order(order_id: int, db: Session = Depends(get_db), current_user: User = Depends(_get_current_user)):
    if getattr(current_user, 'role', '') != 'admin':
        raise HTTPException(status_code=403, detail='관리자만 발주 건을 삭제할 수 있습니다.')
    """발주서관리 삭제.

    화면은 구매의뢰서 번호 기준으로 1건처럼 보이지만 DB에는 업체별 PurchaseOrder가 여러 건일 수 있다.
    따라서 대표 발주 1건을 삭제하면 같은 구매의뢰서의 발주/입고/QR을 함께 삭제하고,
    구매의뢰 상태를 작성완료로 되돌려 재료비관리 사용금액도 원상복구되게 한다.
    """
    order = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not order:
        return {'success': False, 'message': '발주 건을 찾을 수 없습니다.'}

    req = order.purchase_request
    if order.request_id:
        orders = db.query(PurchaseOrder).filter(PurchaseOrder.request_id == order.request_id).all()
        orders = [o for o in orders if _is_cost_order(o)]
    else:
        orders = [order]

    deleted_orders = 0
    deleted_items = 0
    for po in list(orders):
        for item in list(po.items):
            try:
                if item.qr_code_path and os.path.exists(item.qr_code_path):
                    os.remove(item.qr_code_path)
            except Exception:
                pass
            db.delete(item)
            deleted_items += 1
        for path in [po.pdf_path, po.email_draft_path]:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass
        db.delete(po)
        deleted_orders += 1

    if req:
        # 재료비관리 사용금액은 상신완료 이후 상태만 차감한다.
        # 발주/입고 데이터를 삭제하면 다시 작성완료로 되돌려 사용금액에서 제외한다.
        req.status = '작성완료'
        req.bizbox_uploaded = 0
    db.commit()
    _sync_material_budget_master(db)
    return {
        'success': True,
        'message': f'삭제되었습니다. 발주 {deleted_orders}건 / 품목 {deleted_items}개를 삭제했고, 재료비관리 사용금액을 원상복구했습니다.'
    }

@router.delete('/request/{request_id}')
def delete_request(request_id: int, db: Session = Depends(get_db)):
    pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == request_id).first()
    if not pr:
        return {'success': False, 'message': '구매의뢰를 찾을 수 없습니다.'}
    for order in list(pr.orders):
        for item in list(order.items):
            try:
                if item.qr_code_path and os.path.exists(item.qr_code_path):
                    os.remove(item.qr_code_path)
            except Exception:
                pass
            db.delete(item)
        for path in [order.pdf_path, order.email_draft_path]:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass
        db.delete(order)
    try:
        for path in json.loads(pr.attach_files or '[]'):
            if path and os.path.exists(path):
                os.remove(path)
    except Exception:
        pass
    db.delete(pr)
    db.commit()
    _sync_material_budget_master(db)
    return {'success': True}
