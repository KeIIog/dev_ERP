
import re as _re


def _normalize_vendor_name(name: str) -> str:
    """업체명 정규화: 법인어 제거 + 공백/특수문자 제거 + 소문자"""
    name = str(name or '').strip()
    name = _re.sub(r'\(주\)|㈜|\(유\)|주식회사|유한회사|\(사\)', '', name)
    name = _re.sub(r'[\s·・\-_.,/]', '', name).lower()
    name = name.replace('레이져', '레이저').replace('상호', '')
    return name


def get_vendor_email_by_name(query: str, vendor_list: list) -> dict:
    """업체명 퍼지 매칭 (정규화 후 완전일치 → 포함관계 순)"""
    q = _normalize_vendor_name(query)
    if not q:
        return {}
    for v in vendor_list:
        if _normalize_vendor_name(v.get('vendor_name', '')) == q:
            return v
    for v in vendor_list:
        vn = _normalize_vendor_name(v.get('vendor_name', ''))
        if vn and (q in vn or vn in q):
            return v
    return {}

# server/vendor_import.py
import json
import openpyxl
import sys, os, logging
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

logger = logging.getLogger(__name__)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_VENDOR_XLSX = os.path.join(BASE_DIR, 'database', 'vendor_master.xlsx')
CUSTOM_VENDOR_JSON = os.path.join(BASE_DIR, 'database', 'vendor_overrides.json')


def _clean(v):
    return str(v or '').strip()


def _find_header_row(ws):
    """업체관리 엑셀의 헤더 행을 자동으로 찾는다."""
    header_keywords = {'거래처명', '업체명', '대표자명', '사업자등록번호'}
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), start=1):
        vals = [str(v or '').strip() for v in row]
        if any(k in vals for k in header_keywords) and ('거래처명' in vals or '업체명' in vals):
            return idx, vals
    return 1, [str(v or '').strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]


def _norm_header_key(v):
    return _re.sub(r'[\s\r\n\t·・\-_()./]', '', str(v or '')).lower()

def _first_nonempty(*values):
    for v in values:
        if v is not None and str(v).strip():
            return str(v).strip()
    return ''

def _col_map(headers):
    # BIZNO_HEADER_NORMALIZE_FIX_20260509
    mapping = {}
    aliases = {
        'vendor_name': ['거래처명', '업체명', '상호', '회사명', '거래처'],
        'ceo': ['대표자명', '대표명', '대표자', '대표자 성명', '대표이사'],
        'biz_no': ['사업자등록번호', '사업자번호', '사업자 등록번호', '사업자 등록 번호', '사업자등록 번호', '사업자등록', '사업자', '등록번호', '사업자No', '사업자NO', 'BusinessNo', 'Business No', 'BUSINESS_NO'],
        'contact_name': ['거래처담당자', '담당자', '담당자명', '거래처 담당자'],
        'phone': ['핸드폰번호', '연락처', '담당자 연락처', '전화번호', '핸드폰 번호', '휴대폰번호'],
        'email': ['담당자 E-MAIL', '담당자 EMAIL', '담당자 이메일', '이메일', 'E-MAIL', 'email', 'EMAIL'],
        'address': ['본사주소(기본)', '주소', '사업장 소재지', '사업장소재지', '본사주소', '소재지'],
        'maker': ['Maker', '취급메이커', '메이커'],
        'note': ['비고', '메모'],
        'sub_category': ['구분', '세세목', '분류'],
    }
    norm_headers = [_norm_header_key(h) for h in headers]
    for key, names in aliases.items():
        norm_names = [_norm_header_key(n) for n in names]
        for i, nh in enumerate(norm_headers):
            if nh in norm_names or any(nn and nn in nh for nn in norm_names):
                mapping[key] = i
                break
    return mapping

def _get(row, idx):
    if idx is None or idx >= len(row):
        return ''
    return _clean(row[idx])


def _find_biz_no_in_row(row):
    for v in row or []:
        s = str(v or '').strip()
        if not s:
            continue
        m = _re.search(r'\b\d{3}[-\s]?\d{2}[-\s]?\d{5}\b', s)
        if m:
            raw = m.group(0).replace(' ', '').replace('-', '')
            if len(raw) == 10:
                return f'{raw[:3]}-{raw[3:5]}-{raw[5:]}'
            return m.group(0)
    return ''

def parse_vendor_xlsx(file_path: str) -> list:
    """업체관리 엑셀을 읽어 업체 목록으로 변환한다.

    최신 양식: 거래처명 / 대표자명 / 사업자등록번호 / 거래처담당자 / 핸드폰번호 / 담당자 E-MAIL / 본사주소(기본)
    구버전 양식: 가공품/구매품 시트도 함께 지원한다.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    vendors = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, headers = _find_header_row(ws)
        cmap = _col_map(headers)
        category = '' if ('거래처명' in headers or 'ERP 등록 업체' in str(headers)) else _clean(sheet_name)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v is None for v in row):
                continue
            # 최신 양식 우선
            vendor_name = _get(row, cmap.get('vendor_name'))
            if not vendor_name and len(row) > 1:
                vendor_name = _clean(row[1])
            if not vendor_name or vendor_name in ('거래처명', '업체명'):
                continue
            ceo = _get(row, cmap.get('ceo'))
            biz_no = _get(row, cmap.get('biz_no'))
            biz_no = biz_no or _find_biz_no_in_row(row)
            contact_name = _get(row, cmap.get('contact_name'))
            phone = _get(row, cmap.get('phone'))
            email = _get(row, cmap.get('email'))
            address = _get(row, cmap.get('address'))
            note = _get(row, cmap.get('note'))
            maker = _get(row, cmap.get('maker'))
            sub_cat = _get(row, cmap.get('sub_category'))

            # 구버전 가공/구매 양식 보정
            if not any([ceo, biz_no, contact_name, phone, email, address]) and len(row) >= 7:
                if sheet_name == '가공품':
                    sub_cat = _clean(row[0]); vendor_name = _clean(row[1]); ceo = _clean(row[2])
                    contact_name = _clean(row[3]); phone = _clean(row[4]); email = _clean(row[5]); address = _clean(row[6])
                else:
                    vendor_name = _clean(row[1]); sub_cat = _clean(row[2]); contact_name = _clean(row[3])
                    phone = _clean(row[4]); email = _clean(row[5]); address = _clean(row[6]); note = _clean(row[7] if len(row) > 7 else '')
            vendors.append({
                'category': category,
                'sub_category': sub_cat,
                'vendor_name': vendor_name,
                'name': vendor_name,
                'ceo': ceo,
                'representative': ceo,
                'biz_no': biz_no,
                'business_no': biz_no,
                'registration_no': biz_no,
                'contact_name': contact_name,
                'contact': f"{contact_name} / {phone}".strip(' /'),
                'phone': phone,
                'email': email,
                'fax': '',
                'address': address,
                'maker': maker,
                'note': note,
                'source': 'xlsx',
            })
    logger.info('업체 %s개 파싱 완료', len(vendors))

    # VENDOR_BIZNO_ALIAS_FIX_20260509
    # 업체관리 엑셀/JSON의 사업자등록번호 컬럼명이 제각각이어도 biz_no로 통일
    for v in vendors:
        biz = (
            v.get('biz_no') or v.get('business_no') or v.get('registration_no')
            or v.get('사업자등록번호') or v.get('사업자 등록번호') or v.get('사업자번호')
            or v.get('사업자 번호') or v.get('사업자등록') or v.get('사업자')
            or v.get('BUSINESS_NO') or v.get('Business No') or v.get('BusinessNo')
            or v.get('tax_no') or v.get('tax_id') or v.get('corp_no') or v.get('corporate_no')
        )
        if biz:
            biz = str(biz).strip()
            v['biz_no'] = biz
            v['business_no'] = biz
            v['registration_no'] = biz
    return vendors

def load_custom_vendor_entries() -> list:
    if not os.path.exists(CUSTOM_VENDOR_JSON):
        return []
    try:
        with open(CUSTOM_VENDOR_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_custom_vendor_entries(vendors: list[dict]) -> dict:
    os.makedirs(os.path.dirname(CUSTOM_VENDOR_JSON), exist_ok=True)
    norm = []
    for item in vendors or []:
        name = _clean(item.get('vendor_name') or item.get('name'))
        if not name:
            continue
        ceo = _clean(item.get('ceo') or item.get('representative'))
        # CUSTOM_BIZNO_ALIAS_FIX_20260509
        biz_no = _clean(item.get('biz_no') or item.get('business_no') or item.get('registration_no') or item.get('사업자등록번호') or item.get('사업자 등록번호') or item.get('사업자번호') or item.get('사업자 번호') or item.get('사업자등록') or item.get('사업자'))
        contact_name = _clean(item.get('contact_name'))
        phone = _clean(item.get('phone'))
        norm.append({
            'category': _clean(item.get('category')),
            'sub_category': _clean(item.get('sub_category')),
            'vendor_name': name,
            'name': name,
            'ceo': ceo,
            'representative': ceo,
            'biz_no': biz_no,
            'business_no': biz_no,
            'registration_no': biz_no,
            'contact_name': contact_name,
            'contact': _clean(item.get('contact')) or f"{contact_name} / {phone}".strip(' /'),
            'phone': phone,
            'email': _clean(item.get('email')),
            'fax': _clean(item.get('fax')),
            'address': _clean(item.get('address')),
            'maker': _clean(item.get('maker')),
            'note': _clean(item.get('note')),
            'source': 'custom',
        })
    with open(CUSTOM_VENDOR_JSON, 'w', encoding='utf-8') as f:
        json.dump(norm, f, ensure_ascii=False, indent=2)
    return {'success': True, 'count': len(norm)}

def load_vendor_list(xlsx_path: str = DEFAULT_VENDOR_XLSX) -> list:
    # MERGE_KEEP_NONEMPTY_BIZNO_FIX_20260509
    candidates = []
    for p in [xlsx_path, DEFAULT_VENDOR_XLSX]:
        if p and p not in candidates:
            candidates.append(p)
    base = []
    for p in candidates:
        if not p or not os.path.exists(p):
            continue
        try:
            base.extend(parse_vendor_xlsx(p))
        except Exception as e:
            logger.exception('업체관리 엑셀 로드 실패: %s / %s', p, e)
    custom = load_custom_vendor_entries()
    merged = {}
    def _biz(entry):
        return _first_nonempty(entry.get('biz_no'), entry.get('business_no'), entry.get('registration_no'), entry.get('사업자등록번호'), entry.get('사업자 등록번호'), entry.get('사업자번호'), entry.get('사업자 번호'), entry.get('사업자등록'), entry.get('사업자'))
    for entry in base + custom:
        name = _clean(entry.get('vendor_name') or entry.get('name'))
        if not name:
            continue
        key = _normalize_vendor_name(name) or name.lower()
        existing = merged.get(key, {})
        row = {**existing}
        for k, v in entry.items():
            if v is not None and str(v).strip():
                row[k] = v
            elif k not in row:
                row[k] = v
        row['vendor_name'] = name
        row['name'] = name
        biz = _biz(row) or _biz(existing) or _biz(entry)
        if biz:
            row['biz_no'] = biz
            row['business_no'] = biz
            row['registration_no'] = biz
        merged[key] = row
    out = list(merged.values())
    out.sort(key=lambda x: (str(x.get('vendor_name') or ''), str(x.get('category') or '')))
    return out

