# server/inspection_report_generator.py
#
# 검수조사서 자동 생성
# 트리거: 사용자가 프로그램에서 실사진을 해당 구매의뢰서에 업로드할 때
#
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime
from PIL import Image as PILImage

# PDF fallback generator
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

_FONT_REGISTERED = False
def _ensure_pdf_font():
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return 'MalgunGothic'
    candidates = [
        r'C:\Windows\Fonts\malgun.ttf',
        r'C:\Windows\Fonts\malgunbd.ttf',
        '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    for fp in candidates:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('MalgunGothic', fp))
                _FONT_REGISTERED = True
                return 'MalgunGothic'
            except Exception:
                pass
    return 'Helvetica'


# ── 스타일 헬퍼 ───────────────────────────────────────
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _medium():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _font(bold=False, size=10, color="000000"):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color)

def _safe_cell(ws, row, col):
    """병합셀 내부 좌표가 들어와도 실제 값을 쓸 수 있는 좌상단 셀로 변환."""
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            row, col = merged.min_row, merged.min_col
            break
    return ws.cell(row=row, column=col)

def _apply(ws, row, col, value="", bold=False, size=10, fill=None,
           align="center", border=True, color="000000"):
    cell = _safe_cell(ws, row, col)
    cell.value = value
    cell.font  = _font(bold=bold, size=size, color=color)
    cell.alignment = _center() if align == "center" else _left()
    if fill:
        cell.fill = _fill(fill)
    if border:
        cell.border = _thin()
    return cell


def generate_inspection_report(data: dict, photo_paths: list, output_path: str) -> str:
    """
    검수조사서 xlsx 자동 생성
    트리거: 프로그램에서 실사진 업로드 시 자동 호출

    data = {
        "project_name":     "소재부품기술개발사업/2차년도",
        "request_no":       "2025-6537",
        "project_code":     "R24GA01_02",

        # 납품자 (업체)
        "vendor_biz_no":    "138-81-59052",
        "vendor_name":      "케이에스케이",
        "vendor_ceo":       "김인석",
        "vendor_address":   "경기 안양시 동안구 호계동 555-9",

        # 발주자 (당사)
        "buyer_biz_no":     "124-87-31775",
        "buyer_name":       "이노로보틱스 주식회사",
        "buyer_ceo":        "이현철",
        "buyer_address":    "경기도 화성시 정남면 정남산단로 80",

        # 금액
        "supply_amount":    8400,
        "vat_amount":       840,

        # 날짜 (YYYY-MM-DD)
        "order_date":       "2025-03-31",
        "delivery_date":    "2025-06-10",
        "inspection_date":  "",   # 비우면 오늘 날짜 자동

        # 검사 장소
        "inspection_place": "이노로보틱스㈜ 1층 자재보관실(전실)",

        # 검사 결과 (True=예, False=아니오, None=해당없음)
        "check_quantity":   True,
        "check_spec":       True,
        "check_parts":      True,
        "check_condition":  True,

        # 검사 결과 종합
        "inspection_result": "합격",

        # 물품 내역 (발주서 품목과 동일)
        "items": [
            {"item_name": "MAGENT",  "spec": "BY3-56-3",  "order_qty": 2, "recv_qty": 2},
            {"item_name": "HANDLE",  "spec": "BYC-175",   "order_qty": 1, "recv_qty": 1},
        ],

        # 검사자
        "inspector_dept":   "개발1그룹",
        "inspector_rank":   "전임연구원",
        "inspector_name":   "",   # 비우면 로그인 사용자 이름

        # 물품 용도
        "item_purpose":     "데모기 반발력 저감 구조물 관련 구매 건",
    }

    photo_paths: 실사진 파일 경로 리스트 (최대 2장 권장)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검수조사서"

    # ── 열 너비 설정 ──────────────────────────────────
    col_widths = [2, 6, 6, 8, 14, 2, 8, 8, 8, 2, 8, 8, 8, 2, 8, 8, 8, 2, 8, 6]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 행 높이 설정 (기본) ───────────────────────────
    for r in range(1, 80):
        ws.row_dimensions[r].height = 18

    r = 1  # 현재 행 포인터

    # ── 제목 ─────────────────────────────────────────
    ws.merge_cells(f"A{r}:T{r}")
    ws.row_dimensions[r].height = 26
    c = ws.cell(row=r, column=1, value="물품 (용역) 검사(수) 조서")
    c.font = _font(bold=True, size=14)
    c.alignment = _center()
    r += 1

    ws.merge_cells(f"A{r}:T{r}")
    c = ws.cell(row=r, column=1,
                value="검사 : 치수, 외관 등의 일치여부 확인 / 검수 : 수량적인 부분 확인")
    c.font = _font(size=9, color="595959")
    c.alignment = _center()
    r += 2

    # ── 기본 정보 ─────────────────────────────────────
    inspection_date = (data.get("inspection_date") or
                       datetime.now().strftime("%Y-%m-%d"))

    info_rows = [
        ("계약건명 (과제명/연차)", data.get("project_name", ""),
         "구매의뢰서 문서번호", data.get("request_no", "")),
    ]

    def info_row(ws, r, lbl1, val1, lbl2, val2):
        ws.merge_cells(f"B{r}:E{r}")
        ws.merge_cells(f"F{r}:J{r}")
        ws.merge_cells(f"K{r}:M{r}")
        ws.merge_cells(f"N{r}:T{r}")
        _apply(ws, r, 2,  lbl1, bold=True, fill="D9E2F3", size=9)
        _apply(ws, r, 6,  val1, align="left", size=9)
        _apply(ws, r, 11, lbl2, bold=True, fill="D9E2F3", size=9)
        _apply(ws, r, 14, val2, align="left", size=9)

    info_row(ws, r, "계약건명 (과제명/연차)", data.get("project_name",""),
             "구매의뢰서 문서번호", data.get("request_no",""))
    r += 1

    # ── 납품자 / 발주자 ───────────────────────────────
    def party_rows(ws, r, v_biz, v_name, v_ceo, v_addr, b_biz, b_name, b_ceo, b_addr):
        labels = [("납품자", "발주자"),
                  ("사업자등록번호", v_biz,  "사업자등록번호", b_biz),
                  ("상호",          v_name, "상호",          b_name),
                  ("대표자 성명",   v_ceo,  "대표자 성명",   b_ceo),
                  ("사업장 소재지", v_addr, "사업장 소재지", b_addr)]
        for i, row_data in enumerate(labels):
            if i == 0:
                ws.merge_cells(f"B{r}:J{r}")
                ws.merge_cells(f"K{r}:T{r}")
                _apply(ws, r, 2,  "납품자", bold=True, fill="D9E2F3", size=10)
                _apply(ws, r, 11, "발주자", bold=True, fill="D9E2F3", size=10)
            else:
                lbl1, v1, lbl2, v2 = row_data
                ws.merge_cells(f"B{r+i-1}:D{r+i-1}")
                ws.merge_cells(f"E{r+i-1}:J{r+i-1}")
                ws.merge_cells(f"K{r+i-1}:M{r+i-1}")
                ws.merge_cells(f"N{r+i-1}:T{r+i-1}")
                _apply(ws, r+i-1, 2,  lbl1, bold=True, fill="D9E2F3", size=9)
                _apply(ws, r+i-1, 5,  v1,   align="left", size=9)
                _apply(ws, r+i-1, 11, lbl2, bold=True, fill="D9E2F3", size=9)
                _apply(ws, r+i-1, 14, v2,   align="left", size=9)

    party_rows(ws, r,
               data.get("vendor_biz_no",""), data.get("vendor_name",""),
               data.get("vendor_ceo",""),    data.get("vendor_address",""),
               data.get("buyer_biz_no",""),  data.get("buyer_name","이노로보틱스 주식회사"),
               data.get("buyer_ceo","이현철"), data.get("buyer_address","경기도 화성시 정남면 정남산단로 80"))
    r += 5

    # ── 금액 / 날짜 ───────────────────────────────────
    supply = data.get("supply_amount", 0)
    vat    = data.get("vat_amount", int(supply * 0.1))
    total  = supply + vat

    ws.merge_cells(f"B{r}:D{r}")
    ws.merge_cells(f"E{r}:G{r}")
    ws.merge_cells(f"H{r}:I{r}")
    ws.merge_cells(f"J{r}:K{r}")
    ws.merge_cells(f"L{r}:M{r}")
    ws.merge_cells(f"N{r}:O{r}")
    ws.merge_cells(f"P{r}:T{r}")
    _apply(ws, r, 2,  "계약금액 (발주금액)", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 5,  f"{total:,}", size=9)
    _apply(ws, r, 8,  "공급가액", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 10, f"{supply:,}", size=9)
    _apply(ws, r, 12, "부가세액", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 14, f"{vat:,}", size=9)
    r += 1

    ws.merge_cells(f"B{r}:D{r}")
    ws.merge_cells(f"E{r}:H{r}")
    ws.merge_cells(f"I{r}:K{r}")
    ws.merge_cells(f"L{r}:T{r}")
    _apply(ws, r, 2,  "계약체결일(발주일)", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 5,  data.get("order_date",""), size=9)
    _apply(ws, r, 9,  "실제 납품년월일", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 12, data.get("delivery_date",""), align="left", size=9)
    r += 1

    ws.merge_cells(f"B{r}:D{r}")
    ws.merge_cells(f"E{r}:H{r}")
    ws.merge_cells(f"I{r}:K{r}")
    ws.merge_cells(f"L{r}:T{r}")
    _apply(ws, r, 2,  "검사(수) 년월일", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 5,  inspection_date, size=9)
    _apply(ws, r, 9,  "검사(수) 장소", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 12, data.get("inspection_place","이노로보틱스㈜ 1층 자재보관실(전실)"), align="left", size=9)
    r += 1

    # ── 검사 체크리스트 ───────────────────────────────
    checks = [
        ("수량은 일치 하는가?",        data.get("check_quantity",  True)),
        ("계약규격과 외형은 일치 하는가?", data.get("check_spec",   True)),
        ("구성품은 제대로 부착되어 있는가?", data.get("check_parts", True)),
        ("계약 시 요청한 조건에 만족하는가?", data.get("check_condition", True)),
    ]

    def check_mark(val):
        if val is True:   return "예  O", "아니오"
        if val is False:  return "예",    "아니오  O"
        return "예",      "해당없음  O"

    ws.merge_cells(f"B{r}:D{r+len(checks)-1}")
    _apply(ws, r, 2, "검사(수)자 의견", bold=True, fill="D9E2F3", size=9)

    for i, (q, v) in enumerate(checks):
        ws.merge_cells(f"E{r+i}:M{r+i}")
        ws.merge_cells(f"N{r+i}:P{r+i}")
        ws.merge_cells(f"Q{r+i}:T{r+i}")
        yes_txt, no_txt = check_mark(v)
        _apply(ws, r+i, 5,  f"- {q}", align="left", size=9)
        _apply(ws, r+i, 14, yes_txt, size=9, bold=(yes_txt.endswith("O")))
        _apply(ws, r+i, 17, no_txt,  size=9, bold=(no_txt.endswith("O")))
    r += len(checks)

    # ── 검사 결과 ────────────────────────────────────
    ws.merge_cells(f"B{r}:F{r}")
    ws.merge_cells(f"G{r}:T{r}")
    _apply(ws, r, 2, "검사(수) 결과 (불합격 조치사항)", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 7, data.get("inspection_result","합격"), align="left", size=9)
    r += 2

    # ── 물품검사 내역서 ───────────────────────────────
    ws.merge_cells(f"B{r}:T{r}")
    _apply(ws, r, 2, "물품검사(수) 내역서", bold=True, fill="BDD7EE", size=10)
    r += 1

    # 헤더
    hdrs = ["NO.", "품명", "규격", "발주수량", "납품수량", "잔품수량", "비고"]
    merges = [(2,3), (4,7), (8,13), (14,15), (16,17), (18,19), (20,20)]
    for (s, e), h in zip(merges, hdrs):
        if s == e:
            _apply(ws, r, s, h, bold=True, fill="D9E2F3", size=9)
        else:
            ws.merge_cells(f"{get_column_letter(s)}{r}:{get_column_letter(e)}{r}")
            _apply(ws, r, s, h, bold=True, fill="D9E2F3", size=9)
    r += 1

    items = data.get("items", [])
    for i in range(10):
        if i < len(items):
            item = items[i]
            o_qty = item.get("order_qty", 0)
            r_qty = item.get("recv_qty", 0)
            row_data = [str(i+1), item.get("item_name",""), item.get("spec",""),
                        str(o_qty), str(r_qty), str(max(0, o_qty - r_qty)), ""]
        else:
            row_data = [str(i+1), "", "", "", "", "0", ""]

        for (s, e), val in zip(merges, row_data):
            if s == e:
                _apply(ws, r, s, val, size=9, align="center")
            else:
                ws.merge_cells(f"{get_column_letter(s)}{r}:{get_column_letter(e)}{r}")
                _apply(ws, r, s, val, size=9,
                       align="left" if s >= 4 and s <= 13 else "center")
        r += 1

    # ── 확인 문구 ────────────────────────────────────
    ws.merge_cells(f"B{r}:T{r}")
    _apply(ws, r, 2, "위와 같이 검사(수) 하였음.", bold=True, size=10)
    r += 2

    # ── 검사자 ───────────────────────────────────────
    ws.merge_cells(f"B{r}:C{r}")
    ws.merge_cells(f"D{r}:F{r}")
    ws.merge_cells(f"G{r}:I{r}")
    ws.merge_cells(f"J{r}:L{r}")
    ws.merge_cells(f"M{r}:O{r}")
    ws.merge_cells(f"P{r}:R{r}")
    ws.merge_cells(f"S{r}:T{r}")
    _apply(ws, r, 2,  "검사(수)자", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 4,  "소속", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 7,  data.get("inspector_dept","개발1그룹"), size=9)
    _apply(ws, r, 10, "직급", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 13, data.get("inspector_rank",""), size=9)
    _apply(ws, r, 16, "성명", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 19, data.get("inspector_name",""), size=9)
    r += 2

    # ── 물품 사진 섹션 ────────────────────────────────
    ws.merge_cells(f"B{r}:T{r}")
    ws.row_dimensions[r].height = 22
    c = ws.cell(row=r, column=2, value="※ 물품 사진 (입고 시 촬영분) / 수량이 사진상 확인되어야 함.")
    c.font = _font(bold=True, size=9, color="C00000")
    c.alignment = _left()
    r += 1

    ws.merge_cells(f"B{r}:C{r}")
    ws.merge_cells(f"D{r}:K{r}")
    _apply(ws, r, 2, "물품용도", bold=True, fill="D9E2F3", size=9)
    _apply(ws, r, 4, data.get("item_purpose",""), align="left", size=9)
    r += 1

    # 사진 라벨
    ws.merge_cells(f"B{r}:K{r}")
    ws.merge_cells(f"L{r}:T{r}")
    _apply(ws, r, 2,  "사진 01", bold=True, fill="EBF3FB", size=10)
    _apply(ws, r, 12, "사진 02", bold=True, fill="EBF3FB", size=10)
    r += 1

    # 사진 행 (높이 크게)
    photo_start = r
    for pr in range(r, r + 15):
        ws.row_dimensions[pr].height = 14
    ws.merge_cells(f"B{r}:K{r+14}")
    ws.merge_cells(f"L{r}:T{r+14}")
    _apply(ws, r, 2, "", border=True)
    _apply(ws, r, 12, "", border=True)

    # 실사진 삽입
    for i, photo_path in enumerate(photo_paths[:2]):
        if os.path.exists(photo_path):
            try:
                # 이미지 리사이즈 (A4 절반 크기에 맞게)
                img = PILImage.open(photo_path)
                img.thumbnail((400, 300))
                thumb_path = photo_path + "_thumb.jpg"
                img.save(thumb_path, "JPEG")
                xl_img = XLImage(thumb_path)
                xl_img.width  = 300
                xl_img.height = 200
                col_letter = "B" if i == 0 else "L"
                ws.add_image(xl_img, f"{col_letter}{photo_start}")
            except Exception as e:
                print(f"사진 삽입 오류: {e}")
    r += 16

    # ── 첨부 서류 ────────────────────────────────────
    ws.merge_cells(f"B{r}:T{r}")
    _apply(ws, r, 2, "첨부 서류", bold=True, fill="BDD7EE", size=10)
    r += 1

    attachments = [
        "관련 구매의뢰서",
        "관련 거래명세표",
        "관련 발주서 또는 계약서",
        "관련 도면 및 사양서",
    ]
    for i, att in enumerate(attachments, 1):
        ws.merge_cells(f"B{r}:C{r}")
        ws.merge_cells(f"D{r}:T{r}")
        _apply(ws, r, 2, str(i), size=9)
        _apply(ws, r, 4, att, align="left", size=9)
        r += 1

    wb.save(output_path)
    return output_path


def _money(v):
    try:
        return f"{int(float(v or 0)):,}"
    except Exception:
        return "0"

def generate_inspection_report_pdf(data: dict, photo_paths: list, output_path: str) -> str:
    """검수조사서 PDF 직접 생성(A4 1페이지 맞춤). Excel/LibreOffice 없이 동작."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    font = _ensure_pdf_font()
    c = canvas.Canvas(output_path, pagesize=A4)
    page_w, page_h = A4
    margin_x = 11 * mm
    top_y = page_h - 15 * mm
    table_w = page_w - (margin_x * 2)
    blue = colors.HexColor('#D9E2F3')
    blue2 = colors.HexColor('#BDD7EE')

    def _s(v):
        return '' if v is None else str(v)

    def _shorten(text, max_chars):
        text = _s(text)
        return text if len(text) <= max_chars else text[:max_chars-1] + '…'

    def _tw(t, size):
        try:
            return pdfmetrics.stringWidth(_s(t), font, size)
        except Exception:
            return len(_s(t)) * size * 0.5

    def _fit(t, max_w, size, min_size=5.4):
        while size > min_size and _tw(t, size) > max_w:
            size -= 0.35
        return size

    def txt(x, y, text, size=8, bold=False, align='left', max_w=None):
        text = _s(text)
        if max_w:
            size = _fit(text, max_w, size)
        c.setFont(font, size)
        if align == 'center':
            c.drawCentredString(x, y, text)
        elif align == 'right':
            c.drawRightString(x, y, text)
        else:
            c.drawString(x, y, text)

    def rect(x, y, w, h, fill=None, lw=0.75):
        c.setLineWidth(lw)
        if fill:
            c.setFillColor(fill)
            c.rect(x, y, w, h, stroke=1, fill=1)
            c.setFillColor(colors.black)
        else:
            c.rect(x, y, w, h, stroke=1, fill=0)

    def cell(x, y_top, w, h, text='', fill=None, align='left', size=8, bold=False, lw=0.75):
        rect(x, y_top-h, w, h, fill=fill, lw=lw)
        px = x + 2*mm
        if align == 'center':
            px = x + w/2
        elif align == 'right':
            px = x + w - 2*mm
        txt(px, y_top - h/2 - size*0.32, text, size=size, bold=bold, align=align, max_w=w-4*mm)

    def row4(y_top, l1, v1, l2, v2, h=8.0*mm):
        widths = [42*mm, 56*mm, 40*mm, table_w - (42+56+40)*mm]
        xs = [margin_x]
        for ww in widths[:-1]:
            xs.append(xs[-1] + ww)
        vals = [l1, v1, l2, v2]
        for i, (xx, ww, val) in enumerate(zip(xs, widths, vals)):
            cell(xx, y_top, ww, h, val, fill=blue if i in (0, 2) else None,
                 align='center' if i in (0, 2) else 'left', size=7.4, bold=i in (0, 2), lw=0.9)
        return y_top - h

    y = top_y
    txt(page_w/2, y, '물품 (용역) 검사(수) 조서', 15, True, 'center'); y -= 6.2*mm
    txt(page_w/2, y, '검사 : 치수, 외관 등의 일치여부 확인 / 검수 : 수량적인 부분 확인', 6.8, False, 'center'); y -= 6.8*mm

    y = row4(y, '계약건명 (과제명/연차)', data.get('project_name',''), '구매의뢰서 문서번호', data.get('request_no',''))
    y -= 1.0*mm

    half_w = table_w / 2
    cell(margin_x, y, half_w, 8.0*mm, '납품자', fill=blue, align='center', size=8.5, bold=True, lw=1.0)
    cell(margin_x+half_w, y, half_w, 8.0*mm, '발주자', fill=blue, align='center', size=8.5, bold=True, lw=1.0)
    y -= 8.0*mm
    for lab, val, lab2, val2 in [
        ('사업자등록번호', data.get('vendor_biz_no',''), '사업자등록번호', data.get('buyer_biz_no','124-87-31775')),
        ('상호', data.get('vendor_name',''), '상호', data.get('buyer_name','이노로보틱스 주식회사')),
        ('대표자 성명', data.get('vendor_ceo',''), '대표자 성명', data.get('buyer_ceo','이현철')),
        ('사업장 소재지', data.get('vendor_address',''), '사업장 소재지', data.get('buyer_address','경기도 화성시 정남면 정남산단로 80')),
    ]:
        y = row4(y, lab, val, lab2, val2)

    y -= 1.2*mm
    supply = int(float(data.get('supply_amount') or 0))
    vat = int(float(data.get('vat_amount') or 0))
    total = supply + vat
    y = row4(y, '계약금액(발주금액)', _money(total), '공급가액 / 부가세액', f'{_money(supply)} / {_money(vat)}')
    y = row4(y, '계약체결일(발주일)', data.get('order_date',''), '실제 납품년월일', data.get('delivery_date',''))
    y = row4(y, '검사(수) 년월일', data.get('inspection_date',''), '검사(수) 장소', data.get('inspection_place','이노로보틱스㈜ 1층 자재보관실(전실)'))

    y -= 5.0*mm
    cell(margin_x, y, table_w, 8.2*mm, '물품검사(수) 내역서', fill=blue2, align='center', size=9, bold=True, lw=1.0)
    y -= 8.2*mm

    col_mm = [10, 55, 45, 20, 20, 20, 20]
    scale = table_w / (sum(col_mm) * mm)
    cols = [v * mm * scale for v in col_mm]
    headers = ['NO.', '품명', '규격', '발주수량', '납품수량', '잔품수량', '비고']
    x = margin_x
    for cw, hd in zip(cols, headers):
        cell(x, y, cw, 8.0*mm, hd, fill=blue, align='center', size=7.5, bold=True, lw=0.9)
        x += cw
    y -= 8.0*mm

    items = data.get('items') or []
    row_h = 6.8 * mm
    for idx in range(9):
        item = items[idx] if idx < len(items) else {}
        oq = int(float(item.get('order_qty') or 0)) if item else 0
        rq = int(float(item.get('recv_qty') or oq)) if item else 0
        remain = int(float(item.get('remain_qty', max(0, oq-rq)) or 0)) if item else ''
        vals = [idx+1 if idx < len(items) else '', _shorten(item.get('item_name',''), 42) if item else '', _shorten(item.get('spec',''), 32) if item else '', oq if idx < len(items) else '', rq if idx < len(items) else '', remain if idx < len(items) else '', item.get('note','') if item else '']
        x = margin_x
        for ci, (cw, val) in enumerate(zip(cols, vals)):
            cell(x, y, cw, row_h, val, fill=None, align='left' if ci in (1,2,6) else 'center', size=6.7, lw=0.65)
            x += cw
        y -= row_h

    y -= 4.0*mm
    cell(margin_x, y, table_w, 7.4*mm, '검사(수) 결과 (불합격 조치사항) : ' + _s(data.get('inspection_result','합격')), fill=blue, align='left', size=7.5, bold=True, lw=0.9)
    y -= 9.2*mm
    txt(margin_x, y, '위와 같이 검사(수) 하였음.', 8.5, True); y -= 7.0*mm
    txt(margin_x, y, f"검사(수)자  소속: {data.get('inspector_dept','')}    직급: {data.get('inspector_rank','')}    성명: {data.get('inspector_name','')}", 7.2); y -= 8.0*mm
    txt(margin_x, y, '※ 물품 사진 (입고 시 촬영분) / 수량이 사진상 확인되어야 함.', 7.2); y -= 5.0*mm

    box_w = 78 * mm
    box_h = min(45*mm, max(30*mm, y - 18*mm))
    box_x = margin_x
    box_y = y - box_h
    rect(box_x, box_y, box_w, box_h, lw=0.8)
    txt(box_x + box_w/2, y - 4.0*mm, '사진 01', 7.2, True, 'center')
    if photo_paths:
        ph = photo_paths[0]
        if ph and os.path.exists(ph):
            try:
                c.drawImage(ph, box_x + 3*mm, box_y + 3*mm, width=box_w-6*mm, height=box_h-10*mm, preserveAspectRatio=True, anchor='c')
            except Exception:
                pass

    c.showPage()
    c.save()
    return output_path

def open_folder_on_server(path: str):
    try:
        folder = path if os.path.isdir(path) else os.path.dirname(path)
        if os.name == 'nt':
            os.startfile(folder)
        elif sys.platform == 'darwin':
            import subprocess; subprocess.Popen(['open', folder])
        else:
            import subprocess; subprocess.Popen(['xdg-open', folder])
    except Exception as e:
        print(f'폴더 열기 실패: {e}')


def auto_generate_on_photo_upload(request_id: int, photo_paths: list, db) -> dict:
    """
    프로그램에서 실사진 업로드 시 자동 호출되는 진입점

    1. DB에서 구매의뢰/발주/입고 데이터 조회
    2. 검수조사서 xlsx 생성
    3. Bizbox에 자동 첨부 (선택)
    """
    from server.database import PurchaseRequest, PurchaseOrder, GoodsReceipt, InventoryItem

    pr = db.query(PurchaseRequest).filter(PurchaseRequest.id == request_id).first()
    if not pr:
        return {"success": False, "message": "구매의뢰서를 찾을 수 없습니다"}

    po = pr.purchase_order
    receipt = db.query(GoodsReceipt).filter(GoodsReceipt.order_id == po.id).first() if po else None
    item = db.query(InventoryItem).filter(InventoryItem.order_id == po.id).first() if po else None

    today = datetime.now().strftime("%Y-%m-%d")

    data = {
        "project_name":     pr.project_name or "",
        "request_no":       pr.request_no or "",
        "project_code":     pr.project_name.split("]")[0].replace("[","") if pr.project_name else "",

        "vendor_biz_no":    "",  # 업체 DB에서 추가 가능
        "vendor_name":      po.vendor_name if po else "",
        "vendor_ceo":       "",
        "vendor_address":   "",

        "buyer_biz_no":     "124-87-31775",
        "buyer_name":       "이노로보틱스 주식회사",
        "buyer_ceo":        "이현철",
        "buyer_address":    "경기도 화성시 정남면 정남산단로 80",

        "supply_amount":    int(pr.unit_price * pr.quantity) if pr.unit_price else 0,
        "vat_amount":       int(pr.unit_price * pr.quantity * 0.1) if pr.unit_price else 0,

        "order_date":       po.order_date.strftime("%Y-%m-%d") if po and po.order_date else today,
        "delivery_date":    receipt.received_date.strftime("%Y-%m-%d") if receipt else today,
        "inspection_date":  today,

        "inspection_place": "이노로보틱스㈜ 1층 자재보관실(전실)",

        "check_quantity":   True,
        "check_spec":       True,
        "check_parts":      True,
        "check_condition":  True,
        "inspection_result": "합격",

        "items": [{
            "item_name": pr.item_name or "",
            "spec":      pr.spec or "",
            "order_qty": pr.quantity or 0,
            "recv_qty":  receipt.received_qty if receipt else 0,
        }],

        "inspector_dept":   "개발1그룹",
        "inspector_rank":   "",
        "inspector_name":   "",
        "item_purpose":     pr.reason or "",
    }

    # 출력 경로
    out_dir = "./inspection_reports"
    os.makedirs(out_dir, exist_ok=True)
    filename = f"검수조사서_{pr.request_no}_{pr.item_name}_{today}.xlsx"
    output_path = os.path.join(out_dir, filename)

    try:
        generate_inspection_report(data, photo_paths, output_path)
        return {
            "success":     True,
            "output_path": output_path,
            "filename":    filename,
            "message":     f"검수조사서 생성 완료: {filename}"
        }
    except Exception as e:
        return {"success": False, "message": str(e)}


# ─────────────────────────────────────────────────────────────
# 검수조사서: 사용자 제공 Excel 양식 기반 PDF 생성
# - 템플릿의 색상/테두리/병합/이미지/인쇄영역을 그대로 사용
# - 내부적으로 임시 xlsx를 만들고 PDF 변환 후 xlsx는 삭제
# ─────────────────────────────────────────────────────────────
import shutil
import subprocess
import tempfile

INSPECTION_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    'templates',
    'inspection_template.xlsx'
)

def _set_template_value(ws, cell_ref: str, value):
    """병합셀이어도 안전하게 좌상단 셀에 값 입력."""
    cell = ws[cell_ref]
    row, col = cell.row, cell.column
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            row, col = merged.min_row, merged.min_col
            break
    ws.cell(row=row, column=col).value = value if value is not None else ''

def _fmt_date(value):
    if not value:
        return ''
    try:
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        s = str(value)
        return s[:10] if len(s) >= 10 else s
    except Exception:
        return str(value)

def _convert_xlsx_to_pdf(xlsx_path: str, pdf_path: str) -> str:
    """Windows Excel COM 우선, LibreOffice 보조."""
    xlsx_path = os.path.abspath(xlsx_path)
    pdf_path = os.path.abspath(pdf_path)
    out_dir = os.path.dirname(pdf_path)
    os.makedirs(out_dir, exist_ok=True)

    # 1) Windows + Excel 설치 환경
    try:
        import win32com.client  # type: ignore
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(xlsx_path)
        try:
            # 0 = PDF
            wb.ExportAsFixedFormat(0, pdf_path)
        finally:
            wb.Close(False)
            excel.Quit()
        if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
            return pdf_path
    except Exception:
        pass

    # 2) LibreOffice / soffice 설치 환경
    for exe in ('soffice', 'libreoffice'):
        try:
            subprocess.run(
                [exe, '--headless', '--convert-to', 'pdf', '--outdir', out_dir, xlsx_path],
                check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=120
            )
            converted = os.path.join(out_dir, os.path.splitext(os.path.basename(xlsx_path))[0] + '.pdf')
            if os.path.exists(converted):
                if os.path.abspath(converted) != os.path.abspath(pdf_path):
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
                    os.replace(converted, pdf_path)
                if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                    return pdf_path
        except Exception:
            continue

    raise RuntimeError('PDF 변환 실패: Excel 또는 LibreOffice가 설치되어 있는지 확인하세요.')

def generate_inspection_report_from_template_pdf(data: dict, photo_paths: list, pdf_path: str) -> str:
    """
    검수조사서 최종 산출물을 PDF로 직접 생성한다.
    사용자가 제공한 Excel은 양식 참고용이며, 실행 PC에 Excel/LibreOffice가 없어도 동작하도록
    xlsx 생성/변환 과정을 사용하지 않는다.
    """
    return generate_inspection_report_pdf(data, photo_paths, pdf_path)

# ─────────────────────────────────────────────────────────────
# 최신 패치: 검수조사서 PDF 직접 생성 개선
# - 글씨 겹침 방지: 자동 줄바꿈/행 높이 자동 보정
# - 납품자/발주자 표 정렬 보정
# - 검사자 성명 옆 샘플 도장 표시
# - 품목/사진 수가 많으면 자동으로 다음 페이지 생성
# ─────────────────────────────────────────────────────────────
def generate_inspection_report_pdf(data: dict, photo_paths: list, output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    font = _ensure_pdf_font()
    c = canvas.Canvas(output_path, pagesize=A4)
    page_w, page_h = A4
    margin = 12 * mm
    table_w = page_w - margin * 2
    blue = colors.HexColor('#D9E2F3')
    blue2 = colors.HexColor('#BDD7EE')
    red = colors.HexColor('#C00000')

    def _s(v):
        return '' if v is None else str(v)

    def _num(v, default=0):
        try:
            return int(float(v or default))
        except Exception:
            return default

    def _text_width(text, size):
        try:
            return pdfmetrics.stringWidth(_s(text), font, size)
        except Exception:
            return len(_s(text)) * size * 0.5

    def _wrap_text(text, max_w, size, max_lines=None):
        """ReportLab canvas용 간단 줄바꿈. 한글/영문 혼합을 글자 단위로 안전 처리."""
        text = _s(text).replace('\r', ' ').replace('\n', ' ')
        if not text:
            return ['']
        lines, cur = [], ''
        for ch in text:
            test = cur + ch
            if cur and _text_width(test, size) > max_w:
                lines.append(cur)
                cur = ch
            else:
                cur = test
        if cur:
            lines.append(cur)
        if max_lines and len(lines) > max_lines:
            lines = lines[:max_lines]
            if lines:
                lines[-1] = lines[-1][:-1] + '…' if len(lines[-1]) > 1 else '…'
        return lines or ['']

    def _line_height(size):
        return size * 1.35

    def txt(x, y, text, size=8, align='left'):
        c.setFont(font, size)
        s = _s(text)
        if align == 'center':
            c.drawCentredString(x, y, s)
        elif align == 'right':
            c.drawRightString(x, y, s)
        else:
            c.drawString(x, y, s)

    def cell(x, y_top, w, h, text='', fill=None, size=7.2, align='left', bold=False, max_lines=None, lw=0.75):
        c.setLineWidth(lw)
        if fill:
            c.setFillColor(fill)
            c.rect(x, y_top-h, w, h, stroke=1, fill=1)
            c.setFillColor(colors.black)
        else:
            c.rect(x, y_top-h, w, h, stroke=1, fill=0)

        inner_w = max(1, w - 3.0*mm)
        lines = _wrap_text(text, inner_w, size, max_lines=max_lines)
        lh = _line_height(size)
        block_h = lh * len(lines)
        start_y = y_top - (h - block_h) / 2 - size
        for i, line in enumerate(lines):
            yy = start_y - i * lh
            if align == 'center':
                txt(x + w/2, yy, line, size, 'center')
            elif align == 'right':
                txt(x + w - 1.5*mm, yy, line, size, 'right')
            else:
                txt(x + 1.5*mm, yy, line, size, 'left')

    def required_h(*texts, widths, size=7.0, min_h=7.8*mm, max_lines=3):
        max_count = 1
        for t, w in zip(texts, widths):
            max_count = max(max_count, len(_wrap_text(t, max(1, w-3*mm), size, max_lines=max_lines)))
        return max(min_h, (max_count * _line_height(size) + 3.2*mm))

    def header(title='물품 (용역) 검사(수) 조서'):
        y = page_h - 15*mm
        txt(page_w/2, y, title, 15, 'center')
        y -= 7*mm
        txt(page_w/2, y, '검사 : 치수, 외관 등의 일치여부 확인 / 검수 : 수량적인 부분 확인', 6.8, 'center')
        y -= 7*mm
        return y

    def ensure_space(y, need_h, title='물품 (용역) 검사(수) 조서'):
        if y - need_h < margin:
            c.showPage()
            return header(title)
        return y

    # 좌/우 2단 표: 납품자/발주자 제목과 아래 행의 세로줄이 정확히 맞도록 half 기준 사용
    def row4(y, l1, v1, l2, v2, size=6.8, min_h=7.8*mm):
        half = table_w / 2
        label_w = 38 * mm
        value_w = half - label_w
        widths = [label_w, value_w, label_w, value_w]
        h = required_h(l1, v1, l2, v2, widths=widths, size=size, min_h=min_h, max_lines=3)
        y = ensure_space(y, h)
        xs = [margin, margin + label_w, margin + half, margin + half + label_w]
        vals = [l1, v1, l2, v2]
        for i, (x, w, v) in enumerate(zip(xs, widths, vals)):
            cell(x, y, w, h, v, blue if i in (0, 2) else None,
                 size=size, align='center' if i in (0, 2) else 'left', max_lines=3, lw=0.9)
        return y - h

    def draw_stamp(cx, cy, name='관리자'):
        """성명 옆에 들어가는 샘플 도장. 외부 이미지 없이 PDF에 직접 그림."""
        label = _s(name).strip() or '관리자'
        # 너무 긴 이름은 도장 안쪽에 들어가도록 앞 3글자만 사용
        label = label[:3]
        c.saveState()
        c.setStrokeColor(red)
        c.setFillColor(red)
        c.setLineWidth(1.4)
        c.ellipse(cx - 8*mm, cy - 10*mm, cx + 8*mm, cy + 10*mm, stroke=1, fill=0)
        c.setFont(font, 7.2)
        if len(label) >= 3:
            stamp_text = '\n'.join(label[:3])
            for j, ch in enumerate(label[:3]):
                c.drawCentredString(cx, cy + (5-j*5)*mm, ch)
        else:
            c.drawCentredString(cx, cy - 1.5*mm, label)
        c.restoreState()

    # ── 상단 정보 ─────────────────────────────────────
    y = header()
    y = row4(y, '계약건명 (과제명/연차)', data.get('project_name',''), '구매의뢰서 문서번호', data.get('request_no',''))
    y -= 1*mm

    half = table_w / 2
    cell(margin, y, half, 7.5*mm, '납품자', blue, 8, 'center', lw=1.0)
    cell(margin+half, y, half, 7.5*mm, '발주자', blue, 8, 'center', lw=1.0)
    y -= 7.5*mm

    party_rows = [
        ('사업자등록번호', data.get('vendor_biz_no',''), '사업자등록번호', data.get('buyer_biz_no','124-87-31775')),
        ('상호', data.get('vendor_name',''), '상호', data.get('buyer_name','이노로보틱스 주식회사')),
        ('대표자 성명', data.get('vendor_ceo',''), '대표자 성명', data.get('buyer_ceo','이현철')),
        ('사업장 소재지', data.get('vendor_address',''), '사업장 소재지', data.get('buyer_address','경기도 화성시 정남면 정남산단로 80')),
    ]
    for lab, val, lab2, val2 in party_rows:
        y = row4(y, lab, val, lab2, val2, size=6.7)

    y -= 1*mm
    supply = _num(data.get('supply_amount'))
    vat = _num(data.get('vat_amount'))
    total = supply + vat
    y = row4(y, '계약금액(발주금액)', f'{total:,}', '공급가액 / 부가세액', f'{supply:,} / {vat:,}')
    y = row4(y, '계약체결일(발주일)', data.get('order_date',''), '실제 납품년월일', data.get('delivery_date',''))
    y = row4(y, '검사(수) 년월일', data.get('inspection_date',''), '검사(수) 장소', data.get('inspection_place','이노로보틱스㈜ 1층 자재보관실(전실)'))

    # ── 물품검사 내역서 ───────────────────────────────
    y -= 4*mm
    y = ensure_space(y, 16*mm, '물품검사(수) 내역서')
    cell(margin, y, table_w, 7.5*mm, '물품검사(수) 내역서', blue2, 8.5, 'center', lw=1.0)
    y -= 7.5*mm

    col_mm = [10, 54, 44, 20, 20, 20, 21]
    scale = table_w / (sum(col_mm) * mm)
    cols = [v * mm * scale for v in col_mm]
    headers = ['NO.', '품명', '규격', '발주수량', '납품수량', '잔품수량', '비고']

    def draw_item_header(y):
        x = margin
        for cw, hd in zip(cols, headers):
            cell(x, y, cw, 7*mm, hd, blue, 7, 'center', lw=0.9)
            x += cw
        return y - 7*mm

    y = draw_item_header(y)
    row_min_h = 6.8 * mm
    for idx, item in enumerate(data.get('items') or [], 1):
        oq = _num(item.get('order_qty'))
        rq = _num(item.get('recv_qty'), oq)
        remain = item.get('remain_qty', max(0, oq-rq))
        vals = [idx, item.get('item_name',''), item.get('spec',''), oq, rq, remain, item.get('note','')]
        # 품명/규격/비고 길이에 따라 행 높이 증가 → 아래 글씨 겹침 방지
        h = required_h(vals[1], vals[2], vals[6], widths=[cols[1], cols[2], cols[6]], size=6.25, min_h=row_min_h, max_lines=2)
        if y - h < margin + 38*mm:
            c.showPage()
            y = header('물품검사(수) 내역서 - 계속')
            y = draw_item_header(y)
        x = margin
        for ci, (cw, val) in enumerate(zip(cols, vals)):
            cell(x, y, cw, h, val, None, 6.25, 'left' if ci in (1, 2, 6) else 'center', max_lines=2, lw=0.65)
            x += cw
        y -= h

    # ── 검사 결과 / 검사자 ────────────────────────────
    y -= 3*mm
    if y < 40*mm:
        c.showPage()
        y = header()
    result_h = required_h('검사(수) 결과 (불합격 조치사항) : ' + _s(data.get('inspection_result','합격')),
                          widths=[table_w], size=7.2, min_h=8.2*mm, max_lines=2)
    cell(margin, y, table_w, result_h, '검사(수) 결과 (불합격 조치사항) : ' + _s(data.get('inspection_result','합격')),
         blue, 7.2, 'left', max_lines=2, lw=0.9)
    y -= result_h + 3*mm

    txt(margin, y, '위와 같이 검사(수) 하였음.', 8.5)
    y -= 8*mm

    dept = data.get('inspector_dept','')
    rank = data.get('inspector_rank','')
    name = data.get('inspector_name','') or '관리자'
    sig_text = f"검사(수)자  소속: {dept}    직급: {rank}    성명: {name}"
    txt(margin, y, sig_text, 7.2)
    # 성명 오른쪽 여백에 도장 배치. 글씨와 겹치지 않도록 충분히 우측으로 이동.
    stamp_x = min(page_w - margin - 12*mm, margin + max(105*mm, _text_width(sig_text, 7.2) + 12*mm))
    draw_stamp(stamp_x, y + 1.0*mm, name)
    y -= 16*mm

    # ── 사진 ─────────────────────────────────────────
    photos = [p for p in (photo_paths or []) if p and os.path.exists(p)]
    if photos:
        if y < 75*mm:
            c.showPage()
            y = header('물품 사진')
        txt(margin, y, '※ 물품 사진 (입고/검수 시 촬영분)', 7.5)
        y -= 5*mm
        box_w = (table_w - 5*mm) / 2
        box_h = 58*mm
        for i, ph in enumerate(photos, 1):
            if i % 2 == 1 and y - box_h < margin:
                c.showPage()
                y = header('물품 사진 - 계속')
            x = margin if (i % 2 == 1) else margin + box_w + 5*mm
            top = y
            c.rect(x, top-box_h, box_w, box_h, stroke=1, fill=0)
            txt(x + box_w/2, top - 4*mm, f'사진 {i:02d}', 7.2, 'center')
            try:
                c.drawImage(ph, x + 3*mm, top-box_h + 3*mm,
                            width=box_w-6*mm, height=box_h-10*mm,
                            preserveAspectRatio=True, anchor='c')
            except Exception:
                pass
            if i % 2 == 0:
                y -= box_h + 6*mm
        if len(photos) % 2 == 1:
            y -= box_h + 6*mm

    c.showPage()
    c.save()
    return output_path


def generate_inspection_report_from_template_pdf(data: dict, photo_paths: list, pdf_path: str) -> str:
    return generate_inspection_report_pdf(data, photo_paths, pdf_path)


# ─────────────────────────────────────────────────────────────
# V1.0.1 PATCH: 사용자 제공 Excel 양식 기반 XLSX 생성
# - 구매의뢰서-업체별 검수조사서 작성
# - 사진이 많으면 시트(탭) 자동 추가
# - 사진 외 나머지 공통 내용은 동일 입력
# - 업체정보/사업자정보 반영
# - 검수자 성명 기반 도장 이미지 자동 삽입
# ─────────────────────────────────────────────────────────────
def _to_int_safe(v, default=0):
    try:
        return int(float(v or default))
    except Exception:
        return default


def _clear_sheet_images(ws):
    try:
        ws._images = []
    except Exception:
        pass


def _make_name_stamp_image(name: str, output_dir: str) -> str:
    """검수자 성명으로 도장 이미지를 생성한다."""
    os.makedirs(output_dir, exist_ok=True)
    label = str(name or '').strip() or '검수자'
    # 2~3글자 위주 표시. 너무 길면 끝 3글자 사용
    stamp_label = label[-3:] if len(label) > 3 else label
    img = PILImage.new('RGBA', (220, 220), (255, 255, 255, 0))
    from PIL import ImageDraw, ImageFont
    draw = ImageDraw.Draw(img)
    red = (200, 0, 0, 255)
    draw.ellipse((10, 10, 210, 210), outline=red, width=6)

    font = None
    for fp in [r'C:\Windows\Fonts\malgun.ttf', r'C:\Windows\Fonts\malgunbd.ttf', '/usr/share/fonts/truetype/nanum/NanumGothic.ttf', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf']:
        if os.path.exists(fp):
            try:
                font = ImageFont.truetype(fp, 54)
                break
            except Exception:
                pass
    if font is None:
        font = ImageFont.load_default()

    if len(stamp_label) >= 2:
        chars = list(stamp_label)
        y_positions = [44, 92, 140] if len(chars) == 3 else [66, 118]
        for ch, y in zip(chars, y_positions):
            bbox = draw.textbbox((0, 0), ch, font=font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            draw.text(((220 - tw) / 2, y - th / 2), ch, font=font, fill=red)
    else:
        bbox = draw.textbbox((0, 0), stamp_label, font=font)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        draw.text(((220 - tw) / 2, (220 - th) / 2), stamp_label, font=font, fill=red)

    path = os.path.join(output_dir, f"stamp_{label}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.png")
    img.save(path)
    return path


def generate_inspection_report_template_xlsx(data: dict, photo_paths: list, output_path: str) -> str:
    """사용자 제공 inspection_template.xlsx 양식으로 검수조사서를 생성한다."""
    template_path = INSPECTION_TEMPLATE_PATH
    if not os.path.exists(template_path):
        return generate_inspection_report(data, photo_paths, output_path)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.load_workbook(template_path)
    base_ws = wb[wb.sheetnames[0]]

    # 양식 내 추가 시트 제거
    for ws in list(wb.worksheets[1:]):
        wb.remove(ws)
    _clear_sheet_images(base_ws)

    items = list(data.get('items') or [])
    photos = [p for p in (photo_paths or []) if p and os.path.exists(p)]
    items_per_sheet = 6
    photos_per_sheet = 2

    from math import ceil
    item_page_count = max(1, ceil(len(items) / items_per_sheet)) if items else 1
    photo_page_count = max(1, ceil(len(photos) / photos_per_sheet)) if photos else 1
    page_count = max(item_page_count, photo_page_count)

    # 사진 때문에 늘어나는 경우 items는 모든 시트에 동일하게 반복
    repeat_same_items = len(items) <= items_per_sheet

    def safe_set(ws, cell_ref, value):
        try:
            _set_template_value(ws, cell_ref, value)
        except Exception:
            try:
                ws[cell_ref] = value if value is not None else ''
            except Exception:
                pass

    def safe_get_value(item, key, default=''):
        return item.get(key, default) if isinstance(item, dict) else default

    def clear_item_rows(ws):
        for row in range(19, 25):
            for ref in [f'T{row}', f'U{row}', f'X{row}', f'AB{row}', f'AD{row}', f'AF{row}', f'AH{row}']:
                safe_set(ws, ref, '')

    def fill_common(ws):
        vendor_name = data.get('vendor_name', '')
        supply = _to_int_safe(data.get('supply_amount'))
        vat = _to_int_safe(data.get('vat_amount'), int(supply * 0.1))
        total = supply + vat

        safe_set(ws, 'X4', data.get('project_name', ''))
        safe_set(ws, 'AI4', data.get('request_no', ''))

        safe_set(ws, 'X5', data.get('vendor_biz_no') or data.get('biz_no') or data.get('business_no') or data.get('registration_no') or data.get('사업자등록번호') or data.get('사업자번호') or '')
        safe_set(ws, 'X6', vendor_name)
        safe_set(ws, 'X7', data.get('vendor_ceo', ''))
        safe_set(ws, 'X8', data.get('vendor_address', ''))

        safe_set(ws, 'AG5', data.get('buyer_biz_no', '124-87-31775'))
        safe_set(ws, 'AG6', data.get('buyer_name', '이노로보틱스 주식회사'))
        safe_set(ws, 'AG7', data.get('buyer_ceo', '이현철'))
        safe_set(ws, 'AG8', data.get('buyer_address', '경기도 화성시 정남면 정남산단로 80'))

        safe_set(ws, 'X9', total)
        safe_set(ws, 'AD9', supply)
        safe_set(ws, 'AI9', vat)

        safe_set(ws, 'X10', _fmt_date(data.get('order_date', '')))
        safe_set(ws, 'AI10', _fmt_date(data.get('delivery_date', '')))
        safe_set(ws, 'X11', _fmt_date(data.get('inspection_date', '')) or datetime.now().strftime('%Y-%m-%d'))
        safe_set(ws, 'AG11', data.get('inspection_place', '이노로보틱스㈜ 1층 자재보관실(전실)'))

        safe_set(ws, 'X16', data.get('inspection_result', '합격'))
        safe_set(ws, 'X27', data.get('inspector_dept', ''))
        safe_set(ws, 'AC27', data.get('inspector_rank', ''))
        safe_set(ws, 'AG27', data.get('inspector_name', ''))
        safe_set(ws, 'V30', data.get('item_purpose', ''))

    def fill_item_rows(ws, page_index):
        clear_item_rows(ws)
        chunk = items if repeat_same_items else items[page_index * items_per_sheet:(page_index + 1) * items_per_sheet]
        chunk = list(chunk[:items_per_sheet])
        base_no = 0 if repeat_same_items else page_index * items_per_sheet
        for i in range(items_per_sheet):
            row = 19 + i
            if i >= len(chunk):
                safe_set(ws, f'T{row}', i + 1 if repeat_same_items else base_no + i + 1)
                safe_set(ws, f'AF{row}', '')
                continue
            item = chunk[i]
            oq = _to_int_safe(safe_get_value(item, 'order_qty'))
            rq = _to_int_safe(safe_get_value(item, 'recv_qty'), oq)
            remain = safe_get_value(item, 'remain_qty', max(0, oq - rq))
            safe_set(ws, f'T{row}', i + 1 if repeat_same_items else base_no + i + 1)
            safe_set(ws, f'U{row}', safe_get_value(item, 'item_name'))
            safe_set(ws, f'X{row}', safe_get_value(item, 'spec'))
            safe_set(ws, f'AB{row}', oq)
            safe_set(ws, f'AD{row}', rq)
            safe_set(ws, f'AF{row}', remain)
            safe_set(ws, f'AH{row}', safe_get_value(item, 'note'))

    def clear_photo_area(ws):
        _clear_sheet_images(ws)
        for ref in ['T31', 'AC31']:
            safe_set(ws, ref, '')

    def add_photo(ws, img_path, anchor, width_limit=355, height_limit=235):
        if not img_path or not os.path.exists(img_path):
            return
        try:
            img = PILImage.open(img_path)
            if img.mode not in ('RGB', 'L'):
                img = img.convert('RGB')
            img.thumbnail((width_limit, height_limit))
            tmp_dir = os.path.join(os.path.dirname(output_path), '_inspection_tmp')
            os.makedirs(tmp_dir, exist_ok=True)
            tmp_path = os.path.join(tmp_dir, f"photo_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.jpg")
            img.save(tmp_path, 'JPEG', quality=90)
            xl_img = XLImage(tmp_path)
            xl_img.width = img.width
            xl_img.height = img.height
            ws.add_image(xl_img, anchor)
        except Exception as e:
            print(f'검수조사서 사진 삽입 오류: {img_path} / {e}')

    def add_signature_stamp(ws, inspector_name):
        try:
            tmp_dir = os.path.join(os.path.dirname(output_path), '_inspection_tmp')
            stamp_path = _make_name_stamp_image(inspector_name, tmp_dir)
            stamp = XLImage(stamp_path)
            stamp.width = 62
            stamp.height = 62
            ws.add_image(stamp, 'AJ27')
        except Exception as e:
            print(f'검수자 도장 이미지 생성 오류: {e}')

    for page_index in range(page_count):
        if page_index == 0:
            ws = base_ws
            ws.title = '검수조사서_1' if page_count > 1 else '검수조사서'
        else:
            ws = wb.copy_worksheet(base_ws)
            ws.title = f'검수조사서_{page_index + 1}'
            _clear_sheet_images(ws)

        fill_common(ws)
        fill_item_rows(ws, page_index)
        clear_photo_area(ws)

        photo_chunk = photos[page_index * photos_per_sheet:(page_index + 1) * photos_per_sheet]
        if len(photo_chunk) >= 1:
            add_photo(ws, photo_chunk[0], 'T31')
        if len(photo_chunk) >= 2:
            add_photo(ws, photo_chunk[1], 'AC31')

        add_signature_stamp(ws, data.get('inspector_name', ''))
        try:
            ws.sheet_view.showGridLines = False
        except Exception:
            pass

    wb.save(output_path)
    return output_path
